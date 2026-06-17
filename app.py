from __future__ import annotations

import csv
import html
import io
import json
import re
import urllib.request
import uuid
import unicodedata
import xml.etree.ElementTree as ET
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, urlparse
from xml.dom import minidom

import cgi
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


HOST = "127.0.0.1"
PORT = 8765
SETTINGS_FILE = Path(__file__).with_name("settings.json")
ACCOUNTS_FILE = Path(__file__).with_name("own_accounts.json")
COMPANIES_FILE = Path(__file__).with_name("companies.json")
PARTNERS_FILE = Path(__file__).with_name("partners.json")
BANK_REGISTRY_FILE = Path(__file__).with_name("bank_registry.json")

FIELDS = {
    "identifier": {
        "label": "Azonosító",
        "required": False,
        "help": "Ha üres, automatikusan: dátum + sorszám.",
    },
    "sender_account": {"label": "Feladó számlaszám", "required": True},
    "sender_name": {"label": "Feladó neve", "required": True},
    "beneficiary_account": {"label": "Címzett számlaszám", "required": True},
    "beneficiary_name": {"label": "Címzett neve", "required": True},
    "beneficiary_country": {"label": "Címzett országa", "required": False},
    "note": {"label": "Közlemény", "required": False},
    "document_no": {"label": "Bizonylatszám", "required": False},
    "legal_code": {"label": "Jogcímkód", "required": False},
    "currency": {"label": "Deviza", "required": True},
    "decimals": {"label": "Tizedesjegyek száma", "required": True},
    "amount": {"label": "Átutalandó összeg", "required": True},
    "value_date": {"label": "Valutanap", "required": True},
    "status": {"label": "Státusz", "required": False},
}

FX_FIELDS = {
    "identifier": {
        "label": "Azonosító",
        "required": False,
        "help": "Ha üres, automatikusan: dátum + sorszám.",
    },
    "sender_account": {"label": "Feladó számlaszám", "required": True},
    "sender_account_type": {"label": "Feladó számlaszám típusa", "required": True},
    "sender_name": {"label": "Feladó neve", "required": True},
    "beneficiary_account": {"label": "Címzett számlaszám", "required": True},
    "beneficiary_account_type": {"label": "Címzett számlaszám típusa", "required": True},
    "beneficiary_bank_name": {"label": "Címzett bank neve", "required": False},
    "beneficiary_name": {"label": "Címzett neve", "required": True},
    "beneficiary_country": {"label": "Bejegyzés országa", "required": True},
    "beneficiary_bank_swift": {"label": "Partner bank SWIFT kódja", "required": True},
    "correspondent_bank_1": {"label": "Levelező bank", "required": False},
    "swift_copy": {"label": "SWIFT másolat", "required": False},
    "fax_number": {"label": "Fax szám", "required": False},
    "custom_rate_use": {"label": "Egyedi árfolyam használata", "required": False},
    "custom_rate": {"label": "Egyedi árfolyam", "required": False},
    "bank_message": {"label": "Közlemény a bank számára", "required": False},
    "note": {"label": "Közlemény", "required": False},
    "legal_code": {"label": "Jogcímkód", "required": True},
    "permit_no": {"label": "Deviza engedély száma", "required": False},
    "permit_date": {"label": "Deviza engedély dátuma", "required": False},
    "urgent_use": {"label": "Sürgős teljesítés használata", "required": False},
    "urgent_execution": {"label": "Sürgős teljesítés", "required": False},
    "fax_copy_execution": {"label": "Fax másolat teljesítés", "required": False},
    "payout_currency": {"label": "Kifizetendő deviza", "required": True},
    "currency": {"label": "Átutalandó deviza", "required": True},
    "decimals": {"label": "Tizedesjegyek száma", "required": True},
    "amount": {"label": "Átutalandó összeg", "required": True},
    "value_date": {"label": "Feldolgozás kezdő dátuma", "required": False},
    "cost_bearer": {"label": "Költségviselés", "required": False},
    "partner_bank_country": {"label": "Partner bank országkódja", "required": True},
    "correspondent_bank_2": {"label": "Levelező bank 2. sor", "required": False},
    "correspondent_bank_3": {"label": "Levelező bank 3. sor", "required": False},
    "correspondent_bank_4": {"label": "Levelező bank 4. sor", "required": False},
}

TEMPLATE_EXAMPLE = {
    "identifier": "",
    "sender_account": "11111111-22222222-33333333",
    "sender_name": "Minta Feladó Kft",
    "beneficiary_account": "44444444-55555555-66666666",
    "beneficiary_name": "Minta Partner Kft",
    "beneficiary_country": "HU",
    "note": "Számla kiegyenlítés",
    "document_no": "BIZ001",
    "legal_code": "",
    "currency": "HUF",
    "decimals": "0",
    "amount": "12345",
    "value_date": "2026-06-17",
    "status": "",
}

FX_TEMPLATE_EXAMPLE = {
    "identifier": "",
    "sender_account": "HU00111111112222222233333333",
    "sender_account_type": "0",
    "sender_name": "Minta Feladó Kft",
    "beneficiary_account": "DE00123456781234567890",
    "beneficiary_account_type": "0",
    "beneficiary_bank_name": "Minta Bank",
    "beneficiary_name": "Minta Partner GmbH",
    "beneficiary_country": "DE",
    "beneficiary_bank_swift": "DEUTDEFF",
    "correspondent_bank_1": "",
    "swift_copy": "N",
    "fax_number": "",
    "custom_rate_use": "N",
    "custom_rate": "",
    "bank_message": "",
    "note": "Invoice payment",
    "legal_code": "1111",
    "permit_no": "",
    "permit_date": "",
    "urgent_use": "N",
    "urgent_execution": "",
    "fax_copy_execution": "",
    "payout_currency": "EUR",
    "currency": "EUR",
    "decimals": "2",
    "amount": "123.45",
    "value_date": "2026-06-17",
    "cost_bearer": "1",
    "partner_bank_country": "DE",
    "correspondent_bank_2": "",
    "correspondent_bank_3": "",
    "correspondent_bank_4": "",
}

SEPA_FIELDS = {
    "identifier": {
        "label": "Azonosító",
        "required": False,
        "help": "Ha üres, automatikusan: dátum + sorszám.",
    },
    "sender_account": {"label": "Feladó számlaszám", "required": True},
    "sender_account_type": {"label": "Feladó számlaszám típusa", "required": False},
    "sender_name": {"label": "Feladó neve", "required": True},
    "beneficiary_account": {"label": "Címzett IBAN számlaszám", "required": True},
    "beneficiary_account_type": {"label": "Címzett számlaszám típusa", "required": False},
    "beneficiary_name": {"label": "Címzett neve", "required": True},
    "beneficiary_bank_swift": {"label": "Partner bank SWIFT kódja", "required": True},
    "note": {"label": "Közlemény", "required": False},
    "urgent_execution": {"label": "Sürgős teljesítés", "required": False},
    "payout_currency": {"label": "Kifizetendő deviza", "required": True},
    "currency": {"label": "Átutalandó deviza", "required": True},
    "decimals": {"label": "Tizedesjegyek száma", "required": True},
    "amount": {"label": "Átutalandó összeg", "required": True},
}

SEPA_TEMPLATE_EXAMPLE = {
    "identifier": "",
    "sender_account": "HU00111111112222222233333333",
    "sender_account_type": "0",
    "sender_name": "Minta Feladó Kft",
    "beneficiary_account": "AT451100000421840000",
    "beneficiary_account_type": "0",
    "beneficiary_name": "Minta Partner GmbH",
    "beneficiary_bank_swift": "BKAUATWW",
    "note": "Invoice payment",
    "urgent_execution": "",
    "payout_currency": "EUR",
    "currency": "EUR",
    "decimals": "2",
    "amount": "123.45",
}

KH_FX_FIELDS = {
    "identifier": {
        "label": "Azonosító",
        "required": False,
        "help": "Ha üres, automatikusan: dátum + sorszám.",
    },
    "sender_account": {"label": "Feladó számlaszám", "required": True},
    "sender_name": {"label": "Feladó neve", "required": True},
    "beneficiary_account": {"label": "Címzett számlaszám", "required": True},
    "beneficiary_bank_name": {"label": "Címzett bank neve", "required": True},
    "beneficiary_bank_address": {"label": "Címzett bank címe", "required": False},
    "beneficiary_name": {"label": "Címzett neve", "required": True},
    "beneficiary_bank_swift": {"label": "Partner bank SWIFT kódja", "required": True},
    "bank_message": {"label": "Közlemény a bank számára", "required": False},
    "note": {"label": "Közlemény", "required": False},
    "payout_currency": {"label": "Kifizetendő deviza", "required": True},
    "currency": {"label": "Átutalandó deviza", "required": True},
    "decimals": {"label": "Tizedesjegyek száma", "required": True},
    "amount": {"label": "Átutalandó összeg", "required": True},
    "value_date": {"label": "Valutanap", "required": False},
    "cost_bearer": {"label": "Költségviselés", "required": False},
    "urgent_execution": {"label": "Sürgős teljesítés", "required": False},
    "group_transfer": {"label": "Csoportos átutalás", "required": False},
}

KH_FX_TEMPLATE_EXAMPLE = {
    "identifier": "",
    "sender_account": "11111111-22222222-33333333",
    "sender_name": "Minta Feladó Kft",
    "beneficiary_account": "DE00123456781234567890",
    "beneficiary_bank_name": "Minta Bank",
    "beneficiary_bank_address": "Minta bankcím",
    "beneficiary_name": "Minta Partner GmbH",
    "beneficiary_bank_swift": "DEUTDEFF",
    "bank_message": "",
    "note": "Invoice payment",
    "payout_currency": "EUR",
    "currency": "EUR",
    "decimals": "2",
    "amount": "123.45",
    "value_date": "2026-06-17",
    "cost_bearer": "1",
    "urgent_execution": "N",
    "group_transfer": "N",
}

UNICREDIT_PAY_FIELDS = {
    "sender_account": {"label": "Terhelendő számlaszám", "required": True},
    "sender_currency": {"label": "Terhelendő számla devizája", "required": True},
    "beneficiary_account": {"label": "Kedvezményezett számlaszám", "required": True},
    "beneficiary_name": {"label": "Kedvezményezett neve", "required": True},
    "document_no": {"label": "Bizonylatszám", "required": False},
    "note": {"label": "Közlemény", "required": False},
    "amount": {"label": "Összeg", "required": True},
    "currency": {"label": "Összeg devizaneme", "required": True},
    "value_date": {"label": "Értéknap", "required": False},
    "legal_code": {"label": "Jogcímkód", "required": False},
    "beneficiary_country": {"label": "Jogosult országkódja", "required": True},
}

UNICREDIT_PAY_TEMPLATE_EXAMPLE = {
    "sender_account": "11111111-22222222-33333333",
    "sender_currency": "HUF",
    "beneficiary_account": "44444444-55555555-66666666",
    "beneficiary_name": "Minta Partner Kft",
    "document_no": "BIZ001",
    "note": "Számla kiegyenlítés",
    "amount": "12345",
    "currency": "HUF",
    "value_date": "2026-06-17",
    "legal_code": "",
    "beneficiary_country": "HU",
}

UNICREDIT_CCY_FIELDS = {
    "sender_account": {"label": "Terhelendő számlaszám", "required": True},
    "sender_currency": {"label": "Terhelendő számla devizája", "required": True},
    "beneficiary_bank_swift": {"label": "Kedvezményezett bank BIC kódja", "required": True},
    "beneficiary_bank_id": {"label": "Kedvezményezett bank azonosítója", "required": False},
    "beneficiary_bank_name": {"label": "Kedvezményezett bank neve", "required": True},
    "beneficiary_bank_address_1": {"label": "Kedvezményezett bank címe 1", "required": False},
    "beneficiary_bank_address_2": {"label": "Kedvezményezett bank címe 2", "required": False},
    "beneficiary_bank_address_3": {"label": "Kedvezményezett bank címe 3", "required": False},
    "correspondent_bank_swift": {"label": "Levelező bank BIC kódja", "required": False},
    "correspondent_bank_id": {"label": "Levelező bank azonosítója", "required": False},
    "correspondent_bank_name": {"label": "Levelező bank neve", "required": False},
    "correspondent_bank_address_1": {"label": "Levelező bank címe 1", "required": False},
    "correspondent_bank_address_2": {"label": "Levelező bank címe 2", "required": False},
    "correspondent_bank_address_3": {"label": "Levelező bank címe 3", "required": False},
    "beneficiary_account": {"label": "Kedvezményezett számlaszáma", "required": True},
    "beneficiary_name": {"label": "Kedvezményezett neve", "required": True},
    "beneficiary_address_1": {"label": "Kedvezményezett címe 1", "required": False},
    "beneficiary_address_2": {"label": "Kedvezményezett címe 2", "required": False},
    "beneficiary_address_3": {"label": "Kedvezményezett címe 3", "required": False},
    "note": {"label": "Közlemény", "required": False},
    "cost_bearer": {"label": "Banki költségek megosztása", "required": True},
    "payout_currency": {"label": "Teljesítés devizaneme", "required": True},
    "amount": {"label": "Összeg", "required": True},
    "currency": {"label": "Összeg devizaneme", "required": True},
    "value_date": {"label": "Értéknap", "required": False},
    "urgent_execution": {"label": "Sürgősség átutalás flag", "required": False},
    "hold_flag": {"label": "HOLD flag", "required": False},
    "chqb_flag": {"label": "CHQB flag", "required": False},
    "deal_ticket_flag": {"label": "Deal Ticket flag", "required": False},
    "deal_ticket_date": {"label": "Deal Ticket dátuma", "required": False},
    "deal_ticket_sequence": {"label": "Deal Ticket sorszáma", "required": False},
    "beneficiary_country": {"label": "Kedvezményezett országkódja", "required": False},
    "legal_code": {"label": "Statisztikai jogcímkód", "required": True},
}

UNICREDIT_CCY_TEMPLATE_EXAMPLE = {
    "sender_account": "11111111-22222222-33333333",
    "sender_currency": "HUF",
    "beneficiary_bank_swift": "DEUTDEFF",
    "beneficiary_bank_id": "",
    "beneficiary_bank_name": "Minta Bank",
    "beneficiary_bank_address_1": "Minta cím 1",
    "beneficiary_bank_address_2": "",
    "beneficiary_bank_address_3": "",
    "correspondent_bank_swift": "",
    "correspondent_bank_id": "",
    "correspondent_bank_name": "",
    "correspondent_bank_address_1": "",
    "correspondent_bank_address_2": "",
    "correspondent_bank_address_3": "",
    "beneficiary_account": "DE00123456781234567890",
    "beneficiary_name": "Minta Partner GmbH",
    "beneficiary_address_1": "Minta cím 1",
    "beneficiary_address_2": "",
    "beneficiary_address_3": "",
    "note": "Invoice payment",
    "cost_bearer": "SHA",
    "payout_currency": "EUR",
    "amount": "123.45",
    "currency": "EUR",
    "value_date": "2026-06-17",
    "urgent_execution": "N",
    "hold_flag": "N",
    "chqb_flag": "N",
    "deal_ticket_flag": "N",
    "deal_ticket_date": "",
    "deal_ticket_sequence": "",
    "beneficiary_country": "DE",
    "legal_code": "000",
}

OTP_HUF_FIELDS = {
    "identifier": {"label": "Azonosító", "required": False, "help": "Ha üres, automatikusan: dátum + sorszám."},
    "sender_account": {"label": "Feladó számlaszám", "required": True},
    "sender_name": {"label": "Feladó neve", "required": True},
    "beneficiary_account": {"label": "Címzett számlaszám", "required": True},
    "beneficiary_name": {"label": "Címzett neve", "required": True},
    "note": {"label": "Közlemény", "required": False},
    "document_no": {"label": "Bizonylatszám", "required": False},
    "currency": {"label": "Deviza", "required": True},
    "decimals": {"label": "Tizedesjegyek száma", "required": True},
    "amount": {"label": "Átutalandó összeg", "required": True},
    "value_date": {"label": "Értéknap", "required": False},
    "process_mode": {"label": "Feldolgozási mód", "required": False},
}

OTP_HUF_TEMPLATE_EXAMPLE = {
    "identifier": "",
    "sender_account": "11111111-22222222-33333333",
    "sender_name": "Minta Feladó Kft",
    "beneficiary_account": "44444444-55555555-66666666",
    "beneficiary_name": "Minta Partner Kft",
    "note": "Számla kiegyenlítés",
    "document_no": "BIZ001",
    "currency": "HUF",
    "decimals": "0",
    "amount": "12345",
    "value_date": "",
    "process_mode": "",
}

OTP_FX_FIELDS = {
    "identifier": {"label": "Azonosító", "required": False, "help": "Ha üres, automatikusan: dátum + sorszám."},
    "sender_account": {"label": "Feladó számlaszám", "required": True},
    "sender_account_type": {"label": "Feladó számlaszám típusa", "required": True},
    "sender_name": {"label": "Feladó neve", "required": True},
    "beneficiary_account": {"label": "Címzett számlaszám", "required": True},
    "beneficiary_account_type": {"label": "Címzett számlaszám típusa", "required": True},
    "beneficiary_bank_name": {"label": "Címzett bank neve", "required": True},
    "beneficiary_bank_address": {"label": "Címzett bank címe", "required": False},
    "beneficiary_name": {"label": "Címzett neve", "required": True},
    "beneficiary_bank_swift": {"label": "Partner bank azonosítója", "required": True},
    "correspondent_bank_1": {"label": "Levelező bank", "required": False},
    "bank_message": {"label": "Megjegyzés a bank számára", "required": False},
    "note": {"label": "Közlemény", "required": False},
    "permit_no": {"label": "Deviza engedély száma", "required": False},
    "permit_date": {"label": "Deviza engedély dátuma", "required": False},
    "urgent_execution": {"label": "Sürgős teljesítés", "required": False},
    "payout_currency": {"label": "Kifizetendő deviza", "required": True},
    "currency": {"label": "Átutalandó deviza", "required": True},
    "decimals": {"label": "Tizedesjegyek száma", "required": True},
    "amount": {"label": "Átutalandó összeg", "required": True},
    "cost_bearer": {"label": "Költségviselés", "required": True},
}

OTP_FX_TEMPLATE_EXAMPLE = {
    "identifier": "",
    "sender_account": "HU00111111112222222233333333",
    "sender_account_type": "9",
    "sender_name": "Minta Feladó Kft",
    "beneficiary_account": "DE00123456781234567890",
    "beneficiary_account_type": "9",
    "beneficiary_bank_name": "Minta Bank",
    "beneficiary_bank_address": "Minta bankcím",
    "beneficiary_name": "Minta Partner GmbH",
    "beneficiary_bank_swift": "DEUTDEFF",
    "correspondent_bank_1": "",
    "bank_message": "",
    "note": "Invoice payment",
    "permit_no": "",
    "permit_date": "",
    "urgent_execution": "1",
    "payout_currency": "EUR",
    "currency": "EUR",
    "decimals": "2",
    "amount": "123.45",
    "cost_bearer": "1",
}

RAIFFEISEN_HUF_FIELDS = {
    "amount": {"label": "Összeg", "required": True},
    "value_date": {"label": "Értéknap", "required": False},
    "sender_account": {"label": "Kezdeményező pénzforgalmi jelzőszáma", "required": True},
    "beneficiary_account": {"label": "Kedvezményezett pénzforgalmi jelzőszáma", "required": True},
    "beneficiary_name": {"label": "Kedvezményezett neve", "required": True},
    "beneficiary_country": {"label": "Kedvezményezett országkódja", "required": False},
    "legal_code": {"label": "Jogcím", "required": False},
    "note": {"label": "Közlemény", "required": False},
    "external_ref": {"label": "Külső referencia", "required": False},
    "document_no": {"label": "Bizonylatszám", "required": False},
    "internal_note": {"label": "Belső megjegyzés", "required": False},
}

RAIFFEISEN_HUF_TEMPLATE_EXAMPLE = {
    "amount": "12345",
    "value_date": "2026-06-17",
    "sender_account": "11111111-22222222-33333333",
    "beneficiary_account": "44444444-55555555-66666666",
    "beneficiary_name": "Minta Partner Kft",
    "beneficiary_country": "HU",
    "legal_code": "",
    "note": "Számla kiegyenlítés",
    "external_ref": "",
    "document_no": "BIZ001",
    "internal_note": "",
}

RAIFFEISEN_FX_FIELDS = {
    "currency": {"label": "Átutalandó összeg devizaneme", "required": True},
    "sender_currency": {"label": "Megbízó számla devizaneme", "required": False},
    "amount": {"label": "Összeg", "required": True},
    "value_date": {"label": "Értéknap", "required": False},
    "sender_account": {"label": "Terhelendő számlaszám", "required": True},
    "beneficiary_name": {"label": "Kedvezményezett neve", "required": True},
    "beneficiary_bank_name": {"label": "Kedvezményezett bank neve", "required": True},
    "beneficiary_bank_country": {"label": "Kedvezményezett bank ország", "required": False},
    "beneficiary_bank_address_1": {"label": "Kedvezményezett bank címe 1", "required": False},
    "beneficiary_bank_address_2": {"label": "Kedvezményezett bank címe 2", "required": False},
    "beneficiary_account": {"label": "Kedvezményezett számlaszáma", "required": True},
    "legal_code": {"label": "Jogcímkód", "required": False},
    "note": {"label": "Közlemény", "required": False},
    "commission_bearer": {"label": "Ügyfél jutalék fizetője", "required": True},
    "other_fee_bearer": {"label": "Egyéb jutalék fizetője", "required": True},
    "permit_no": {"label": "Engedély szám", "required": False},
    "document_no": {"label": "Bizonylatszám", "required": False},
    "beneficiary_bank_swift": {"label": "Kedvezményezett swift címe", "required": False},
    "amount_currency_mode": {"label": "Összeg devizaneme jel", "required": False},
    "payment_method": {"label": "Teljesítés módja", "required": False},
    "priority": {"label": "Prioritás", "required": False},
    "item_type": {"label": "Tétel típusa", "required": False},
    "iban_flag": {"label": "IBAN jelzés", "required": False},
    "beneficiary_country": {"label": "Kedvezményezett országkódja", "required": False},
    "external_ref": {"label": "Külső referencia", "required": False},
}

RAIFFEISEN_FX_TEMPLATE_EXAMPLE = {
    "currency": "EUR",
    "sender_currency": "HUF",
    "amount": "123.45",
    "value_date": "2026-06-17",
    "sender_account": "11111111-22222222-33333333",
    "beneficiary_name": "Minta Partner GmbH",
    "beneficiary_bank_name": "Minta Bank",
    "beneficiary_bank_country": "DE",
    "beneficiary_bank_address_1": "Minta cím 1",
    "beneficiary_bank_address_2": "",
    "beneficiary_account": "DE00123456781234567890",
    "legal_code": "",
    "note": "Invoice payment",
    "commission_bearer": "0",
    "other_fee_bearer": "0",
    "permit_no": "",
    "document_no": "BIZ001",
    "beneficiary_bank_swift": "DEUTDEFF",
    "amount_currency_mode": " ",
    "payment_method": " ",
    "priority": " ",
    "item_type": " ",
    "iban_flag": "1",
    "beneficiary_country": "DE",
    "external_ref": "",
}

PAIN_FIELDS = {
    "message_id": {"label": "Üzenet azonosító", "required": False},
    "payment_id": {"label": "Fizetési blokk azonosító", "required": False},
    "debtor_name": {"label": "Terhelendő fél neve", "required": True},
    "sender_account": {"label": "Terhelendő számla", "required": True},
    "sender_currency": {"label": "Terhelendő számla devizája", "required": False},
    "beneficiary_name": {"label": "Kedvezményezett neve", "required": True},
    "beneficiary_account": {"label": "Kedvezményezett számlaszáma / IBAN", "required": True},
    "beneficiary_bank_swift": {"label": "Kedvezményezett bank BIC/SWIFT", "required": False},
    "beneficiary_country": {"label": "Kedvezményezett országkódja", "required": False},
    "amount": {"label": "Összeg", "required": True},
    "currency": {"label": "Deviza", "required": True},
    "value_date": {"label": "Értéknap", "required": True},
    "end_to_end_id": {"label": "EndToEnd azonosító", "required": False},
    "cost_bearer": {"label": "Költségviselés", "required": False},
    "note": {"label": "Közlemény", "required": False},
}

PAIN_TEMPLATE_EXAMPLE = {
    "message_id": "",
    "payment_id": "",
    "debtor_name": "Minta Feladó Kft",
    "sender_account": "HU00111111112222222233333333",
    "sender_currency": "EUR",
    "beneficiary_name": "Minta Partner GmbH",
    "beneficiary_account": "AT451100000421840000",
    "beneficiary_bank_swift": "BKAUATWW",
    "beneficiary_country": "AT",
    "amount": "123.45",
    "currency": "EUR",
    "value_date": "2026-06-17",
    "end_to_end_id": "",
    "cost_bearer": "SLEV",
    "note": "Invoice payment",
}

BANKS = {
    "erste": {"label": "Erste Bank"},
    "kh": {"label": "K&H Bank"},
    "mbh": {"label": "MBH Bank"},
    "unicredit": {"label": "UniCredit Bank"},
    "raiffeisen": {"label": "Raiffeisen Bank"},
    "otp": {"label": "OTP Bank"},
}

FORMATS = {
    "erste_huf_payord": {
        "bank": "erste",
        "label": "Forint átutalás - PAYORD (DO)",
        "short_label": "PAYORD / DO",
        "badge": "PAYORD / DO · 941 karakter soronként CR/LF-fel",
        "description": "Erste forint átutalási import, 939 adatkarakter + CR/LF, tehát 941 karakter soronként.",
        "default_encoding": "cp1250",
        "line_length": 941,
        "fields": FIELDS,
        "template_example": TEMPLATE_EXAMPLE,
    },
    "erste_fx_payord": {
        "bank": "erste",
        "label": "Deviza átutalás - PAYORD (IN)",
        "short_label": "PAYORD / IN",
        "badge": "PAYORD / IN · 1048 karakter soronként CR/LF-fel",
        "description": "Erste deviza átutalási import, 1046 adatkarakter + CR/LF, tehát 1048 karakter soronként.",
        "default_encoding": "cp1250",
        "line_length": 1048,
        "fields": FX_FIELDS,
        "template_example": FX_TEMPLATE_EXAMPLE,
    },
    "erste_sepa_payord": {
        "bank": "erste",
        "label": "SEPA átutalás - PAYORD (IN / SCE)",
        "short_label": "PAYORD / IN SEPA",
        "badge": "PAYORD / IN · SEPA .SCE · 1048 karakter soronként CR/LF-fel",
        "description": "Erste SEPA (.SCE) átutalási import: a deviza PAYORD (IN) formátum szűkített mezőkészlete, EUR SEPA utaláshoz.",
        "default_encoding": "cp1250",
        "line_length": 1048,
        "fields": SEPA_FIELDS,
        "template_example": SEPA_TEMPLATE_EXAMPLE,
    },
    "kh_huf_payord": {
        "bank": "kh",
        "label": "Forint átutalás - PAYORD (DO)",
        "short_label": "PAYORD / DO",
        "badge": "K&H PAYORD / DO · 941 karakter soronként CR/LF-fel",
        "description": "K&H Electra forint átutalási import, EDIFACT PAYORD DO rekord.",
        "default_encoding": "cp1250",
        "line_length": 941,
        "fields": FIELDS,
        "template_example": TEMPLATE_EXAMPLE,
    },
    "kh_fx_payord": {
        "bank": "kh",
        "label": "Deviza átutalás - PAYORD (IN)",
        "short_label": "PAYORD / IN",
        "badge": "K&H PAYORD / IN · 941 karakter soronként CR/LF-fel",
        "description": "K&H Electra deviza átutalási import, bank-specifikus 939 adatkarakter + CR/LF rekorddal.",
        "default_encoding": "cp1250",
        "line_length": 941,
        "fields": KH_FX_FIELDS,
        "template_example": KH_FX_TEMPLATE_EXAMPLE,
    },
    "unicredit_huf_pay": {
        "bank": "unicredit",
        "label": "Forint átutalás - PAY",
        "short_label": "PAY",
        "badge": "UniCredit PAY · 256 byte rekordok · header/data/trailer",
        "description": "UniCredit SpectraNet forintátutalási import: 256 byte-os PAY header, data és trailer rekordokkal.",
        "default_encoding": "cp1250",
        "line_length": 256,
        "fields": UNICREDIT_PAY_FIELDS,
        "template_example": UNICREDIT_PAY_TEMPLATE_EXAMPLE,
    },
    "unicredit_fx_ccy": {
        "bank": "unicredit",
        "label": "Deviza átutalás - CCY",
        "short_label": "CCY",
        "badge": "UniCredit CCY · 800 byte rekordok · header/data/trailer",
        "description": "UniCredit SpectraNet devizaátutalási import: 800 byte-os CCY header, data és trailer rekordokkal.",
        "default_encoding": "cp1250",
        "line_length": 800,
        "fields": UNICREDIT_CCY_FIELDS,
        "template_example": UNICREDIT_CCY_TEMPLATE_EXAMPLE,
    },
    "kh_sepa_pain": {
        "bank": "kh",
        "label": "SEPA átutalás - PAIN XML",
        "short_label": "PAIN XML SEPA",
        "badge": "K&H SEPA · PAIN XML",
        "description": "K&H Electra SEPA import PAIN XML formátumban.",
        "default_encoding": "utf-8",
        "line_length": None,
        "fields": PAIN_FIELDS,
        "template_example": PAIN_TEMPLATE_EXAMPLE,
        "builder": "pain_sepa",
    },
    "mbh_huf": {
        "bank": "mbh",
        "label": "HUF átutalás - PAIN XML",
        "short_label": "MBH HUF",
        "badge": "MBH HUF · ISO20022 PAIN XML",
        "description": "MBH Vállalati Netbank / Direct Bank ISO20022 pain.001 import HUF utaláshoz.",
        "default_encoding": "utf-8",
        "line_length": None,
        "fields": PAIN_FIELDS,
        "template_example": {**PAIN_TEMPLATE_EXAMPLE, "currency": "HUF", "sender_currency": "HUF", "beneficiary_account": "44444444-55555555-66666666", "beneficiary_country": "HU", "cost_bearer": "SHAR"},
        "builder": "pain_huf",
    },
    "mbh_fx": {
        "bank": "mbh",
        "label": "Deviza átutalás - PAIN XML",
        "short_label": "MBH Deviza",
        "badge": "MBH Deviza · ISO20022 PAIN XML",
        "description": "MBH Vállalati Netbank / Direct Bank ISO20022 pain.001 import deviza utaláshoz.",
        "default_encoding": "utf-8",
        "line_length": None,
        "fields": PAIN_FIELDS,
        "template_example": PAIN_TEMPLATE_EXAMPLE,
        "builder": "pain_fx",
    },
    "mbh_sepa": {
        "bank": "mbh",
        "label": "SEPA átutalás - PAIN XML",
        "short_label": "MBH SEPA",
        "badge": "MBH SEPA · ISO20022 PAIN XML",
        "description": "MBH Vállalati Netbank / Direct Bank ISO20022 pain.001 import SEPA utaláshoz.",
        "default_encoding": "utf-8",
        "line_length": None,
        "fields": PAIN_FIELDS,
        "template_example": PAIN_TEMPLATE_EXAMPLE,
        "builder": "pain_sepa",
    },
    "otp_huf": {
        "bank": "otp",
        "label": "HUF átutalás - PAYORD (DO)",
        "short_label": "PAYORD / DO",
        "badge": "OTP PAYORD / DO · 941 karakter soronként CR/LF-fel",
        "description": "OTPdirekt Electra forint átutalási EDIFACT PAYORD import.",
        "default_encoding": "cp1250",
        "line_length": 941,
        "fields": OTP_HUF_FIELDS,
        "template_example": OTP_HUF_TEMPLATE_EXAMPLE,
    },
    "otp_fx": {
        "bank": "otp",
        "label": "Deviza átutalás - PAYORD (IN)",
        "short_label": "PAYORD / IN",
        "badge": "OTP PAYORD / IN · 941 karakter soronként CR/LF-fel",
        "description": "OTPdirekt Electra bankon kívüli deviza átutalási EDIFACT PAYORD import.",
        "default_encoding": "cp1250",
        "line_length": 941,
        "fields": OTP_FX_FIELDS,
        "template_example": OTP_FX_TEMPLATE_EXAMPLE,
    },
    "raiffeisen_huf": {
        "bank": "raiffeisen",
        "label": "HUF átutalás - Raiffeisen DBF",
        "short_label": "Raiffeisen DBF HUF",
        "badge": "Raiffeisen DBF HUF · 251 karakter soronként CR/LF-fel",
        "description": "Raiffeisen Electra forint átutalási DBF import.",
        "default_encoding": "cp852",
        "line_length": 251,
        "fields": RAIFFEISEN_HUF_FIELDS,
        "template_example": RAIFFEISEN_HUF_TEMPLATE_EXAMPLE,
    },
    "raiffeisen_fx": {
        "bank": "raiffeisen",
        "label": "Deviza átutalás - Raiffeisen DBF",
        "short_label": "Raiffeisen DBF Deviza",
        "badge": "Raiffeisen DBF Deviza · 495 karakter soronként CR/LF-fel",
        "description": "Raiffeisen Electra deviza átutalási DBF import.",
        "default_encoding": "cp852",
        "line_length": 495,
        "fields": RAIFFEISEN_FX_FIELDS,
        "template_example": RAIFFEISEN_FX_TEMPLATE_EXAMPLE,
    },
    "raiffeisen_sepa": {
        "bank": "raiffeisen",
        "label": "SEPA átutalás - Raiffeisen DBF",
        "short_label": "Raiffeisen DBF SEPA",
        "badge": "Raiffeisen DBF SEPA CT · 495 karakter soronként CR/LF-fel",
        "description": "Raiffeisen Electra SEPA CT import a deviza DBF formátum FTMOD=2 jelölésével.",
        "default_encoding": "cp852",
        "line_length": 495,
        "fields": RAIFFEISEN_FX_FIELDS,
        "template_example": {**RAIFFEISEN_FX_TEMPLATE_EXAMPLE, "currency": "EUR", "payment_method": "2", "iban_flag": "1"},
    },
}

CURRENCIES = [
    "HUF", "EUR", "AUD", "BGN", "BRL", "CAD", "CHF", "CNY", "CZK", "DKK", "GBP", "HKD",
    "HRK", "IDR", "ILS", "INR", "ISK", "JPY", "KRW", "MXN", "MYR", "NOK", "NZD", "PHP",
    "PLN", "RON", "RSD", "RUB", "SEK", "SGD", "THB", "TRY", "UAH", "USD", "ZAR", "ATS",
    "AUP", "BEF", "BGL", "CSD", "CSK", "DDM", "DEM", "EEK", "EGP", "ESP", "FIM", "FRF",
    "GHP", "GRD", "IEP", "ITL", "KPW", "KWD", "LBP", "LTL", "LUF", "LVL", "MNT", "NLG",
    "OAL", "OBL", "OFR", "ORB", "PKR", "PTE", "ROL", "SDP", "SIT", "SKK", "SUR", "VND",
    "XEU", "XTR", "YUD",
]

ALIASES = {
    "identifier": ["azonosító", "azonosito", "id", "megbízás azonosító", "megbizas azonosito"],
    "sender_account": ["feladó számlaszám", "felado szamlaszam", "forrás számla", "forras szamla"],
    "sender_currency": ["terhelendő számla devizája", "terhelendo szamla devizaja", "feladó deviza", "felado deviza"],
    "sender_account_type": ["feladó számlaszám típusa", "felado szamlaszam tipusa", "feladó számlatípus"],
    "sender_name": ["feladó neve", "felado neve", "megbízó", "megbizo"],
    "beneficiary_account": [
        "címzett számlaszám",
        "cimzett szamlaszam",
        "kedvezményezett számlaszám",
        "kedvezmenyezett szamlaszam",
        "számlaszám",
        "szamlaszam",
    ],
    "beneficiary_account_type": ["címzett számlaszám típusa", "cimzett szamlaszam tipusa"],
    "beneficiary_bank_name": ["címzett bank neve", "cimzett bank neve", "partner bank neve", "bank neve"],
    "beneficiary_bank_id": ["kedvezményezett bank azonosítója", "kedvezmenyezett bank azonositoja", "bank azonosító"],
    "beneficiary_bank_address": ["címzett bank címe", "cimzett bank cime", "kedvezményezett bank címe"],
    "beneficiary_bank_address_1": ["kedvezményezett bank címe 1", "kedvezmenyezett bank cime 1", "címzett bank címe 1"],
    "beneficiary_bank_address_2": ["kedvezményezett bank címe 2", "kedvezmenyezett bank cime 2", "címzett bank címe 2"],
    "beneficiary_bank_address_3": ["kedvezményezett bank címe 3", "kedvezmenyezett bank cime 3", "címzett bank címe 3"],
    "beneficiary_name": [
        "címzett neve",
        "cimzett neve",
        "kedvezményezett neve",
        "kedvezmenyezett neve",
        "partner neve",
    ],
    "beneficiary_country": ["ország", "orszag", "címzett országa", "cimzett orszaga"],
    "beneficiary_bank_swift": ["swift", "swift kód", "swift kod", "partner bank swift", "partner bank swift kódja"],
    "correspondent_bank_swift": ["levelező bank bic", "levelezo bank bic", "levelező bank swift"],
    "correspondent_bank_id": ["levelező bank azonosítója", "levelezo bank azonositoja"],
    "correspondent_bank_name": ["levelező bank neve", "levelezo bank neve"],
    "beneficiary_address_1": ["kedvezményezett címe 1", "kedvezmenyezett cime 1", "címzett címe 1"],
    "beneficiary_address_2": ["kedvezményezett címe 2", "kedvezmenyezett cime 2", "címzett címe 2"],
    "beneficiary_address_3": ["kedvezményezett címe 3", "kedvezmenyezett cime 3", "címzett címe 3"],
    "correspondent_bank_1": ["levelező bank", "levelezo bank"],
    "swift_copy": ["swift másolat", "swift masolat"],
    "fax_number": ["fax szám", "fax szam"],
    "custom_rate_use": ["egyedi árfolyam használata", "egyedi arfolyam hasznalata"],
    "custom_rate": ["egyedi árfolyam", "egyedi arfolyam"],
    "bank_message": ["közlemény a bank számára", "kozlemeny a bank szamara"],
    "note": ["közlemény", "kozlemeny", "megjegyzés", "megjegyzes"],
    "document_no": ["bizonylatszám", "bizonylatszam", "bizonylat", "sorszám", "sorszam"],
    "legal_code": ["jogcímkód", "jogcimkod", "jogcím", "jogcim"],
    "permit_no": ["deviza engedély száma", "deviza engedely szama"],
    "permit_date": ["deviza engedély dátuma", "deviza engedely datuma"],
    "urgent_use": ["sürgős teljesítés használata", "surgos teljesites hasznalata"],
    "urgent_execution": ["sürgős teljesítés", "surgos teljesites"],
    "fax_copy_execution": ["fax másolat teljesítés", "fax masolat teljesites"],
    "payout_currency": ["kifizetendő deviza", "kifizetendo deviza"],
    "currency": ["deviza", "currency", "pénznem", "penznem"],
    "decimals": ["tizedesjegyek száma", "tizedesjegy", "decimals"],
    "amount": ["összeg", "osszeg", "átutalandó összeg", "atutalando osszeg", "amount"],
    "value_date": ["valutanap", "értéknap", "erteknap", "dátum", "datum"],
    "cost_bearer": ["költségviselés", "koltsegviseles"],
    "partner_bank_country": ["partner bank országkódja", "partner bank orszagkodja"],
    "correspondent_bank_2": ["levelező bank 2", "levelezo bank 2", "levelező bank 2. sor"],
    "correspondent_bank_3": ["levelező bank 3", "levelezo bank 3", "levelező bank 3. sor"],
    "correspondent_bank_4": ["levelező bank 4", "levelezo bank 4", "levelező bank 4. sor"],
    "correspondent_bank_address_1": ["levelező bank címe 1", "levelezo bank cime 1"],
    "correspondent_bank_address_2": ["levelező bank címe 2", "levelezo bank cime 2"],
    "correspondent_bank_address_3": ["levelező bank címe 3", "levelezo bank cime 3"],
    "hold_flag": ["hold flag", "hold"],
    "chqb_flag": ["chqb flag", "chqb"],
    "deal_ticket_flag": ["deal ticket flag"],
    "deal_ticket_date": ["deal ticket dátuma", "deal ticket datuma"],
    "deal_ticket_sequence": ["deal ticket sorszáma", "deal ticket sorszama"],
    "status": ["státusz", "statusz", "status"],
    "external_ref": ["külső referencia", "kulso referencia", "kulsref", "external ref"],
    "internal_note": ["belső megjegyzés", "belso megjegyzes"],
    "beneficiary_bank_country": ["kedvezményezett bank ország", "kedvezmenyezett bank orszag", "bank ország"],
    "process_mode": ["feldolgozási mód", "feldolgozasi mod", "process mode"],
    "commission_bearer": ["ügyfél jutalék fizetője", "ugyfel jutalek fizetoje", "fbkmk"],
    "other_fee_bearer": ["egyéb jutalék fizetője", "egyeb jutalek fizetoje", "fekmk"],
    "amount_currency_mode": ["összeg devizaneme jel", "osszeg devizaneme jel", "fodev"],
    "payment_method": ["teljesítés módja", "teljesites modja", "ftmod"],
    "priority": ["prioritás", "prioritas", "fprior"],
    "item_type": ["tétel típusa", "tetel tipusa", "fttip"],
    "iban_flag": ["iban jelzés", "iban jelzes", "fiban"],
    "message_id": ["üzenet azonosító", "uzenet azonosito", "message id", "msgid"],
    "payment_id": ["fizetési blokk azonosító", "fizetesi blokk azonosito", "payment id", "pmtinfid"],
    "debtor_name": ["terhelendő fél neve", "terhelendo fel neve", "adós neve", "ados neve", "debtor name"],
    "end_to_end_id": ["endtoend azonosító", "endtoend azonosito", "end to end id"],
}

HTML_PAGE = r"""<!doctype html>
<html lang="hu">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Banki TXT konverter</title>
  <style>
    :root {
      --bg: #f3f5f8;
      --surface: #ffffff;
      --surface-2: #f8fafc;
      --ink: #162033;
      --muted: #475569;
      --line: #d9e1ec;
      --line-strong: #c9d3df;
      --accent: #a3192c;
      --accent-dark: #841623;
      --accent-soft: #fff1f3;
      --accent-2: #0f766e;
      --accent-2-soft: #edf9f6;
      --warn: #9a6400;
      --bad: #b42331;
      --bad-soft: #fff3f4;
      --shadow: 0 18px 44px rgba(17, 24, 39, 0.08);
      --shadow-soft: 0 8px 24px rgba(17, 24, 39, 0.06);
      font-family: Inter, Segoe UI, system-ui, -apple-system, sans-serif;
    }
    * { box-sizing: border-box; }
    body {
      margin: 0;
      color: var(--ink);
      background: var(--bg);
      font-size: 14px;
    }
    .shell { width: 100%; max-width: none; margin: 0; padding: 22px clamp(16px, 3vw, 40px) 36px; }
    header {
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 20px;
      padding: 4px 0 22px;
    }
    h1 { margin: 0 0 6px; font-size: 30px; letter-spacing: 0; line-height: 1.08; }
    p { margin: 0; color: var(--muted); line-height: 1.45; }
    .eyebrow {
      color: var(--accent);
      font-weight: 800;
      font-size: 12px;
      letter-spacing: .08em;
      text-transform: uppercase;
      margin-bottom: 7px;
    }
    .header-meta {
      display: flex;
      align-items: center;
      justify-content: flex-end;
      flex-wrap: wrap;
      gap: 8px;
    }
    .badge, .status-pill {
      border: 1px solid var(--line);
      border-radius: 6px;
      background: #fff;
      padding: 10px 12px;
      color: #475569;
      font-size: 13px;
      font-weight: 650;
      white-space: nowrap;
      box-shadow: var(--shadow-soft);
    }
    .status-pill {
      display: inline-flex;
      align-items: center;
      gap: 6px;
      font-size: 12px;
      box-shadow: none;
    }
    .status-pill::before {
      content: "";
      width: 8px;
      height: 8px;
      border-radius: 50%;
      background: var(--muted);
    }
    .status-pill.ok { color: var(--accent-2); border-color: rgba(15,118,110,.28); background: var(--accent-2-soft); }
    .status-pill.ok::before { background: var(--accent-2); }
    .status-pill.warn { color: var(--warn); border-color: rgba(146,89,10,.28); background: #fff9ec; }
    .status-pill.warn::before { background: var(--warn); }
    .layout { display: grid; gap: 18px; align-items: start; }
    .panel {
      background: var(--surface);
      border: 1px solid var(--line);
      border-radius: 8px;
      box-shadow: var(--shadow);
      overflow: hidden;
    }
    .commandbar {
      display: grid;
      grid-template-columns: 1fr auto;
      gap: 12px;
      align-items: center;
      margin-bottom: 18px;
    }
    .command-summary {
      border: 1px solid var(--line);
      border-radius: 8px;
      background: #fff;
      padding: 12px 14px;
      min-height: 52px;
      display: flex;
      align-items: center;
      gap: 12px;
      box-shadow: var(--shadow-soft);
    }
    .command-summary strong { font-size: 14px; }
    .command-summary span { color: var(--muted); font-size: 12px; }
    .command-summary span[data-kind="ok"] { color: var(--accent-2); }
    .command-summary span[data-kind="warn"] { color: var(--warn); }
    .command-summary span[data-kind="bad"] { color: var(--bad); }
    .command-actions {
      display: flex;
      flex-wrap: wrap;
      justify-content: flex-end;
      gap: 8px;
      align-items: center;
    }
    .command-actions select { width: auto; min-width: 190px; }
    .import-grid {
      display: grid;
      grid-template-columns: repeat(2, minmax(220px, 1fr));
      gap: 12px;
      align-items: end;
      margin-bottom: 16px;
    }
    .import-grid .full-row { grid-column: 1 / -1; }
    .import-actions {
      display: flex;
      gap: 8px;
      align-items: center;
      flex-wrap: wrap;
      margin: 6px 0 18px;
    }
    .import-section-title {
      display: flex;
      justify-content: space-between;
      gap: 12px;
      align-items: center;
      padding-top: 14px;
      border-top: 1px solid #edf0f3;
      margin-top: 6px;
    }
    .import-section-title h3 { margin: 0; font-size: 14px; }
    .advanced-mapping {
      border-top: 1px solid #edf0f3;
      margin-top: 10px;
      padding-top: 12px;
    }
    .advanced-mapping summary {
      cursor: pointer;
      font-weight: 850;
      font-size: 14px;
      color: var(--ink);
      list-style-position: inside;
    }
    .advanced-mapping[open] summary { margin-bottom: 10px; }
    .panel h2 {
      margin: 0;
      padding: 14px 16px;
      font-size: 14px;
      letter-spacing: .01em;
      border-bottom: 1px solid var(--line);
      background: var(--surface-2);
    }
    .panel-body { padding: 16px; }
    label { display: block; font-weight: 750; font-size: 12px; margin-bottom: 6px; color: #263449; }
    input, select, button, textarea {
      font: inherit;
      color: var(--ink);
    }
    input[type="file"], input[type="text"], input[type="number"], input[type="date"], select {
      width: 100%;
      border: 1px solid var(--line);
      background: #fff;
      border-radius: 6px;
      padding: 9px 10px;
      min-height: 40px;
      font-size: 14px;
      outline: none;
      transition: border-color .15s ease, box-shadow .15s ease, background .15s ease;
    }
    input:focus, select:focus {
      border-color: rgba(15, 118, 110, .55);
      box-shadow: 0 0 0 3px rgba(15, 118, 110, .12);
    }
    input[type="file"] {
      padding: 6px;
      color: var(--muted);
      background: var(--surface-2);
    }
    input[type="file"]::file-selector-button {
      border: 0;
      border-radius: 5px;
      background: #263449;
      color: #fff;
      padding: 8px 10px;
      margin-right: 10px;
      font-weight: 750;
      cursor: pointer;
    }
    .grid-2 { display: grid; grid-template-columns: 1fr 1fr; gap: 10px; }
    .format-stack { display: grid; grid-template-columns: 1fr; gap: 10px; }
    .stack { display: grid; gap: 12px; }
    .hint { font-size: 11px; color: var(--muted); margin-top: 6px; line-height: 1.4; }
    .field-note { display: none; }
    .intro-note {
      border: 1px solid var(--line);
      border-radius: 6px;
      padding: 10px 12px;
      color: var(--muted);
      background: #fbfcfd;
      font-size: 13px;
      line-height: 1.45;
    }
    .sr-only {
      position: absolute;
      width: 1px;
      height: 1px;
      padding: 0;
      margin: -1px;
      overflow: hidden;
      clip: rect(0, 0, 0, 0);
      white-space: nowrap;
      border: 0;
    }
    .empty-state {
      border: 1px dashed var(--line-strong);
      border-radius: 8px;
      padding: 18px;
      background: #fbfcfd;
      color: var(--muted);
      display: grid;
      gap: 5px;
    }
    .empty-state strong {
      color: var(--ink);
      font-size: 14px;
    }
    .empty-state span {
      font-size: 13px;
      line-height: 1.45;
    }
    .empty-state ol {
      margin: 4px 0 0;
      padding-left: 20px;
      color: var(--muted);
      font-size: 13px;
      line-height: 1.55;
    }
    .empty-state-actions {
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
      margin-top: 6px;
    }
    .loading-row {
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 12px;
      background: var(--surface-2);
      display: grid;
      gap: 8px;
    }
    .skeleton {
      height: 12px;
      border-radius: 999px;
      background: linear-gradient(90deg, #edf2f7 0%, #f8fafc 50%, #edf2f7 100%);
      background-size: 200% 100%;
      animation: shimmer 1.1s linear infinite;
    }
    .skeleton.short { width: 42%; }
    .skeleton.medium { width: 68%; }
    @keyframes shimmer { to { background-position: -200% 0; } }
    .actions { display: grid; grid-template-columns: 1fr 1fr; gap: 8px; margin-top: 12px; }
    button, .button-link {
      border: 1px solid transparent;
      border-radius: 6px;
      padding: 9px 11px;
      font-weight: 800;
      min-height: 40px;
      cursor: pointer;
      background: #eef2f5;
      display: inline-flex;
      align-items: center;
      justify-content: center;
      text-decoration: none;
      transition: transform .12s ease, box-shadow .12s ease, border-color .12s ease, background .12s ease;
    }
    button:hover:not(:disabled), .button-link:hover { transform: translateY(-1px); box-shadow: var(--shadow-soft); }
    button:focus-visible, .button-link:focus-visible {
      outline: 3px solid rgba(15, 118, 110, .22);
      outline-offset: 2px;
    }
    button.primary { background: var(--accent); color: #fff; border-color: var(--accent); }
    button.primary:hover:not(:disabled) { background: var(--accent-dark); }
    button.secondary, .button-link.secondary { background: #fff; border-color: var(--line-strong); color: var(--ink); }
    button.ghost, .button-link.ghost { background: var(--surface-2); border-color: var(--line); color: #334155; }
    button:disabled { cursor: not-allowed; opacity: .55; transform: none; box-shadow: none; }
    button[data-loading="true"]::after {
      content: "";
      width: 12px;
      height: 12px;
      margin-left: 8px;
      border: 2px solid currentColor;
      border-right-color: transparent;
      border-radius: 50%;
      animation: spin .65s linear infinite;
    }
    @keyframes spin { to { transform: rotate(360deg); } }
    .mapping {
      display: grid;
      grid-template-columns: minmax(190px, 250px) minmax(180px, 1fr) minmax(160px, 230px);
      gap: 10px;
      align-items: center;
      padding: 10px 0;
      border-bottom: 1px solid #edf0f3;
    }
    .mapping:last-child { border-bottom: 0; }
    .req { color: var(--accent); font-weight: 800; }
    .map-help { color: var(--muted); font-size: 12px; }
    .default-input { min-width: 0; }
    .status {
      margin-top: 12px;
      border-radius: 6px;
      border: 1px solid var(--line);
      padding: 11px 12px;
      font-size: 12px;
      color: var(--muted);
      background: var(--surface-2);
      white-space: pre-wrap;
    }
    .status.ok { border-color: rgba(13,107,87,.35); color: var(--accent-2); background: #f2fbf7; }
    .status.warn { border-color: rgba(148,98,0,.35); color: var(--warn); background: #fff9ec; }
    .status.bad { border-color: rgba(165,29,42,.35); color: var(--bad); background: #fff4f5; }
    .sample-wrap { overflow: auto; border: 1px solid var(--line); border-radius: 6px; background:#fff; }
    table { border-collapse: collapse; width: 100%; min-width: 680px; font-size: 13px; }
    th, td { border-bottom: 1px solid #edf0f3; padding: 9px 10px; text-align: left; vertical-align: top; }
    th { background: #fbfcfd; font-weight: 750; color: #384554; position: sticky; top: 0; }
    td { color: #394757; }
    pre {
      margin: 0;
      max-height: 180px;
      overflow: auto;
      white-space: pre;
      border: 1px solid var(--line);
      border-radius: 6px;
      padding: 12px;
      background: #101923;
      color: #f3f7fb;
      font-size: 12px;
      min-height: 58px;
    }
    .spec {
      display: none;
      grid-template-columns: repeat(3, 1fr);
      gap: 10px;
      margin-top: 13px;
      font-size: 12px;
      color: var(--muted);
    }
    .spec div { border: 1px solid var(--line); border-radius: 6px; padding: 9px; background: #fbfcfd; }
    .compact-actions { display: grid; grid-template-columns: 1fr 1fr; gap: 8px; }
    .wide-action { grid-column: 1 / -1; }
    .result-grid { display: grid; grid-template-columns: repeat(3, minmax(160px, 1fr)); gap: 10px; }
    .metric {
      border: 1px solid var(--line);
      border-radius: 8px;
      background: linear-gradient(180deg, #fff 0%, var(--surface-2) 100%);
      padding: 13px;
      min-height: 74px;
      display: flex;
      flex-direction: column;
      justify-content: center;
      border-top: 3px solid var(--accent-2);
    }
    .metric strong { display: block; font-size: 22px; margin-bottom: 3px; line-height: 1.05; }
    .metric span { color: var(--muted); font-size: 12px; font-weight: 650; }
    .error-list { display: grid; gap: 8px; }
    .error-summary {
      border: 1px solid rgba(165,29,42,.28);
      border-radius: 8px;
      padding: 12px;
      background: var(--bad-soft);
      color: var(--bad);
    }
    .error-summary strong { display:block; margin-bottom: 8px; }
    .error-summary ul { margin: 0; padding-left: 18px; }
    .error-item { border: 1px solid rgba(165,29,42,.25); border-radius: 6px; padding: 9px; background: #fff4f5; color: var(--bad); font-size: 13px; }
    dialog {
      width: min(1120px, calc(100vw - 28px));
      border: 1px solid var(--line);
      border-radius: 8px;
      box-shadow: var(--shadow);
      padding: 0;
    }
    dialog::backdrop { background: rgba(17,24,32,.34); }
    .dialog-head { display:flex; justify-content:space-between; align-items:center; gap:12px; padding:14px 16px; border-bottom:1px solid var(--line); background:var(--surface-2); }
    .dialog-head h2 { margin:0; font-size:16px; }
    .dialog-body { padding:16px; max-height:70vh; overflow:auto; }
    .help-grid { display:grid; gap:12px; }
    .help-grid section { border-bottom:1px solid #edf0f3; padding-bottom:12px; }
    .help-grid h3 { margin:0 0 6px; font-size:14px; }
    .help-grid p { font-size:13px; }
    .account-grid { display:grid; grid-template-columns: 1fr 1fr; gap:14px; align-items:start; }
    .account-card { border:1px solid var(--line); border-radius:8px; background:#fff; padding:14px; }
    .account-card h3 { margin:0 0 12px; font-size:14px; }
    .account-actions { display:flex; flex-wrap:wrap; gap:8px; align-items:center; margin-top:10px; }
    .account-list { display:grid; gap:8px; margin-top:12px; }
    .account-row { display:grid; grid-template-columns: 1.2fr 1fr auto; gap:10px; align-items:center; border:1px solid var(--line); border-radius:8px; padding:10px; background:var(--surface-2); }
    .account-row strong { display:block; font-size:13px; }
    .account-row span { display:block; color:var(--muted); font-size:12px; margin-top:2px; }
    .validation-line { min-height:20px; font-size:12px; color:var(--muted); margin-top:6px; white-space:pre-wrap; }
    .validation-line.ok { color:var(--accent-2); }
    .validation-line.bad { color:var(--bad); }
    .registry-sample { margin-top:10px; max-height:180px; overflow:auto; border:1px solid var(--line); border-radius:6px; }
    .mapping-toolbar { display:flex; justify-content:space-between; gap:10px; align-items:center; margin-bottom:12px; }
    .muted-small { color: var(--muted); font-size: 12px; }
    @media (prefers-reduced-motion: reduce) {
      *, *::before, *::after {
        animation-duration: .01ms !important;
        animation-iteration-count: 1 !important;
        scroll-behavior: auto !important;
        transition-duration: .01ms !important;
      }
    }
    @media (max-width: 880px) {
      .shell { padding: 16px 12px 28px; }
      .commandbar { grid-template-columns: 1fr; }
      .command-actions { justify-content: stretch; }
      .command-actions > * { flex: 1 1 calc(50% - 8px); }
      .command-actions select, .command-actions .primary { flex-basis: 100%; width: 100%; }
      .import-grid { grid-template-columns: 1fr; }
      header { display: block; }
      .header-meta { justify-content: flex-start; margin-top: 12px; }
      .badge { display: inline-block; }
      .mapping { grid-template-columns: 1fr; gap: 7px; }
      .spec { grid-template-columns: 1fr; }
      .result-grid { grid-template-columns: 1fr; }
      .account-grid { grid-template-columns: 1fr; }
      .account-row { grid-template-columns: 1fr; }
    }
    @media (max-width: 720px) {
      .sample-wrap { border: 0; background: transparent; overflow: visible; }
      .sample-wrap table, .sample-wrap thead, .sample-wrap tbody, .sample-wrap tr, .sample-wrap th, .sample-wrap td {
        display: block;
        min-width: 0;
        width: 100%;
      }
      .sample-wrap thead { display: none; }
      .sample-wrap tr {
        border: 1px solid var(--line);
        border-radius: 8px;
        background: #fff;
        margin-bottom: 8px;
        overflow: hidden;
      }
      .sample-wrap td {
        display: grid;
        grid-template-columns: minmax(110px, 38%) 1fr;
        gap: 8px;
        border-bottom: 1px solid #edf0f3;
      }
      .sample-wrap td::before {
        content: attr(data-label);
        color: var(--muted);
        font-weight: 750;
      }
    }
  </style>
</head>
<body>
  <div class="shell">
    <header>
      <div>
        <div class="eyebrow">Erste EDIFACT import</div>
        <h1>Banki TXT konverter</h1>
        <p>Excel feltöltése, banki formátum választása, majd fix hosszúságú importfájl letöltése.</p>
      </div>
      <div class="header-meta">
        <div id="registryPill" class="status-pill warn" role="status" aria-live="polite">MNB tábla: betöltés</div>
        <div id="formatBadge" class="badge">PAYORD / DO · 941 karakter soronként CR/LF-fel</div>
      </div>
    </header>

    <section class="commandbar">
      <div class="command-summary">
        <strong id="commandFormatName">PAYORD / DO</strong>
        <span id="statusBox" role="status" aria-live="polite">Nyisd meg az Import menüt, válassz fájlt, majd olvasd be.</span>
      </div>
      <div class="command-actions">
        <select id="companySelect" title="Aktív cég"></select>
        <button id="companiesBtn" class="secondary" type="button">Cégek</button>
        <button id="openImportBtn" class="secondary" type="button">Import</button>
        <button id="accountsBtn" class="secondary" type="button">Saját bankszámlák</button>
        <button id="partnersBtn" class="secondary" type="button">Partnerek</button>
        <button id="convertBtn" class="primary" disabled aria-describedby="convertDisabledReason">TXT letöltése</button>
        <span id="convertDisabledReason" class="sr-only">Tölts fel fájlt és olvasd be az importot.</span>
        <a id="templateLink" class="button-link ghost" href="/template.xlsx" download>Excel sablon</a>
        <button id="helpBtn" class="ghost" type="button">? Súgó</button>
      </div>
    </section>

    <div class="layout">
      <main class="panel">
        <h2>Beolvasás eredménye</h2>
        <div class="panel-body">
          <div id="resultSummary" class="result-grid">
            <div class="metric"><strong>-</strong><span>Fejléc</span></div>
            <div class="metric"><strong>-</strong><span>Adatsor</span></div>
            <div class="metric"><strong>-</strong><span>Formátum</span></div>
          </div>
          <div id="errorArea" class="error-list" style="margin-top:14px;"></div>
          <div style="margin-top:14px;">
            <h3 style="font-size:14px;margin:0 0 8px;">Beolvasott minta</h3>
            <div id="sampleArea" class="sample-wrap"></div>
          </div>
        </div>
      </main>
    </div>

    <section class="panel" style="margin-top:16px;">
      <h2>TXT előnézet</h2>
      <div class="panel-body stack">
        <pre id="previewBox">A sikeres konvertálás után itt látszik az első rekord eleje.</pre>
      </div>
    </section>
  </div>

  <dialog id="importDialog" aria-labelledby="importDialogTitle">
    <div class="dialog-head">
      <h2 id="importDialogTitle">Import</h2>
      <button id="closeImportBtn" class="secondary" type="button">Bezárás</button>
    </div>
    <div class="dialog-body">
      <div class="import-grid">
        <div>
          <label for="bankSelect">Bank</label>
          <select id="bankSelect"></select>
        </div>
        <div>
          <label for="formatSelect">Formátum</label>
          <select id="formatSelect"></select>
          <div id="formatHelp" class="hint">A bank által elvárt TXT rekordforma.</div>
        </div>
        <div class="full-row">
          <label for="fileInput">Fájl</label>
          <input id="fileInput" type="file" accept=".xlsx,.xlsm,.csv">
        </div>
        <div>
          <label for="encoding">TXT kódolás</label>
          <select id="encoding">
            <option value="cp1250" selected>windows-1250</option>
            <option value="cp852">IBM CP852</option>
            <option value="utf-8">UTF-8</option>
          </select>
        </div>
        <div>
          <label for="identifierDate">Azonosító dátuma</label>
          <input id="identifierDate" type="date">
        </div>
      </div>
      <div class="import-actions">
        <button id="inspectBtn" class="primary">Beolvasás</button>
        <a id="templateLinkInDialog" class="button-link ghost" href="/template.xlsx" download>Excel sablon</a>
      </div>
      <details id="mappingDetails" class="advanced-mapping">
        <summary>Haladó: Excel oszlopmegfeleltetés</summary>
        <div class="import-section-title">
          <div>
            <h3>Excel oszlopok hozzárendelése</h3>
            <div class="muted-small">Nem a banki szabványt módosítja, csak azt állítja be, melyik Excel oszlop melyik fix mezőbe kerüljön.</div>
          </div>
          <button id="useGuessesBtn" class="secondary" type="button">Automatikus felismerés</button>
        </div>
        <div id="mappingArea" class="stack">
          <p>Még nincs beolvasott fejléc.</p>
        </div>
      </details>
    </div>
  </dialog>

  <dialog id="accountsDialog" aria-labelledby="accountsDialogTitle">
    <div class="dialog-head">
      <h2 id="accountsDialogTitle">Saját bankszámlák listája - <span id="accountsCompanyName"></span></h2>
      <button id="closeAccountsBtn" class="secondary" type="button">Bezárás</button>
    </div>
    <div class="dialog-body">
      <div class="account-grid">
        <section class="account-card">
          <h3>Kézi rögzítés</h3>
          <div class="grid-2">
            <div>
              <label for="ownBankCountry">Bank országa</label>
              <select id="ownBankCountry">
                <option value="HU" selected>HU - Magyarország</option>
              </select>
            </div>
            <div>
              <label for="ownBankName">Bank neve</label>
              <input id="ownBankName" type="text" placeholder="Automatikus is lehet a hitelesítő táblából">
            </div>
            <div>
              <label for="ownCurrency">Deviza</label>
              <select id="ownCurrency"></select>
            </div>
          </div>
          <div style="margin-top:10px;">
            <label for="ownAccountNumber">Számlaszám vagy HU IBAN</label>
            <input id="ownAccountNumber" type="text" placeholder="12345678-12345678-12345678 vagy HU...">
          </div>
          <div class="account-actions">
            <button id="saveAccountBtn" class="primary" type="button">Mentés</button>
          </div>
          <div id="accountStatus" class="status" role="status" aria-live="polite">Nincs kiválasztott bankszámla.</div>
        </section>

        <section class="account-card">
          <h3>Import</h3>
          <label for="accountImportFile">Bankszámla import</label>
          <input id="accountImportFile" type="file" accept=".xlsx,.xlsm,.csv">
          <div class="account-actions">
            <button id="importAccountsBtn" class="secondary" type="button">Importálás</button>
            <a class="button-link ghost" href="/accounts-template.xlsx" download>Import sablon</a>
          </div>
          <div id="registryMeta" class="status" role="status" aria-live="polite">Hitelesítő tábla állapota betöltés alatt.</div>
        </section>
      </div>

      <section class="account-card" style="margin-top:14px;">
        <h3>Rögzített bankszámlák</h3>
        <div id="accountsList" class="account-list"></div>
      </section>
    </div>
  </dialog>

  <dialog id="companiesDialog" aria-labelledby="companiesDialogTitle">
    <div class="dialog-head">
      <h2 id="companiesDialogTitle">Cégek</h2>
      <button id="closeCompaniesBtn" class="secondary" type="button">Bezárás</button>
    </div>
    <div class="dialog-body">
      <div class="account-grid">
        <section class="account-card">
          <h3>Új cég</h3>
          <label for="companyName">Cég neve</label>
          <input id="companyName" type="text" placeholder="pl. Minta Kft">
          <div class="account-actions">
            <button id="saveCompanyBtn" class="primary" type="button">Mentés</button>
          </div>
          <div id="companyStatus" class="status" role="status" aria-live="polite">A cégválasztó az import célcégét adja meg.</div>
        </section>
        <section class="account-card">
          <h3>Rögzített cégek</h3>
          <div id="companiesList" class="account-list"></div>
        </section>
      </div>
    </div>
  </dialog>

  <dialog id="partnersDialog" aria-labelledby="partnersDialogTitle">
    <div class="dialog-head">
      <h2 id="partnersDialogTitle">Partnerlista</h2>
      <button id="closePartnersBtn" class="secondary" type="button">Bezárás</button>
    </div>
    <div class="dialog-body">
      <div class="account-grid">
        <section class="account-card">
          <h3>Kézi rögzítés</h3>
          <div class="grid-2">
            <div><label for="partnerCode">Partner kód</label><input id="partnerCode" type="text"></div>
            <div><label for="partnerName">Név</label><input id="partnerName" type="text"></div>
            <div><label for="partnerAccount">Magyar számlaszám</label><input id="partnerAccount" type="text" placeholder="12345678-12345678-12345678"></div>
            <div><label for="partnerIban">IBAN</label><input id="partnerIban" type="text"></div>
            <div><label for="partnerSwift">SWIFT/BIC</label><input id="partnerSwift" type="text"></div>
            <div><label for="partnerCountry">Ország</label><input id="partnerCountry" type="text" value="HU"></div>
          </div>
          <label for="partnerAddress">Partner címe</label>
          <input id="partnerAddress" type="text">
          <div class="grid-2" style="margin-top:10px;">
            <div><label for="partnerBankName">Bank neve</label><input id="partnerBankName" type="text"></div>
            <div><label for="partnerBankAddress">Bank címe</label><input id="partnerBankAddress" type="text"></div>
          </div>
          <div class="account-actions">
            <button id="savePartnerBtn" class="primary" type="button">Mentés</button>
            <button id="lookupPartnerBankBtn" class="secondary" type="button">Bankadat keresés</button>
          </div>
          <div id="partnerStatus" class="status" role="status" aria-live="polite">Magyar számlánál az első 8 számjegyből tölt banknevet.</div>
        </section>
        <section class="account-card">
          <h3>Import</h3>
          <label for="partnerImportFile">Partner import</label>
          <input id="partnerImportFile" type="file" accept=".xlsx,.xlsm,.csv">
          <div class="account-actions">
            <button id="importPartnersBtn" class="secondary" type="button">Importálás</button>
            <a class="button-link ghost" href="/partners-template.xlsx" download>Import sablon</a>
          </div>
        </section>
      </div>
      <section class="account-card" style="margin-top:14px;">
        <h3>Rögzített partnerek</h3>
        <div id="partnersList" class="account-list"></div>
      </section>
    </div>
  </dialog>

  <dialog id="accountEditDialog" aria-labelledby="accountEditDialogTitle">
    <div class="dialog-head">
      <h2 id="accountEditDialogTitle">Bankszámla szerkesztése</h2>
      <button id="closeAccountEditBtn" class="secondary" type="button">Bezárás</button>
    </div>
    <div class="dialog-body">
      <div class="account-card">
        <div class="grid-2">
          <div>
            <label for="editBankCountry">Bank országa</label>
            <select id="editBankCountry">
              <option value="HU" selected>HU - Magyarország</option>
            </select>
          </div>
          <div>
            <label for="editBankName">Bank neve</label>
            <input id="editBankName" type="text">
          </div>
          <div>
            <label for="editCurrency">Deviza</label>
            <select id="editCurrency"></select>
          </div>
        </div>
        <div style="margin-top:10px;">
          <label for="editAccountNumber">Számlaszám vagy HU IBAN</label>
          <input id="editAccountNumber" type="text">
        </div>
        <div class="account-actions">
          <button id="saveAccountEditBtn" class="primary" type="button">Mentés</button>
          <button id="cancelAccountEditBtn" class="secondary" type="button">Mégse</button>
        </div>
        <div id="accountEditStatus" class="status" role="status" aria-live="polite">Szerkesztésre megnyitva.</div>
      </div>
    </div>
  </dialog>

  <dialog id="partnerEditDialog" aria-labelledby="partnerEditDialogTitle">
    <div class="dialog-head">
      <h2 id="partnerEditDialogTitle">Partner szerkesztése</h2>
      <button id="closePartnerEditBtn" class="secondary" type="button">Bezárás</button>
    </div>
    <div class="dialog-body">
      <div class="account-card">
        <div class="grid-2">
          <div><label for="editPartnerCode">Partner kód</label><input id="editPartnerCode" type="text"></div>
          <div><label for="editPartnerName">Név</label><input id="editPartnerName" type="text"></div>
          <div><label for="editPartnerAccount">Magyar számlaszám</label><input id="editPartnerAccount" type="text"></div>
          <div><label for="editPartnerIban">IBAN</label><input id="editPartnerIban" type="text"></div>
          <div><label for="editPartnerSwift">SWIFT/BIC</label><input id="editPartnerSwift" type="text"></div>
          <div><label for="editPartnerCountry">Ország</label><input id="editPartnerCountry" type="text"></div>
        </div>
        <label for="editPartnerAddress">Partner címe</label>
        <input id="editPartnerAddress" type="text">
        <div class="grid-2" style="margin-top:10px;">
          <div><label for="editPartnerBankName">Bank neve</label><input id="editPartnerBankName" type="text"></div>
          <div><label for="editPartnerBankAddress">Bank címe</label><input id="editPartnerBankAddress" type="text"></div>
        </div>
        <div class="account-actions">
          <button id="savePartnerEditBtn" class="primary" type="button">Mentés</button>
          <button id="cancelPartnerEditBtn" class="secondary" type="button">Mégse</button>
        </div>
        <div id="partnerEditStatus" class="status" role="status" aria-live="polite">Szerkesztésre megnyitva.</div>
      </div>
    </div>
  </dialog>

  <dialog id="helpDialog" aria-labelledby="helpDialogTitle">
    <div class="dialog-head">
      <h2 id="helpDialogTitle">Súgó</h2>
      <button id="closeHelpBtn" class="secondary" type="button">Bezárás</button>
    </div>
    <div class="dialog-body">
      <div class="help-grid">
        <section><h3>Bank és formátum</h3><p>A bank kiválasztása szűri a választható import formátumokat. Most az Erste EDIFACT forint PAYORD (DO) és deviza PAYORD (IN) formátumai érhetők el.</p></section>
        <section><h3>TXT kódolás</h3><p>A letöltött TXT karakterkészlete. Magyar banki importnál általában a windows-1250 a jó alapérték.</p></section>
        <section><h3>Azonosító dátuma</h3><p>Ha az Excelben nincs 14 számjegyű azonosító, a konverter ebből generál azonosítót: ÉÉÉÉHHNN + 6 jegyű sorszám.</p></section>
        <section><h3>Excel oszlopmegfeleltetés</h3><p>Ez nem a banki szabvány szerkesztése. Csak azt határozza meg, hogy a feltöltött Excel melyik oszlopa kerüljön a kiválasztott fix banki formátum adott mezőjébe.</p></section>
        <section><h3>Saját bankszámlák</h3><p>A listába jelenleg magyar bankszámlák rögzíthetők. A mentés ellenőrzi a 3x8-as belföldi ellenőrzőszámokat, HU IBAN esetén a MOD 97-10 IBAN szabályt és a belföldi számlarészt is.</p></section>
        <section><h3>Hibakezelés</h3><p>Konvertálás előtt az app ellenőrzi a kötelező mezőket, dátumokat, összegeket és a mezőhosszokat. A hibák a Beolvasás eredménye panelen jelennek meg.</p></section>
      </div>
    </div>
  </dialog>

<script>
const BANKS = __BANKS_JSON__;
const FORMATS = __FORMATS_JSON__;
const FIELDS = __FIELDS_JSON__;
const CURRENCIES = __CURRENCIES_JSON__;
let currentInspect = null;
let currentSettings = { active_bank: "erste", active_format: "erste_huf_payord", formats: {} };
let saveTimer = null;
let editingAccountId = "";
let editingPartnerId = "";
let accountsState = [];
let companiesState = [];
let partnersState = [];
let registryRows = [];
let lastAutoBankName = "";
let lastEditAutoBankName = "";
const dialogTriggers = new WeakMap();

const el = id => document.getElementById(id);
const today = new Date();
el("identifierDate").value = today.toISOString().slice(0, 10);

function focusableElements(root) {
  return [...root.querySelectorAll('a[href], button:not([disabled]), input:not([disabled]), select:not([disabled]), textarea:not([disabled]), [tabindex]:not([tabindex="-1"])')]
    .filter(item => item.offsetParent !== null || item === document.activeElement);
}

function openDialog(dialogId, trigger = document.activeElement) {
  const dialog = el(dialogId);
  if (!dialog) return;
  dialogTriggers.set(dialog, trigger);
  dialog.showModal();
  requestAnimationFrame(() => {
    const first = focusableElements(dialog)[0];
    (first || dialog).focus();
  });
}

function closeDialog(dialogId) {
  const dialog = el(dialogId);
  if (!dialog?.open) return;
  dialog.close();
}

function setupDialogA11y() {
  document.querySelectorAll("dialog").forEach(dialog => {
    if (dialog.dataset.a11yReady) return;
    dialog.dataset.a11yReady = "1";
    dialog.setAttribute("tabindex", "-1");
    dialog.addEventListener("keydown", event => {
      if (event.key !== "Tab") return;
      const items = focusableElements(dialog);
      if (!items.length) {
        event.preventDefault();
        dialog.focus();
        return;
      }
      const first = items[0];
      const last = items[items.length - 1];
      if (event.shiftKey && document.activeElement === first) {
        event.preventDefault();
        last.focus();
      } else if (!event.shiftKey && document.activeElement === last) {
        event.preventDefault();
        first.focus();
      }
    });
    dialog.addEventListener("close", () => {
      const trigger = dialogTriggers.get(dialog);
      if (trigger && document.contains(trigger)) {
        trigger.focus();
      }
    });
  });
}

function populateBanks() {
  const bankSelect = el("bankSelect");
  bankSelect.innerHTML = "";
  for (const [id, bank] of Object.entries(BANKS)) {
    const option = document.createElement("option");
    option.value = id;
    option.textContent = bank.label;
    bankSelect.appendChild(option);
  }
}

function populateCurrencySelect(select, selected = "HUF") {
  select.innerHTML = "";
  for (const code of CURRENCIES) {
    const option = document.createElement("option");
    option.value = code;
    option.textContent = code;
    select.appendChild(option);
  }
  select.value = CURRENCIES.includes(selected) ? selected : "HUF";
}

function populateCurrencySelects() {
  populateCurrencySelect(el("ownCurrency"), "HUF");
  populateCurrencySelect(el("editCurrency"), "HUF");
}

function activeCompanyId() {
  return el("companySelect")?.value || "default";
}

function activeCompanyName() {
  const found = companiesState.find(company => company.id === activeCompanyId());
  return found?.name || "Alap cég";
}

async function loadCompanies() {
  if (el("companiesList")) el("companiesList").innerHTML = loadingRows("Cégek betöltése...");
  const data = await fetchJson("/api/companies");
  companiesState = data.companies || [];
  const select = el("companySelect");
  const current = data.active_company_id || currentSettings.active_company_id || activeCompanyId();
  select.innerHTML = "";
  for (const company of companiesState) {
    const option = document.createElement("option");
    option.value = company.id;
    option.textContent = company.name;
    select.appendChild(option);
  }
  if ([...select.options].some(option => option.value === current)) select.value = current;
  else if (select.options.length) select.selectedIndex = 0;
  renderCompanies();
}

function renderCompanies() {
  const list = el("companiesList");
  if (!list) return;
  renderListState("companiesList", companiesState || [], company => `
    <div class="account-row">
      <div><strong>${escapeHtml(company.name)}</strong><span>${company.id === activeCompanyId() ? "Aktív" : "Cég"}</span></div>
      <div><strong>${escapeHtml(company.id)}</strong><span>Azonosító</span></div>
      <div class="account-actions" style="margin:0;">
        <button class="secondary" type="button" data-set-company="${escapeHtml(company.id)}">Kiválasztás</button>
      </div>
    </div>
  `, "Nincs rögzített cég", "Adj hozzá céget, hogy cégenként külön számlákat és partnereket kezelj.");
}

async function saveCompany() {
  const name = el("companyName").value.trim();
  if (!name) {
    setCompanyStatus("Adj meg cégnevet.", "bad");
    return;
  }
  setButtonLoading("saveCompanyBtn", true, "Mentés...");
  try {
    const data = await fetchJson("/api/companies", {
      method: "POST",
      headers: { "Content-Type": "application/json" },
      body: JSON.stringify({ name })
    });
    companiesState = data.companies || [];
    el("companyName").value = "";
    await loadCompanies();
    el("companySelect").value = data.company.id;
    await changeCompany(data.company.id);
    setCompanyStatus("Cég mentve.", "ok");
  } finally {
    setButtonLoading("saveCompanyBtn", false);
  }
}

function setCompanyStatus(text, kind = "") {
  const box = el("companyStatus");
  box.textContent = text;
  box.className = `status ${kind}`;
}

async function changeCompany(companyId) {
  currentSettings.active_company_id = companyId || activeCompanyId();
  await saveSettings();
  renderCompanies();
  if (el("accountsDialog")?.open) await loadAccounts();
  if (el("partnersDialog")?.open) await loadPartners();
  setStatus(`Aktív cég: ${activeCompanyName()}`, "ok");
}

function populateFormats() {
  const bankId = el("bankSelect").value;
  const desired = currentSettings.active_format || el("formatSelect").value;
  const formatSelect = el("formatSelect");
  formatSelect.innerHTML = "";
  for (const [id, format] of Object.entries(FORMATS)) {
    if (format.bank !== bankId) continue;
    const option = document.createElement("option");
    option.value = id;
    option.textContent = format.label;
    formatSelect.appendChild(option);
  }
  if ([...formatSelect.options].some(o => o.value === desired)) {
    formatSelect.value = desired;
  } else if (formatSelect.options.length) {
    formatSelect.selectedIndex = 0;
  }
  applySelectedFormat();
}

function selectedFormat() {
  return FORMATS[el("formatSelect").value] || Object.values(FORMATS)[0];
}

function currentFields() {
  return selectedFormat()?.fields || FIELDS;
}

function applySelectedFormat() {
  const format = selectedFormat();
  if (!format) return;
  el("formatBadge").textContent = format.badge || format.short_label || format.label;
  el("commandFormatName").textContent = format.short_label || format.label;
  el("formatHelp").textContent = format.description;
  el("templateLink").href = `/template.xlsx?format=${encodeURIComponent(el("formatSelect").value)}`;
  el("templateLinkInDialog").href = `/template.xlsx?format=${encodeURIComponent(el("formatSelect").value)}`;
  const settings = formatSettings();
  el("encoding").value = settings.encoding || format.default_encoding || "cp1250";
  if (settings.identifier_date) el("identifierDate").value = settings.identifier_date;
  buildMappings();
  updateConvertAction();
  saveSettingsDebounced();
}

function formatSettings(formatId = el("formatSelect").value) {
  currentSettings.formats ||= {};
  currentSettings.formats[formatId] ||= { mapping: {}, defaults: {} };
  currentSettings.formats[formatId].mapping ||= {};
  currentSettings.formats[formatId].defaults ||= {};
  return currentSettings.formats[formatId];
}

async function loadSettings() {
  try {
    const res = await fetch("/api/settings");
    if (res.ok) currentSettings = await res.json();
  } catch {}
}

function saveSettingsDebounced() {
  clearTimeout(saveTimer);
  saveTimer = setTimeout(saveSettings, 350);
}

async function saveSettings() {
  const formatId = el("formatSelect").value || "erste_huf_payord";
  currentSettings.active_company_id = activeCompanyId();
  currentSettings.active_bank = el("bankSelect").value || "erste";
  currentSettings.active_format = formatId;
  currentSettings.formats ||= {};
  const existing = formatSettings(formatId);
  const collected = collectConfig();
  if (!document.querySelector("[data-map]")) {
    collected.mapping = existing.mapping || {};
    collected.defaults = existing.defaults || {};
  }
  currentSettings.formats[formatId] = collected;
  try {
    await fetch("/api/settings", {
      method: "POST",
      headers: { "Content-Type": "application/json" },
      body: JSON.stringify(currentSettings)
    });
  } catch {
    setStatus("A beállítás mentése nem sikerült.", "warn");
  }
}

function setStatus(text, kind = "") {
  const box = el("statusBox");
  box.textContent = text;
  box.dataset.kind = kind;
}

function setDisabledReason(buttonId, reason = "") {
  const button = el(buttonId);
  const reasonBox = el(`${buttonId.replace("Btn", "")}DisabledReason`);
  if (!button) return;
  const disabled = Boolean(reason);
  button.disabled = disabled;
  button.title = reason || "";
  button.setAttribute("aria-disabled", disabled ? "true" : "false");
  if (reasonBox) reasonBox.textContent = reason || "";
}

function getConvertDisabledReason() {
  if (!el("fileInput").files[0]) return "Tölts fel egy Excel vagy CSV fájlt az Import panelen.";
  if (!currentInspect) return "Olvasd be a fájlt, hogy ellenőrizni lehessen a fejlécet és a sorokat.";
  if (requiredMappingMissing()) return "Ellenőrizd a kötelező Excel oszlop-hozzárendeléseket a Haladó részben.";
  return "";
}

function updateConvertAction() {
  setDisabledReason("convertBtn", getConvertDisabledReason());
}

function normalise(s) {
  return String(s || "")
    .toLowerCase()
    .normalize("NFD")
    .replace(/[\u0300-\u036f]/g, "")
    .replace(/\s+/g, " ")
    .trim();
}

function optionList(headers, selected) {
  let out = `<option value="">-- nincs oszlop --</option>`;
  for (const h of headers) {
    const safe = escapeHtml(h);
    out += `<option value="${safe}" ${h === selected ? "selected" : ""}>${safe}</option>`;
  }
  return out;
}

function escapeHtml(v) {
  return String(v ?? "").replace(/[&<>"']/g, c => ({
    "&": "&amp;", "<": "&lt;", ">": "&gt;", '"': "&quot;", "'": "&#39;"
  }[c]));
}

async function fetchJson(url, options = {}) {
  const res = await fetch(url, options);
  const contentType = res.headers.get("content-type") || "";
  const data = contentType.includes("application/json") ? await res.json() : { error: await res.text() };
  if (!res.ok) {
    throw new Error(data.error || "A kérés nem sikerült.");
  }
  return data;
}

function emptyState(title, description) {
  return `<div class="empty-state" role="status"><strong>${escapeHtml(title)}</strong><span>${escapeHtml(description)}</span></div>`;
}

function importEmptyState() {
  return `
    <div class="empty-state" role="status">
      <strong>Még nincs beolvasott adat</strong>
      <span>A konvertálás három rövid lépésből áll. Először válaszd ki az aktív céget, utána tölts fel egy Excel vagy CSV fájlt az Import panelen.</span>
      <ol>
        <li>Cég kiválasztása vagy létrehozása.</li>
        <li>Import panel megnyitása, bank és formátum ellenőrzése.</li>
        <li>Fájl beolvasása, majd TXT letöltése.</li>
      </ol>
      <div class="empty-state-actions">
        <button class="secondary" type="button" data-open-import>Import megnyitása</button>
        <a class="button-link ghost" href="/template.xlsx" download>Sablon letöltése</a>
      </div>
    </div>
  `;
}

function loadingRows(label = "Betöltés...") {
  return `
    <div class="loading-row" role="status" aria-live="polite" aria-label="${escapeHtml(label)}">
      <div class="skeleton medium"></div>
      <div class="skeleton"></div>
      <div class="skeleton short"></div>
    </div>
  `;
}

function setButtonLoading(buttonId, isLoading, labelWhenLoading = "Dolgozom...") {
  const button = el(buttonId);
  if (!button) return;
  if (isLoading) {
    if (!button.dataset.originalText) button.dataset.originalText = button.textContent;
    button.textContent = labelWhenLoading;
    button.dataset.loading = "true";
    button.disabled = true;
  } else {
    button.textContent = button.dataset.originalText || button.textContent;
    delete button.dataset.originalText;
    button.dataset.loading = "false";
    button.disabled = false;
  }
}

function renderListState(listId, rows, renderer, emptyTitle, emptyDescription) {
  const list = el(listId);
  if (!rows?.length) {
    list.innerHTML = emptyState(emptyTitle, emptyDescription);
    return;
  }
  list.innerHTML = rows.map(renderer).join("");
}

function buildMappings() {
  const headers = currentInspect?.headers || [];
  const guesses = currentInspect?.guesses || {};
  const fields = currentFields();
  const settings = formatSettings();
  const area = el("mappingArea");
  area.innerHTML = "";
  for (const [key, info] of Object.entries(fields)) {
    const row = document.createElement("div");
    row.className = "mapping";
    const required = info.required ? `<span class="req">*</span>` : "";
    const guessed = settings.mapping[key] || guesses[key] || "";
    const defaultValue = settings.defaults[key] ?? defaultFor(key);
    row.innerHTML = `
      <div>
        <strong>${escapeHtml(info.label)} ${required}</strong>
        <div class="map-help">${escapeHtml(info.help || "Excel oszlop hozzárendelése ehhez a fix banki mezőhöz.")}</div>
      </div>
      <select data-map="${escapeHtml(key)}">${optionList(headers, guessed)}</select>
      <input class="default-input" data-default="${escapeHtml(key)}" type="text"
        value="${escapeHtml(defaultValue)}" placeholder="Fix érték, ha nincs Excel oszlop">
    `;
    area.appendChild(row);
  }
}

function requiredMappingMissing() {
  const headers = currentInspect?.headers || [];
  const guesses = currentInspect?.guesses || {};
  const settings = formatSettings();
  return Object.entries(currentFields()).some(([key, info]) => {
    if (!info.required) return false;
    const mapped = settings.mapping[key] || guesses[key] || "";
    const fallback = settings.defaults[key] ?? defaultFor(key);
    return !mapped && !fallback && headers.length > 0;
  });
}

function defaultFor(key) {
  if (el("formatSelect")?.value === "erste_sepa_payord" && ["currency", "payout_currency"].includes(key)) {
    return "EUR";
  }
  if (el("formatSelect")?.value === "erste_sepa_payord" && key === "decimals") {
    return "2";
  }
  if (el("formatSelect")?.value === "unicredit_fx_ccy" && key === "cost_bearer") {
    return "SHA";
  }
  if (el("formatSelect")?.value === "unicredit_fx_ccy" && key === "legal_code") {
    return "000";
  }
  const defaults = {
    sender_currency: "HUF",
    currency: "HUF",
    payout_currency: "EUR",
    decimals: "0",
    status: "",
    beneficiary_country: "HU",
    beneficiary_bank_country: "",
    sender_account_type: "0",
    beneficiary_account_type: "0",
    swift_copy: "N",
    custom_rate_use: "N",
    urgent_use: "N",
    urgent_execution: "N",
    process_mode: "",
    group_transfer: "N",
    hold_flag: "N",
    chqb_flag: "N",
    deal_ticket_flag: "N",
    cost_bearer: "1",
    commission_bearer: "0",
    other_fee_bearer: "0",
    amount_currency_mode: " ",
    payment_method: " ",
    priority: " ",
    item_type: " ",
    iban_flag: " ",
    document_no: "",
    legal_code: "",
    debtor_name: "",
    external_ref: "",
    internal_note: ""
  };
  return defaults[key] ?? "";
}

function renderSample() {
  const headers = currentInspect?.headers || [];
  const rows = currentInspect?.sample || [];
  if (!headers.length) {
    el("sampleArea").innerHTML = importEmptyState();
    return;
  }
  let html = "<table><thead><tr>";
  for (const h of headers) html += `<th>${escapeHtml(h)}</th>`;
  html += "</tr></thead><tbody>";
  for (const r of rows) {
    html += "<tr>";
    for (let i = 0; i < headers.length; i++) html += `<td data-label="${escapeHtml(headers[i])}">${escapeHtml(r[i] ?? "")}</td>`;
    html += "</tr>";
  }
  html += "</tbody></table>";
  el("sampleArea").innerHTML = html;
}

async function inspectFile() {
  const file = el("fileInput").files[0];
  if (!file) {
    setStatus("Előbb válassz ki egy fájlt.", "warn");
    updateConvertAction();
    return;
  }
  const form = new FormData();
  form.append("file", file);
  setStatus("Beolvasás...");
  setButtonLoading("inspectBtn", true, "Beolvasás...");
  try {
    const data = await fetchJson("/api/inspect", { method: "POST", body: form });
    currentInspect = data;
    buildMappings();
    el("mappingDetails").open = requiredMappingMissing();
    renderSample();
    updateConvertAction();
    renderResultSummary(data);
    renderErrors([]);
    setStatus(`${data.headers.length} fejléc beolvasva, ${data.data_rows} adatsor észlelve.`, "ok");
    saveSettingsDebounced();
  } catch (err) {
    currentInspect = null;
    updateConvertAction();
    setStatus(err.message || "Nem sikerült beolvasni a fájlt.", "bad");
    renderErrors([err.message || "Nem sikerült beolvasni a fájlt."]);
  } finally {
    setButtonLoading("inspectBtn", false);
    updateConvertAction();
  }
}

function renderResultSummary(data = null) {
  const format = selectedFormat();
  const rows = [
    [data?.headers?.length ?? "-", "Fejléc"],
    [data?.data_rows ?? "-", "Adatsor"],
    [data?.errors?.length ?? 0, "Hibás sor"],
    [format?.short_label ?? "-", "Formátum"]
  ];
  el("resultSummary").innerHTML = rows.map(([value, label]) =>
    `<div class="metric"><strong>${escapeHtml(value)}</strong><span>${escapeHtml(label)}</span></div>`
  ).join("");
}

function renderErrors(errors) {
  const list = Array.isArray(errors) ? errors : String(errors || "").split("\n").filter(Boolean);
  if (!list.length) {
    el("errorArea").innerHTML = "";
    return;
  }
  el("errorArea").innerHTML = `
    <div class="error-summary" role="alert" tabindex="-1">
      <strong>${list.length} javítandó hiba</strong>
      <ul>${list.map(e => `<li>${escapeHtml(e)}</li>`).join("")}</ul>
    </div>
  `;
  el("errorArea").querySelector(".error-summary")?.focus();
}

function clientFormatHuAccount(value) {
  const compact = String(value || "").replace(/[^0-9]/g, "");
  const groups = compact.slice(0, 24).match(/.{1,8}/g) || [];
  return groups.join("-");
}

function formatAccountNumberElement(input) {
  const value = input.value;
  const compact = value.replace(/[\s-]+/g, "").toUpperCase();
  if (/^[A-Z]/.test(compact)) {
    input.value = compact.replace(/[^A-Z0-9]/g, "").slice(0, 34).match(/.{1,4}/g)?.join(" ") || "";
  } else {
    input.value = clientFormatHuAccount(value);
  }
}

function formatAccountNumberInput() {
  formatAccountNumberElement(el("ownAccountNumber"));
}

function accountPrefixFromInput(value) {
  const compact = String(value || "").replace(/[\s-]+/g, "").toUpperCase();
  if (compact.startsWith("HU") && compact.length >= 12) {
    return compact.slice(4, 12).replace(/\D/g, "");
  }
  if (/^[A-Z]/.test(compact)) {
    return "";
  }
  return compact.replace(/\D/g, "").slice(0, 8);
}

function autoFillBankNameFor(accountInput, bankInput, countryValue, lastValue) {
  if (countryValue !== "HU") return lastValue;
  const prefix = accountPrefixFromInput(accountInput.value);
  if (prefix.length !== 8) {
    if (bankInput.value === lastValue) bankInput.value = "";
    return "";
  }
  const found = registryRows.find(row => String(row.prefix || "") === prefix);
  if (!found?.bank_name) return lastValue;
  if (!bankInput.value || bankInput.value === lastValue) {
    bankInput.value = found.bank_name;
    return found.bank_name;
  }
  return lastValue;
}

function autoFillBankNameFromAccount() {
  lastAutoBankName = autoFillBankNameFor(el("ownAccountNumber"), el("ownBankName"), el("ownBankCountry").value, lastAutoBankName);
}

function autoFillBankNameFromEditAccount() {
  lastEditAutoBankName = autoFillBankNameFor(el("editAccountNumber"), el("editBankName"), el("editBankCountry").value, lastEditAutoBankName);
}

function updateAccountValidationHint() {
  const value = el("ownAccountNumber").value.trim();
  const compact = value.replace(/[\s-]+/g, "").toUpperCase();
  if (!value || /^[A-Z]{2}/.test(compact)) {
    return;
  }
  el("ownAccountNumber").value = clientFormatHuAccount(value);
}

function updateEditAccountFormatting() {
  const value = el("editAccountNumber").value.trim();
  const compact = value.replace(/[\s-]+/g, "").toUpperCase();
  if (!value || /^[A-Z]{2}/.test(compact)) {
    return;
  }
  el("editAccountNumber").value = clientFormatHuAccount(value);
}

function setAccountStatus(text, kind = "") {
  const box = el("accountStatus");
  box.textContent = text;
  box.className = `status ${kind}`;
}

function setAccountEditStatus(text, kind = "") {
  const box = el("accountEditStatus");
  box.textContent = text;
  box.className = `status ${kind}`;
}

function clearAccountForm() {
  lastAutoBankName = "";
  el("ownBankCountry").value = "HU";
  el("ownBankName").value = "";
  el("ownCurrency").value = "HUF";
  el("ownAccountNumber").value = "";
  updateAccountValidationHint();
  setAccountStatus("Nincs kiválasztott bankszámla.");
}

function renderAccounts(accounts) {
  accountsState = accounts || [];
  el("accountsCompanyName").textContent = activeCompanyName();
  renderListState("accountsList", accountsState, account => `
    <div class="account-row">
      <div>
        <strong>${escapeHtml(account.bank_name || "Nincs banknév")}</strong>
        <span>${escapeHtml(account.bank_country || "HU")}</span>
      </div>
      <div>
        <strong>${escapeHtml(account.account_number || "")}</strong>
        <span>${escapeHtml(account.currency || "HUF")}</span>
      </div>
      <div class="account-actions" style="margin:0;">
        <button class="secondary" type="button" data-edit-account="${escapeHtml(account.id)}">Szerkesztés</button>
        <button class="ghost" type="button" data-delete-account="${escapeHtml(account.id)}">Törlés</button>
      </div>
    </div>
  `, "Nincs saját bankszámla", "Az aktív céghez még nincs rögzített bankszámla. Rögzíts kézzel vagy importálj sablonból.");
}

async function loadAccounts() {
  el("accountsList").innerHTML = loadingRows("Saját bankszámlák betöltése...");
  const data = await fetchJson(`/api/accounts?company_id=${encodeURIComponent(activeCompanyId())}`);
  renderAccounts(data.accounts || []);
}

async function saveAccount() {
  const payload = {
    id: "",
    company_id: activeCompanyId(),
    bank_country: el("ownBankCountry").value,
    bank_name: el("ownBankName").value,
    currency: el("ownCurrency").value,
    account_number: el("ownAccountNumber").value
  };
  setAccountStatus("Mentés...");
  setButtonLoading("saveAccountBtn", true, "Mentés...");
  try {
    const data = await fetchJson(`/api/accounts?company_id=${encodeURIComponent(activeCompanyId())}`, {
      method: "POST",
      headers: { "Content-Type": "application/json" },
      body: JSON.stringify(payload)
    });
    renderAccounts(data.accounts || []);
    clearAccountForm();
    setAccountStatus("Bankszámla mentve és validálva.", "ok");
  } finally {
    setButtonLoading("saveAccountBtn", false);
  }
}

async function deleteAccount(id) {
  const res = await fetch(`/api/accounts/delete?company_id=${encodeURIComponent(activeCompanyId())}`, {
    method: "POST",
    headers: { "Content-Type": "application/json" },
    body: JSON.stringify({ id })
  });
  const data = await res.json();
  renderAccounts(data.accounts || []);
  setAccountStatus("Bankszámla törölve.", "ok");
}

function editAccount(id) {
  const account = accountsState.find(item => item.id === id);
  if (!account) return;
  editingAccountId = id;
  el("editBankCountry").value = account.bank_country || "HU";
  el("editBankName").value = account.bank_name || "";
  el("editCurrency").value = account.currency || "HUF";
  el("editAccountNumber").value = account.account_number || "";
  lastEditAutoBankName = account.bank_name || "";
  updateEditAccountFormatting();
  setAccountEditStatus("Szerkesztésre megnyitva.");
  openDialog("accountEditDialog");
}

async function saveEditedAccount() {
  if (!editingAccountId) {
    setAccountEditStatus("Nincs szerkesztésre kiválasztott bankszámla.", "bad");
    return;
  }
  const payload = {
    id: editingAccountId,
    company_id: activeCompanyId(),
    bank_country: el("editBankCountry").value,
    bank_name: el("editBankName").value,
    currency: el("editCurrency").value,
    account_number: el("editAccountNumber").value
  };
  setAccountEditStatus("Mentés...");
  const res = await fetch(`/api/accounts?company_id=${encodeURIComponent(activeCompanyId())}`, {
    method: "POST",
    headers: { "Content-Type": "application/json" },
    body: JSON.stringify(payload)
  });
  const data = await res.json();
  if (!res.ok) {
    setAccountEditStatus(data.error || "Nem sikerült menteni.", "bad");
    return;
  }
  renderAccounts(data.accounts || []);
  editingAccountId = "";
  lastEditAutoBankName = "";
  el("accountEditDialog").close();
  setAccountStatus("Bankszámla módosítva.", "ok");
}

async function importAccounts() {
  const file = el("accountImportFile").files[0];
  if (!file) {
    setAccountStatus("Válassz ki egy import fájlt.", "warn");
    return;
  }
  const form = new FormData();
  form.append("file", file);
  setAccountStatus("Importálás...");
  const res = await fetch(`/api/accounts/import?company_id=${encodeURIComponent(activeCompanyId())}`, { method: "POST", body: form });
  const data = await res.json();
  if (!res.ok) {
    setAccountStatus(data.error || "Nem sikerült importálni.", "bad");
    return;
  }
  renderAccounts(data.accounts || []);
  const errorText = (data.errors || []).length ? ` Hibák: ${(data.errors || []).join("; ")}` : "";
  setAccountStatus(`${data.added || 0} bankszámla importálva.${errorText}`, errorText ? "warn" : "ok");
}

function setPartnerStatus(text, kind = "") {
  const box = el("partnerStatus");
  box.textContent = text;
  box.className = `status ${kind}`;
}

function clearPartnerForm() {
  for (const id of ["partnerCode", "partnerName", "partnerAccount", "partnerIban", "partnerSwift", "partnerAddress", "partnerBankName", "partnerBankAddress"]) {
    el(id).value = "";
  }
  el("partnerCountry").value = "HU";
}

function autoFillPartnerBankFromHuAccount() {
  const prefix = accountPrefixFromInput(el("partnerAccount").value || el("partnerIban").value);
  if (prefix.length !== 8) return;
  const found = registryRows.find(row => String(row.prefix || "") === prefix);
  if (found?.bank_name && !el("partnerBankName").value) {
    el("partnerBankName").value = found.bank_name;
  }
}

function renderPartners(partners) {
  partnersState = partners || [];
  renderListState("partnersList", partnersState, partner => `
    <div class="account-row">
      <div>
        <strong>${escapeHtml(partner.name || "Nincs név")}</strong>
        <span>${escapeHtml(partner.partner_code || "")}</span>
      </div>
      <div>
        <strong>${escapeHtml(partner.account_number || partner.iban || "")}</strong>
        <span>${escapeHtml(partner.swift_bic || partner.bank_name || "")}</span>
      </div>
      <div class="account-actions" style="margin:0;">
        <button class="secondary" type="button" data-edit-partner="${escapeHtml(partner.id)}">Szerkesztés</button>
        <button class="ghost" type="button" data-delete-partner="${escapeHtml(partner.id)}">Törlés</button>
      </div>
    </div>
  `, "Nincs partner", "Az aktív cég partnerlistája üres. Rögzíts új partnert vagy importáld Excelből.");
}

async function loadPartners() {
  el("partnersList").innerHTML = loadingRows("Partnerek betöltése...");
  const data = await fetchJson(`/api/partners?company_id=${encodeURIComponent(activeCompanyId())}`);
  renderPartners(data.partners || []);
}

function collectPartnerPayload() {
  return {
    company_id: activeCompanyId(),
    partner_code: el("partnerCode").value,
    name: el("partnerName").value,
    account_number: el("partnerAccount").value,
    iban: el("partnerIban").value,
    swift_bic: el("partnerSwift").value,
    country: el("partnerCountry").value,
    address: el("partnerAddress").value,
    bank_name: el("partnerBankName").value,
    bank_address: el("partnerBankAddress").value
  };
}

async function savePartner() {
  setPartnerStatus("Mentés...");
  setButtonLoading("savePartnerBtn", true, "Mentés...");
  try {
    const data = await fetchJson(`/api/partners?company_id=${encodeURIComponent(activeCompanyId())}`, {
      method: "POST",
      headers: { "Content-Type": "application/json" },
      body: JSON.stringify(collectPartnerPayload())
    });
    renderPartners(data.partners || []);
    clearPartnerForm();
    setPartnerStatus("Partner mentve.", "ok");
  } finally {
    setButtonLoading("savePartnerBtn", false);
  }
}

async function deletePartner(id) {
  const res = await fetch(`/api/partners/delete?company_id=${encodeURIComponent(activeCompanyId())}`, {
    method: "POST",
    headers: { "Content-Type": "application/json" },
    body: JSON.stringify({ id, company_id: activeCompanyId() })
  });
  const data = await res.json();
  renderPartners(data.partners || []);
  setPartnerStatus("Partner törölve.", "ok");
}

async function importPartners() {
  const file = el("partnerImportFile").files[0];
  if (!file) {
    setPartnerStatus("Válassz ki import fájlt.", "warn");
    return;
  }
  const form = new FormData();
  form.append("file", file);
  setPartnerStatus("Importálás...");
  const res = await fetch(`/api/partners/import?company_id=${encodeURIComponent(activeCompanyId())}`, { method: "POST", body: form });
  const data = await res.json();
  if (!res.ok) {
    setPartnerStatus(data.error || "Nem sikerült importálni.", "bad");
    return;
  }
  renderPartners(data.partners || []);
  const errorText = (data.errors || []).length ? ` Hibák: ${(data.errors || []).slice(0, 3).join("; ")}` : "";
  setPartnerStatus(`${data.added || 0} partner importálva.${errorText}`, errorText ? "warn" : "ok");
}

function setPartnerEditStatus(text, kind = "") {
  const box = el("partnerEditStatus");
  box.textContent = text;
  box.className = `status ${kind}`;
}

function editPartner(id) {
  const partner = partnersState.find(item => item.id === id);
  if (!partner) return;
  editingPartnerId = id;
  el("editPartnerCode").value = partner.partner_code || "";
  el("editPartnerName").value = partner.name || "";
  el("editPartnerAccount").value = partner.account_number || "";
  el("editPartnerIban").value = partner.iban || "";
  el("editPartnerSwift").value = partner.swift_bic || "";
  el("editPartnerCountry").value = partner.country || "HU";
  el("editPartnerAddress").value = partner.address || "";
  el("editPartnerBankName").value = partner.bank_name || "";
  el("editPartnerBankAddress").value = partner.bank_address || "";
  setPartnerEditStatus("Szerkesztésre megnyitva.");
  openDialog("partnerEditDialog");
}

function collectEditedPartnerPayload() {
  return {
    id: editingPartnerId,
    company_id: activeCompanyId(),
    partner_code: el("editPartnerCode").value,
    name: el("editPartnerName").value,
    account_number: el("editPartnerAccount").value,
    iban: el("editPartnerIban").value,
    swift_bic: el("editPartnerSwift").value,
    country: el("editPartnerCountry").value,
    address: el("editPartnerAddress").value,
    bank_name: el("editPartnerBankName").value,
    bank_address: el("editPartnerBankAddress").value
  };
}

async function saveEditedPartner() {
  if (!editingPartnerId) {
    setPartnerEditStatus("Nincs szerkesztésre kiválasztott partner.", "bad");
    return;
  }
  setPartnerEditStatus("Mentés...");
  setButtonLoading("savePartnerEditBtn", true, "Mentés...");
  try {
    const data = await fetchJson(`/api/partners?company_id=${encodeURIComponent(activeCompanyId())}`, {
      method: "POST",
      headers: { "Content-Type": "application/json" },
      body: JSON.stringify(collectEditedPartnerPayload())
    });
    renderPartners(data.partners || []);
    editingPartnerId = "";
    el("partnerEditDialog").close();
    setPartnerStatus("Partner módosítva.", "ok");
  } finally {
    setButtonLoading("savePartnerEditBtn", false);
  }
}

async function lookupPartnerBank() {
  const params = new URLSearchParams({ bic: el("partnerSwift").value, iban: el("partnerIban").value });
  autoFillPartnerBankFromHuAccount();
  if (!params.get("bic") && !params.get("iban")) {
    setPartnerStatus("Adj meg SWIFT/BIC kódot vagy IBAN-t.", "warn");
    return;
  }
  setPartnerStatus("Bankadat keresés...");
  const res = await fetch(`/api/bic-lookup?${params.toString()}`);
  const data = await res.json();
  if (!res.ok || !data.found) {
    setPartnerStatus(data.error || "Nem találtam online bankadatot. Kézzel megadható.", "warn");
    return;
  }
  if (data.bank_name) el("partnerBankName").value = data.bank_name;
  if (data.bank_address) el("partnerBankAddress").value = data.bank_address;
  if (data.bic) el("partnerSwift").value = data.bic;
  setPartnerStatus(data.source_message || "Bankadat kitöltve.", "ok");
}

function renderRegistry(registry) {
  registryRows = registry.rows || [];
  const meta = [];
  const count = registry.row_count || (registry.rows || []).length || 0;
  if (registry.startup_message) meta.push(registry.startup_message);
  else if (registry.updated_at) meta.push(`Hitelesítő tábla betöltve: ${count} prefix, frissítve: ${registry.updated_at}.`);
  else meta.push("Hitelesítő tábla még nincs betöltve.");
  const box = el("registryMeta");
  box.textContent = meta.join(" ");
  box.className = `status ${registry.startup_ok === false ? "warn" : "ok"}`;
  const pill = el("registryPill");
  if (pill) {
    pill.textContent = registry.startup_ok === false
      ? "MNB tábla: helyi adat"
      : `MNB tábla: ${count || 0} prefix`;
    pill.className = `status-pill ${registry.startup_ok === false ? "warn" : "ok"}`;
    pill.title = meta.join(" ");
  }
}

async function loadRegistry() {
  const res = await fetch("/api/bank-registry");
  const data = await res.json();
  renderRegistry(data);
  autoFillBankNameFromAccount();
  autoFillBankNameFromEditAccount();
}

function collectConfig() {
  const mapping = {};
  const defaults = {};
  document.querySelectorAll("[data-map]").forEach(sel => mapping[sel.dataset.map] = sel.value);
  document.querySelectorAll("[data-default]").forEach(inp => defaults[inp.dataset.default] = inp.value);
  return {
    company_id: activeCompanyId(),
    bank: el("bankSelect").value || "",
    format: el("formatSelect").value || "",
    encoding: el("encoding").value,
    identifier_date: el("identifierDate").value,
    mapping,
    defaults
  };
}

async function convertFile() {
  const file = el("fileInput").files[0];
  if (!file || !currentInspect) {
    setStatus("Előbb olvasd be a fájlt.", "warn");
    return;
  }
  const form = new FormData();
  form.append("file", file);
  form.append("config", JSON.stringify(collectConfig()));
  setStatus("Konvertálás...");
  setButtonLoading("convertBtn", true, "TXT készül...");
  try {
    const res = await fetch("/api/convert", { method: "POST", body: form });
    const contentType = res.headers.get("content-type") || "";
    if (!res.ok) {
      const data = contentType.includes("application/json") ? await res.json() : {error: await res.text()};
      setStatus(data.error || "Nem sikerült konvertálni.", "bad");
      renderErrors(data.errors || data.error || []);
      return;
    }
    const buffer = await res.arrayBuffer();
    const blob = new Blob([buffer], { type: contentType || "text/plain" });
    const url = URL.createObjectURL(blob);
    const a = document.createElement("a");
    a.href = url;
    a.download = res.headers.get("x-filename") || "payord_import.txt";
    document.body.appendChild(a);
    a.click();
    a.remove();
    setTimeout(() => URL.revokeObjectURL(url), 1000);

    const decoderLabel = el("encoding").value === "cp1250" ? "windows-1250" : (el("encoding").value === "cp852" ? "ibm852" : "utf-8");
    let text = "";
    try { text = new TextDecoder(decoderLabel).decode(buffer); }
    catch { text = new TextDecoder("utf-8").decode(buffer); }
    el("previewBox").textContent = text.slice(0, 700).replace(/\r/g, "\\r").replace(/\n/g, "\\n\n");
    const rows = res.headers.get("x-record-count") || "?";
    setStatus(`${rows} rekord elkészült és letöltődött.`, "ok");
    renderErrors([]);
    saveSettingsDebounced();
  } finally {
    setButtonLoading("convertBtn", false);
    updateConvertAction();
  }
}

el("inspectBtn").addEventListener("click", () => inspectFile().catch(err => setStatus(err.message, "bad")));
el("convertBtn").addEventListener("click", () => convertFile().catch(err => setStatus(err.message, "bad")));
el("bankSelect").addEventListener("change", populateFormats);
el("formatSelect").addEventListener("change", applySelectedFormat);
el("fileInput").addEventListener("change", () => {
  currentInspect = null;
  renderSample();
  renderResultSummary();
  renderErrors([]);
  updateConvertAction();
});
el("sampleArea").addEventListener("click", event => {
  if (event.target?.closest("[data-open-import]")) {
    openDialog("importDialog", event.target.closest("[data-open-import]"));
  }
});
el("openImportBtn").addEventListener("click", event => openDialog("importDialog", event.currentTarget));
el("companiesBtn").addEventListener("click", async event => {
  openDialog("companiesDialog", event.currentTarget);
  await loadCompanies();
});
el("partnersBtn").addEventListener("click", async event => {
  openDialog("partnersDialog", event.currentTarget);
  await Promise.all([loadPartners(), loadRegistry()]);
});
el("accountsBtn").addEventListener("click", async event => {
  openDialog("accountsDialog", event.currentTarget);
  await Promise.all([loadAccounts(), loadRegistry()]);
});
el("helpBtn").addEventListener("click", event => openDialog("helpDialog", event.currentTarget));
el("closeImportBtn").addEventListener("click", () => closeDialog("importDialog"));
el("closeAccountsBtn").addEventListener("click", () => closeDialog("accountsDialog"));
el("closeCompaniesBtn").addEventListener("click", () => closeDialog("companiesDialog"));
el("closePartnersBtn").addEventListener("click", () => closeDialog("partnersDialog"));
el("closePartnerEditBtn").addEventListener("click", () => {
  editingPartnerId = "";
  closeDialog("partnerEditDialog");
});
el("cancelPartnerEditBtn").addEventListener("click", () => {
  editingPartnerId = "";
  closeDialog("partnerEditDialog");
});
el("closeAccountEditBtn").addEventListener("click", () => {
  editingAccountId = "";
  closeDialog("accountEditDialog");
});
el("cancelAccountEditBtn").addEventListener("click", () => {
  editingAccountId = "";
  closeDialog("accountEditDialog");
});
el("closeHelpBtn").addEventListener("click", () => closeDialog("helpDialog"));
el("saveAccountBtn").addEventListener("click", () => saveAccount().catch(err => setAccountStatus(err.message, "bad")));
el("saveAccountEditBtn").addEventListener("click", () => saveEditedAccount().catch(err => setAccountEditStatus(err.message, "bad")));
el("importAccountsBtn").addEventListener("click", () => importAccounts().catch(err => setAccountStatus(err.message, "bad")));
el("saveCompanyBtn").addEventListener("click", () => saveCompany().catch(err => setCompanyStatus(err.message, "bad")));
el("savePartnerBtn").addEventListener("click", () => savePartner().catch(err => setPartnerStatus(err.message, "bad")));
el("savePartnerEditBtn").addEventListener("click", () => saveEditedPartner().catch(err => setPartnerEditStatus(err.message, "bad")));
el("importPartnersBtn").addEventListener("click", () => importPartners().catch(err => setPartnerStatus(err.message, "bad")));
el("lookupPartnerBankBtn").addEventListener("click", () => lookupPartnerBank().catch(err => setPartnerStatus(err.message, "bad")));
el("companySelect").addEventListener("change", () => changeCompany(activeCompanyId()).catch(err => setStatus(err.message, "bad")));
el("ownAccountNumber").addEventListener("input", () => {
  formatAccountNumberInput();
  updateAccountValidationHint();
  autoFillBankNameFromAccount();
});
el("ownBankName").addEventListener("input", () => {
  if (el("ownBankName").value !== lastAutoBankName) lastAutoBankName = "";
});
el("editAccountNumber").addEventListener("input", () => {
  formatAccountNumberElement(el("editAccountNumber"));
  updateEditAccountFormatting();
  autoFillBankNameFromEditAccount();
});
el("editBankName").addEventListener("input", () => {
  if (el("editBankName").value !== lastEditAutoBankName) lastEditAutoBankName = "";
});
el("accountsList").addEventListener("click", event => {
  const editId = event.target?.dataset?.editAccount;
  const deleteId = event.target?.dataset?.deleteAccount;
  if (editId) editAccount(editId);
  if (deleteId) deleteAccount(deleteId).catch(err => setAccountStatus(err.message, "bad"));
});
el("companiesList").addEventListener("click", event => {
  const companyId = event.target?.dataset?.setCompany;
  if (companyId) {
    el("companySelect").value = companyId;
    changeCompany(companyId).catch(err => setCompanyStatus(err.message, "bad"));
  }
});
el("partnersList").addEventListener("click", event => {
  const editId = event.target?.dataset?.editPartner;
  const deleteId = event.target?.dataset?.deletePartner;
  if (editId) editPartner(editId);
  if (deleteId) deletePartner(deleteId).catch(err => setPartnerStatus(err.message, "bad"));
});
el("partnerAccount").addEventListener("input", () => {
  formatAccountNumberElement(el("partnerAccount"));
  autoFillPartnerBankFromHuAccount();
});
el("partnerIban").addEventListener("input", autoFillPartnerBankFromHuAccount);
el("partnerSwift").addEventListener("blur", () => {
  if (el("partnerSwift").value.trim()) lookupPartnerBank().catch(() => {});
});
el("editPartnerAccount").addEventListener("input", () => formatAccountNumberElement(el("editPartnerAccount")));

function makeDialogsDraggable() {
  document.querySelectorAll("dialog").forEach(dialog => {
    const handle = dialog.querySelector(".dialog-head");
    if (!handle || handle.dataset.draggableReady) return;
    handle.dataset.draggableReady = "1";
    handle.style.cursor = "move";
    handle.addEventListener("pointerdown", event => {
      if (event.target.closest("button")) return;
      const rect = dialog.getBoundingClientRect();
      const offsetX = event.clientX - rect.left;
      const offsetY = event.clientY - rect.top;
      dialog.style.margin = "0";
      dialog.style.left = `${rect.left}px`;
      dialog.style.top = `${rect.top}px`;
      dialog.style.position = "fixed";
      handle.setPointerCapture(event.pointerId);
      const move = moveEvent => {
        const maxLeft = Math.max(0, window.innerWidth - rect.width);
        const maxTop = Math.max(0, window.innerHeight - 56);
        const left = Math.min(Math.max(0, moveEvent.clientX - offsetX), maxLeft);
        const top = Math.min(Math.max(0, moveEvent.clientY - offsetY), maxTop);
        dialog.style.left = `${left}px`;
        dialog.style.top = `${top}px`;
      };
      const stop = () => {
        handle.removeEventListener("pointermove", move);
        handle.removeEventListener("pointerup", stop);
        handle.removeEventListener("pointercancel", stop);
      };
      handle.addEventListener("pointermove", move);
      handle.addEventListener("pointerup", stop);
      handle.addEventListener("pointercancel", stop);
    });
  });
}
el("useGuessesBtn").addEventListener("click", () => {
  const settings = formatSettings();
  settings.mapping = {...(currentInspect?.guesses || {})};
  buildMappings();
  saveSettingsDebounced();
});
el("mappingArea").addEventListener("change", saveSettingsDebounced);
el("mappingArea").addEventListener("change", updateConvertAction);
el("mappingArea").addEventListener("input", () => {
  saveSettingsDebounced();
  updateConvertAction();
});
el("encoding").addEventListener("change", saveSettingsDebounced);
el("identifierDate").addEventListener("change", saveSettingsDebounced);

(async function initApp() {
  populateBanks();
  populateCurrencySelects();
  await loadSettings();
  await loadCompanies();
  if (currentSettings.active_company_id && [...el("companySelect").options].some(option => option.value === currentSettings.active_company_id)) {
    el("companySelect").value = currentSettings.active_company_id;
  }
  if (BANKS[currentSettings.active_bank]) el("bankSelect").value = currentSettings.active_bank;
  populateFormats();
  if (FORMATS[currentSettings.active_format]) el("formatSelect").value = currentSettings.active_format;
  applySelectedFormat();
  renderResultSummary();
  renderSample();
  updateConvertAction();
  setupDialogA11y();
  makeDialogsDraggable();
  loadRegistry().catch(() => {
    const pill = el("registryPill");
    if (pill) {
      pill.textContent = "MNB tábla: nem elérhető";
      pill.className = "status-pill warn";
    }
  });
})();
</script>
</body>
</html>
"""


def json_response(handler: BaseHTTPRequestHandler, data: dict[str, Any], status: int = 200) -> None:
    body = json.dumps(data, ensure_ascii=False).encode("utf-8")
    handler.send_response(status)
    handler.send_header("Content-Type", "application/json; charset=utf-8")
    handler.send_header("Content-Length", str(len(body)))
    handler.end_headers()
    handler.wfile.write(body)


def text_response(handler: BaseHTTPRequestHandler, body: str, status: int = 200) -> None:
    data = body.encode("utf-8")
    handler.send_response(status)
    handler.send_header("Content-Type", "text/html; charset=utf-8")
    handler.send_header("Content-Length", str(len(data)))
    handler.end_headers()
    handler.wfile.write(data)


def get_upload(handler: BaseHTTPRequestHandler) -> cgi.FieldStorage:
    return cgi.FieldStorage(
        fp=handler.rfile,
        headers=handler.headers,
        environ={
            "REQUEST_METHOD": "POST",
            "CONTENT_TYPE": handler.headers.get("Content-Type", ""),
            "CONTENT_LENGTH": handler.headers.get("Content-Length", "0"),
        },
    )


def field_value(form: cgi.FieldStorage, name: str, default: str = "") -> str:
    item = form[name] if name in form else None
    if item is None:
        return default
    if isinstance(item, list):
        item = item[0]
    if getattr(item, "filename", None):
        return default
    return str(item.value or default)


def upload_bytes(form: cgi.FieldStorage) -> tuple[str, bytes]:
    if "file" not in form:
        raise ValueError("Hiányzik a feltöltött fájl.")
    item = form["file"]
    filename = Path(item.filename or "upload.xlsx").name
    data = item.file.read()
    if not data:
        raise ValueError("A feltöltött fájl üres.")
    return filename, data


def clean_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y%m%d")
    if isinstance(value, date):
        return value.strftime("%Y%m%d")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def read_csv(data: bytes) -> tuple[list[str], list[list[Any]], str]:
    try:
        text = data.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = data.decode("cp1250")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,	,")
        rows = list(csv.reader(io.StringIO(text), dialect))
    except csv.Error:
        rows = list(csv.reader(io.StringIO(text), delimiter=";"))
    return ["CSV"], rows, "CSV"


def workbook_rows(data: bytes, filename: str) -> tuple[list[str], str, list[str], list[list[Any]]]:
    header_row = 1
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        sheets, rows, selected = read_csv(data)
        return sheets, selected, header_from_rows(rows, header_row), rows[header_row:]
    if suffix not in {".xlsx", ".xlsm"}:
        raise ValueError("Csak .xlsx, .xlsm vagy .csv fájl támogatott.")
    workbook = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    sheets = workbook.sheetnames
    selected = sheets[0]
    ws = workbook[selected]
    all_rows = list(ws.iter_rows(values_only=True))
    return sheets, selected, header_from_rows(all_rows, header_row), [list(r) for r in all_rows[header_row:]]


def header_from_rows(rows: list[Any], header_row: int) -> list[str]:
    index = max(header_row - 1, 0)
    if index >= len(rows):
        return []
    raw = list(rows[index])
    headers: list[str] = []
    seen: dict[str, int] = {}
    for i, value in enumerate(raw, start=1):
        label = clean_cell(value) or f"Oszlop {i}"
        if label in seen:
            seen[label] += 1
            label = f"{label} ({seen[label]})"
        else:
            seen[label] = 1
        headers.append(label)
    return headers


def normalise(value: str) -> str:
    repl = str.maketrans("áéíóöőúüűÁÉÍÓÖŐÚÜŰ", "aeiooouuuAEIOOOUUU")
    return re.sub(r"\s+", " ", value.translate(repl).lower()).strip()


def guess_mappings(headers: list[str]) -> dict[str, str]:
    norm_headers = {normalise(h): h for h in headers}
    guesses: dict[str, str] = {}
    for key, aliases in ALIASES.items():
        norm_aliases = [normalise(a) for a in aliases]
        for alias in norm_aliases:
            if alias in norm_headers:
                guesses[key] = norm_headers[alias]
                break
        if key not in guesses:
            for nh, original in norm_headers.items():
                if any(alias in nh for alias in norm_aliases):
                    guesses[key] = original
                    break
    return guesses


def row_is_empty(row: list[Any]) -> bool:
    return all(clean_cell(v) == "" for v in row)


def row_dict(headers: list[str], row: list[Any]) -> dict[str, str]:
    result: dict[str, str] = {}
    for i, header in enumerate(headers):
        result[header] = clean_cell(row[i] if i < len(row) else "")
    return result


def parse_date(value: str, field_name: str) -> str:
    raw = clean_cell(value)
    if not raw:
        raise ValueError(f"Hiányzik a dátum: {field_name}.")
    digits = re.sub(r"\D", "", raw)
    if len(digits) == 8:
        return digits
    for fmt in ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(raw, fmt).strftime("%Y%m%d")
        except ValueError:
            pass
    raise ValueError(f"Hibás dátumformátum ({field_name}): {raw}")


def value_for(row: dict[str, str], key: str, config: dict[str, Any]) -> str:
    header = (config.get("mapping") or {}).get(key) or ""
    defaults = config.get("defaults") or {}
    value = row.get(header, "") if header else ""
    return clean_cell(value) or clean_cell(defaults.get(key, ""))


def digits_only(value: str) -> str:
    return re.sub(r"\D", "", clean_cell(value))


def amount_to_field(value: str, decimals: str) -> str:
    raw = clean_cell(value).replace(" ", "").replace("\u00a0", "")
    if not raw:
        raise ValueError("Hiányzik az összeg.")
    raw = raw.replace(",", ".")
    try:
        decimal_amount = Decimal(raw)
        decimal_places = int(clean_cell(decimals) or "0")
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"Hibás összeg vagy tizedesjegy: {value}") from exc
    multiplier = Decimal(10) ** decimal_places
    number = int((decimal_amount * multiplier).quantize(Decimal("1"), rounding=ROUND_HALF_UP))
    if number < 0:
        raise ValueError("Az összeg nem lehet negatív.")
    output = str(number).zfill(13)
    if len(output) > 13:
        raise ValueError(f"Az összeg túl hosszú a 13 karakteres mezőhöz: {value}")
    return output


def amount_to_fixed(value: str, decimals: int, length: int) -> str:
    raw = clean_cell(value).replace(" ", "").replace("\u00a0", "")
    if not raw:
        raise ValueError("Hiányzik az összeg.")
    raw = raw.replace(",", ".")
    try:
        decimal_amount = Decimal(raw)
    except InvalidOperation as exc:
        raise ValueError(f"Hibás összeg: {value}") from exc
    multiplier = Decimal(10) ** decimals
    number = int((decimal_amount * multiplier).quantize(Decimal("1"), rounding=ROUND_HALF_UP))
    if number < 0:
        raise ValueError("Az összeg nem lehet negatív.")
    output = str(number).zfill(length)
    if len(output) > length:
        raise ValueError(f"Az összeg túl hosszú a {length} karakteres mezőhöz: {value}")
    return output


def amount_with_dot(value: str, decimals: int, length: int) -> str:
    raw = clean_cell(value).replace(" ", "").replace("\u00a0", "").replace(",", ".")
    if not raw:
        raise ValueError("Hiányzik az összeg.")
    try:
        decimal_amount = Decimal(raw)
    except InvalidOperation as exc:
        raise ValueError(f"Hibás összeg: {value}") from exc
    quant = Decimal("1").scaleb(-decimals)
    formatted = f"{decimal_amount.quantize(quant, rounding=ROUND_HALF_UP):.{decimals}f}"
    if decimal_amount < 0:
        raise ValueError("Az összeg nem lehet negatív.")
    if len(formatted) > length:
        raise ValueError(f"Az összeg túl hosszú a {length} karakteres mezőhöz: {value}")
    return formatted.rjust(length, "0")


def decimal_amount_string(value: str, decimals: int = 2) -> str:
    raw = clean_cell(value).replace(" ", "").replace("\u00a0", "").replace(",", ".")
    if not raw:
        raise ValueError("Hiányzik az összeg.")
    try:
        decimal_amount = Decimal(raw)
    except InvalidOperation as exc:
        raise ValueError(f"Hibás összeg: {value}") from exc
    if decimal_amount < 0:
        raise ValueError("Az összeg nem lehet negatív.")
    quant = Decimal("1").scaleb(-decimals)
    return f"{decimal_amount.quantize(quant, rounding=ROUND_HALF_UP):.{decimals}f}"


def fixed_local_account(value: str, length: int = 24) -> str:
    digits = digits_only(value)
    if len(digits) not in {16, 24}:
        raise ValueError("A belföldi számlaszám 16 vagy 24 számjegy legyen.")
    return digits.ljust(length, " ")


def checksum_number(value: int, length: int) -> str:
    return str(value)[-length:].zfill(length)


def checksum_text_number(value: str, length: int) -> int:
    normalized = "".join(ch if ch.isdigit() else "0" for ch in value)
    return int(normalized or "0") % (10 ** length)


def optional_date_or_zero(value: str) -> str:
    if not clean_cell(value):
        return "00000000"
    return parse_date(value, "értéknap")


def yn_flag(value: str, default: str = "N") -> str:
    raw = clean_cell(value).upper()
    if raw in {"", "0", "N", "NO", "NEM", "FALSE"}:
        return default if raw == "" else "N"
    if raw in {"1", "Y", "YES", "IGEN", "TRUE"}:
        return "Y"
    raise ValueError(f"Y/N értéket vártam, ezt kaptam: {value}")


def fit_text(value: str, length: int, pad: str = " ", left: bool = False) -> str:
    text = clean_cell(value)
    if len(text) > length:
        text = text[:length]
    if left:
        return text.rjust(length, pad)
    return text.ljust(length, pad)


def fit_digits(value: str, length: int, pad_left: bool = False) -> str:
    text = digits_only(value)
    if len(text) > length:
        text = text[:length]
    return text.zfill(length) if pad_left else text.ljust(length, " ")


def set_field(buffer: list[str], start: int, length: int, value: str) -> None:
    text = value
    if len(text) != length:
        raise ValueError(f"Belső mezőhossz hiba a {start}. pozíciónál: {len(text)} != {length}")
    begin = start - 1
    buffer[begin:begin + length] = list(text)


def normalise_status(value: str) -> str:
    raw = clean_cell(value).upper()
    if raw in {"", "NORMAL", "N", "0"}:
        return " "
    if raw in {"V", "VIBER"}:
        return "V"
    if raw in {"T", "TPLUSONE", "T+1"}:
        return "T"
    raise ValueError(f"Ismeretlen státusz: {value}")


def optional_date(value: str, field_name: str) -> str:
    if not clean_cell(value):
        return " " * 8
    return parse_date(value, field_name)


def build_payord_line(values: dict[str, str], sequence: int, config: dict[str, Any]) -> str:
    line = [" "] * 939
    decimals = clean_cell(values["decimals"] or "0")
    identifier = clean_cell(values["identifier"])
    if not identifier:
        identifier_date = parse_date(config.get("identifier_date") or date.today().strftime("%Y-%m-%d"), "azonosító dátuma")
        identifier = f"{identifier_date}{sequence:06d}"
    identifier = digits_only(identifier)
    if len(identifier) != 14:
        raise ValueError("Az azonosító 14 számjegy legyen, vagy hagyd üresen az automatikus generáláshoz.")

    set_field(line, 1, 6, "PAYORD")
    set_field(line, 7, 2, "DO")
    set_field(line, 9, 14, identifier)
    set_field(line, 23, 47, fit_digits(values["sender_account"], 47))
    set_field(line, 70, 1, "0")
    set_field(line, 71, 32, fit_text(values["sender_name"], 32))
    set_field(line, 220, 47, fit_digits(values["beneficiary_account"], 47))
    set_field(line, 267, 1, "0")
    set_field(line, 332, 32, fit_text(values["beneficiary_name"], 32))
    set_field(line, 472, 2, fit_text(values["beneficiary_country"], 2))
    set_field(line, 593, 96, fit_text(values["note"], 96))
    set_field(line, 689, 6, fit_text(values["document_no"], 6))
    set_field(line, 695, 4, fit_text(values["legal_code"], 4))
    set_field(line, 806, 3, fit_text(values["currency"].upper(), 3))
    set_field(line, 809, 1, fit_digits(decimals, 1))
    set_field(line, 810, 13, amount_to_field(values["amount"], decimals))
    set_field(line, 835, 8, parse_date(values["value_date"], "valutanap"))
    set_field(line, 843, 1, normalise_status(values["status"]))
    set_field(line, 938, 2, "00")
    return "".join(line)


def build_fx_payord_line(values: dict[str, str], sequence: int, config: dict[str, Any]) -> str:
    line = [" "] * 1046
    decimals = clean_cell(values["decimals"] or "2")
    identifier = clean_cell(values["identifier"])
    if not identifier:
        identifier_date = parse_date(config.get("identifier_date") or date.today().strftime("%Y-%m-%d"), "azonosító dátuma")
        identifier = f"{identifier_date}{sequence:06d}"
    identifier = digits_only(identifier)
    if len(identifier) != 14:
        raise ValueError("Az azonosító 14 számjegy legyen, vagy hagyd üresen az automatikus generáláshoz.")

    set_field(line, 1, 6, "PAYORD")
    set_field(line, 7, 2, "IN")
    set_field(line, 9, 14, identifier)
    set_field(line, 23, 47, fit_text(values["sender_account"], 47))
    set_field(line, 70, 1, fit_text(values["sender_account_type"] or "0", 1))
    set_field(line, 71, 140, fit_text(values["sender_name"], 140))
    set_field(line, 220, 47, fit_text(values["beneficiary_account"], 47))
    set_field(line, 267, 1, fit_text(values["beneficiary_account_type"] or "0", 1))
    set_field(line, 268, 64, fit_text(values["beneficiary_bank_name"], 64))
    set_field(line, 332, 140, fit_text(values["beneficiary_name"], 140))
    set_field(line, 472, 2, fit_text(values["beneficiary_country"], 2))
    set_field(line, 474, 14, fit_text(values["beneficiary_bank_swift"], 14))
    set_field(line, 489, 35, fit_text(values["correspondent_bank_1"], 35))
    set_field(line, 525, 1, fit_text(values["swift_copy"] or "N", 1))
    set_field(line, 526, 15, fit_text(values["fax_number"], 15))
    set_field(line, 541, 1, fit_text(values["custom_rate_use"] or "N", 1))
    set_field(line, 542, 15, fit_text(values["custom_rate"], 15))
    set_field(line, 558, 35, fit_text(values["bank_message"], 35))
    set_field(line, 593, 96, fit_text(values["note"], 96))
    set_field(line, 689, 4, fit_text(values["legal_code"], 4))
    set_field(line, 753, 24, fit_text(values["permit_no"], 24))
    set_field(line, 777, 8, optional_date(values["permit_date"], "deviza engedély dátuma"))
    set_field(line, 785, 1, fit_text(values["urgent_use"] or "N", 1))
    set_field(line, 786, 1, fit_text(values["urgent_execution"], 1))
    set_field(line, 788, 1, fit_text(values["fax_copy_execution"], 1))
    set_field(line, 789, 3, fit_text(values["payout_currency"].upper(), 3))
    set_field(line, 806, 3, fit_text(values["currency"].upper(), 3))
    set_field(line, 809, 1, fit_digits(decimals, 1))
    set_field(line, 810, 13, amount_to_field(values["amount"], decimals))
    set_field(line, 835, 8, optional_date(values["value_date"], "feldolgozás kezdő dátuma"))
    set_field(line, 855, 1, fit_text(values["cost_bearer"] or "1", 1))
    set_field(line, 938, 2, "00")
    set_field(line, 940, 2, fit_text(values["partner_bank_country"], 2))
    set_field(line, 942, 35, fit_text(values["correspondent_bank_2"], 35))
    set_field(line, 977, 35, fit_text(values["correspondent_bank_3"], 35))
    set_field(line, 1012, 35, fit_text(values["correspondent_bank_4"], 35))
    return "".join(line)


def build_sepa_payord_line(values: dict[str, str], sequence: int, config: dict[str, Any]) -> str:
    line = [" "] * 1046
    decimals = clean_cell(values.get("decimals") or "2")
    payout_currency = clean_cell(values.get("payout_currency") or "EUR").upper()
    currency = clean_cell(values.get("currency") or "EUR").upper()
    if payout_currency != "EUR" or currency != "EUR":
        raise ValueError("SEPA átutalásnál a kifizetendő és az átutalandó deviza is EUR kell legyen.")
    iban_result = validate_iban_syntax(values.get("beneficiary_account", ""))
    if not iban_result["ok"]:
        raise ValueError("Címzett IBAN: " + "; ".join(iban_result["errors"]))

    identifier = clean_cell(values.get("identifier", ""))
    if not identifier:
        identifier_date = parse_date(config.get("identifier_date") or date.today().strftime("%Y-%m-%d"), "azonosító dátuma")
        identifier = f"{identifier_date}{sequence:06d}"
    identifier = digits_only(identifier)
    if len(identifier) != 14:
        raise ValueError("Az azonosító 14 számjegy legyen, vagy hagyd üresen az automatikus generáláshoz.")

    urgent_execution = clean_cell(values.get("urgent_execution", ""))
    set_field(line, 1, 6, "PAYORD")
    set_field(line, 7, 2, "IN")
    set_field(line, 9, 14, identifier)
    set_field(line, 23, 47, fit_text(values.get("sender_account", ""), 47))
    set_field(line, 70, 1, fit_text(values.get("sender_account_type") or "0", 1))
    set_field(line, 71, 140, fit_text(values.get("sender_name", ""), 140))
    set_field(line, 220, 47, fit_text(iban_result["iban"], 47))
    set_field(line, 267, 1, fit_text(values.get("beneficiary_account_type") or "0", 1))
    set_field(line, 332, 140, fit_text(values.get("beneficiary_name", ""), 140))
    set_field(line, 474, 14, fit_text(values.get("beneficiary_bank_swift", ""), 14))
    set_field(line, 593, 96, fit_text(values.get("note", ""), 96))
    set_field(line, 785, 1, "Y" if urgent_execution else "N")
    set_field(line, 786, 1, fit_text(urgent_execution, 1))
    set_field(line, 789, 3, "EUR")
    set_field(line, 806, 3, "EUR")
    set_field(line, 809, 1, fit_digits(decimals, 1))
    set_field(line, 810, 13, amount_to_field(values.get("amount", ""), decimals))
    set_field(line, 938, 2, "00")
    return "".join(line)


def build_kh_fx_payord_line(values: dict[str, str], sequence: int, config: dict[str, Any]) -> str:
    line = [" "] * 939
    decimals = clean_cell(values.get("decimals") or "2")
    identifier = clean_cell(values.get("identifier", ""))
    if not identifier:
        identifier_date = parse_date(config.get("identifier_date") or date.today().strftime("%Y-%m-%d"), "azonosító dátuma")
        identifier = f"{identifier_date}{sequence:06d}"
    identifier = digits_only(identifier)
    if len(identifier) != 14:
        raise ValueError("Az azonosító 14 számjegy legyen, vagy hagyd üresen az automatikus generáláshoz.")

    set_field(line, 1, 6, "PAYORD")
    set_field(line, 7, 2, "IN")
    set_field(line, 9, 14, identifier)
    set_field(line, 23, 47, fit_text(values.get("sender_account", ""), 47))
    set_field(line, 70, 1, " ")
    set_field(line, 71, 32, fit_text(values.get("sender_name", ""), 32))
    set_field(line, 220, 47, fit_text(values.get("beneficiary_account", ""), 47))
    set_field(line, 267, 1, " ")
    set_field(line, 268, 35, fit_text(values.get("beneficiary_bank_name", ""), 35))
    set_field(line, 303, 29, fit_text(values.get("beneficiary_bank_address", ""), 29))
    set_field(line, 332, 105, fit_text(values.get("beneficiary_name", ""), 105))
    set_field(line, 474, 15, fit_text(values.get("beneficiary_bank_swift", ""), 15))
    set_field(line, 558, 35, fit_text(values.get("bank_message", ""), 35))
    set_field(line, 593, 96, fit_text(values.get("note", ""), 96))
    set_field(line, 789, 3, fit_text(values.get("payout_currency", "").upper(), 3))
    set_field(line, 806, 3, fit_text(values.get("currency", "").upper(), 3))
    set_field(line, 809, 1, fit_digits(decimals, 1))
    set_field(line, 810, 13, amount_to_field(values.get("amount", ""), decimals))
    set_field(line, 835, 8, optional_date(values.get("value_date", ""), "valutanap"))
    set_field(line, 855, 1, fit_text(values.get("cost_bearer") or "1", 1))
    set_field(line, 888, 1, yn_flag(values.get("urgent_execution", ""), "N"))
    set_field(line, 889, 1, yn_flag(values.get("group_transfer", ""), "N"))
    set_field(line, 938, 2, "00")
    return "".join(line)


def build_otp_huf_payord_line(values: dict[str, str], sequence: int, config: dict[str, Any]) -> str:
    line = [" "] * 939
    decimals = clean_cell(values.get("decimals") or "0")
    identifier = clean_cell(values.get("identifier", ""))
    if not identifier:
        identifier_date = parse_date(config.get("identifier_date") or date.today().strftime("%Y-%m-%d"), "azonosító dátuma")
        identifier = f"{identifier_date}{sequence:06d}"
    identifier = digits_only(identifier)
    if len(identifier) != 14:
        raise ValueError("Az azonosító 14 számjegy legyen, vagy hagyd üresen az automatikus generáláshoz.")

    process_mode = clean_cell(values.get("process_mode", "")).upper()
    if process_mode in {"NORMAL", "N", "0"}:
        process_mode = " "
    if process_mode not in {"", " ", "V", "T", "W", "P"}:
        raise ValueError("OTP feldolgozási mód csak üres, V, T, W vagy P lehet.")

    set_field(line, 1, 6, "PAYORD")
    set_field(line, 7, 2, "DO")
    set_field(line, 9, 14, identifier)
    set_field(line, 23, 47, fit_digits(values.get("sender_account", ""), 47))
    set_field(line, 70, 1, "0")
    set_field(line, 71, 32, fit_text(values.get("sender_name", ""), 32))
    set_field(line, 220, 47, fit_digits(values.get("beneficiary_account", ""), 47))
    set_field(line, 267, 1, "0")
    set_field(line, 332, 32, fit_text(values.get("beneficiary_name", ""), 32))
    set_field(line, 472, 2, "  ")
    set_field(line, 593, 96, fit_text(values.get("note", ""), 96))
    set_field(line, 689, 6, fit_text(values.get("document_no", ""), 6))
    set_field(line, 695, 4, "    ")
    set_field(line, 806, 3, fit_text(values.get("currency", "").upper(), 3))
    set_field(line, 809, 1, fit_digits(decimals, 1))
    set_field(line, 810, 13, amount_to_field(values.get("amount", ""), decimals))
    set_field(line, 835, 8, optional_date(values.get("value_date", ""), "értéknap"))
    set_field(line, 843, 1, process_mode or " ")
    set_field(line, 938, 2, "00")
    return "".join(line)


def build_otp_fx_payord_line(values: dict[str, str], sequence: int, config: dict[str, Any]) -> str:
    line = [" "] * 939
    decimals = clean_cell(values.get("decimals") or "2")
    identifier = clean_cell(values.get("identifier", ""))
    if not identifier:
        identifier_date = parse_date(config.get("identifier_date") or date.today().strftime("%Y-%m-%d"), "azonosító dátuma")
        identifier = f"{identifier_date}{sequence:06d}"
    identifier = digits_only(identifier)
    if len(identifier) != 14:
        raise ValueError("Az azonosító 14 számjegy legyen, vagy hagyd üresen az automatikus generáláshoz.")

    urgent = clean_cell(values.get("urgent_execution", ""))
    if urgent and urgent not in {"0", "1", "2"}:
        raise ValueError("OTP sürgős teljesítés mező csak üres, 0, 1 vagy 2 lehet.")

    set_field(line, 1, 6, "PAYORD")
    set_field(line, 7, 2, "IN")
    set_field(line, 9, 14, identifier)
    set_field(line, 23, 47, fit_text(values.get("sender_account", ""), 47))
    set_field(line, 70, 1, fit_text(values.get("sender_account_type") or "0", 1))
    set_field(line, 71, 140, fit_text(values.get("sender_name", ""), 140))
    set_field(line, 220, 47, fit_text(values.get("beneficiary_account", ""), 47))
    set_field(line, 267, 1, fit_text(values.get("beneficiary_account_type") or "0", 1))
    set_field(line, 268, 35, fit_text(values.get("beneficiary_bank_name", ""), 35))
    set_field(line, 303, 29, fit_text(values.get("beneficiary_bank_address", ""), 29))
    set_field(line, 332, 140, fit_text(values.get("beneficiary_name", ""), 140))
    set_field(line, 474, 14, fit_text(values.get("beneficiary_bank_swift", ""), 14))
    set_field(line, 489, 35, fit_text(values.get("correspondent_bank_1", ""), 35))
    set_field(line, 558, 35, fit_text(values.get("bank_message", ""), 35))
    set_field(line, 593, 96, fit_text(values.get("note", ""), 96))
    set_field(line, 753, 24, fit_text(values.get("permit_no", ""), 24))
    set_field(line, 777, 8, optional_date(values.get("permit_date", ""), "deviza engedély dátuma"))
    set_field(line, 786, 1, urgent or " ")
    set_field(line, 789, 3, fit_text(values.get("payout_currency", "").upper(), 3))
    set_field(line, 806, 3, fit_text(values.get("currency", "").upper(), 3))
    set_field(line, 809, 1, fit_digits(decimals, 1))
    set_field(line, 810, 13, amount_to_field(values.get("amount", ""), decimals))
    set_field(line, 855, 1, fit_text(values.get("cost_bearer") or "1", 1))
    set_field(line, 938, 2, "00")
    return "".join(line)


def build_raiffeisen_huf_line(values: dict[str, str], sequence: int, config: dict[str, Any]) -> str:
    line = [" "] * 249
    set_field(line, 1, 15, amount_with_dot(values.get("amount", ""), 2, 15))
    set_field(line, 16, 8, optional_date(values.get("value_date", ""), "értéknap"))
    set_field(line, 24, 13, " " * 13)
    set_field(line, 37, 24, fixed_local_account(values.get("sender_account", ""), 24))
    set_field(line, 61, 24, fixed_local_account(values.get("beneficiary_account", ""), 24))
    set_field(line, 85, 32, fit_text(values.get("beneficiary_name", ""), 32))
    set_field(line, 117, 2, fit_text(values.get("beneficiary_country", ""), 2))
    set_field(line, 119, 3, fit_text(values.get("legal_code", ""), 3))
    set_field(line, 137, 70, fit_text(values.get("note", ""), 70))
    set_field(line, 213, 10, fit_text(values.get("external_ref", ""), 10))
    set_field(line, 223, 6, fit_text(values.get("document_no", ""), 6))
    set_field(line, 229, 20, fit_text(values.get("internal_note", ""), 20))
    return "".join(line)


def build_raiffeisen_fx_line(values: dict[str, str], sequence: int, config: dict[str, Any]) -> str:
    line = [" "] * 493
    note = fit_text(values.get("note", ""), 120)
    beneficiary_name = fit_text(values.get("beneficiary_name", ""), 105)
    set_field(line, 1, 3, fit_text(values.get("currency", "").upper(), 3))
    set_field(line, 4, 3, fit_text(values.get("sender_currency", "").upper(), 3))
    set_field(line, 7, 15, amount_with_dot(values.get("amount", ""), 3, 15))
    set_field(line, 22, 8, optional_date(values.get("value_date", ""), "értéknap"))
    set_field(line, 30, 13, fit_digits(values.get("sender_account", ""), 13))
    set_field(line, 43, 30, beneficiary_name[0:30])
    set_field(line, 73, 30, beneficiary_name[30:60])
    set_field(line, 103, 30, beneficiary_name[60:90])
    set_field(line, 133, 15, beneficiary_name[90:105])
    set_field(line, 148, 30, fit_text(values.get("beneficiary_bank_name", ""), 30))
    set_field(line, 178, 30, " " * 30)
    set_field(line, 208, 26, fit_text(values.get("beneficiary_bank_country", ""), 26))
    set_field(line, 234, 30, fit_text(values.get("beneficiary_bank_address_1", ""), 30))
    set_field(line, 264, 20, fit_text(values.get("beneficiary_bank_address_2", ""), 20))
    set_field(line, 284, 35, fit_text(values.get("beneficiary_account", ""), 35))
    set_field(line, 319, 4, fit_text(values.get("legal_code", ""), 4))
    set_field(line, 323, 30, note[0:30])
    set_field(line, 353, 30, note[30:60])
    set_field(line, 383, 30, note[60:90])
    set_field(line, 413, 30, note[90:120])
    set_field(line, 443, 1, fit_text(values.get("commission_bearer") or "0", 1))
    set_field(line, 444, 1, fit_text(values.get("other_fee_bearer") or "0", 1))
    set_field(line, 445, 10, fit_text(values.get("permit_no", ""), 10))
    set_field(line, 455, 6, fit_text(values.get("document_no", ""), 6))
    set_field(line, 461, 11, fit_text(values.get("beneficiary_bank_swift", ""), 11))
    set_field(line, 472, 1, fit_text(values.get("amount_currency_mode") or " ", 1))
    set_field(line, 473, 1, fit_text(values.get("payment_method") or " ", 1))
    set_field(line, 474, 1, " ")
    set_field(line, 475, 1, fit_text(values.get("priority") or " ", 1))
    set_field(line, 476, 1, fit_text(values.get("item_type") or " ", 1))
    set_field(line, 477, 1, fit_text(values.get("iban_flag") or " ", 1))
    set_field(line, 478, 2, fit_text(values.get("beneficiary_country", ""), 2))
    set_field(line, 480, 3, "   ")
    set_field(line, 483, 10, fit_text(values.get("external_ref", ""), 10))
    return "".join(line)


def xml_text(parent: ET.Element, tag: str, text: str | None = None) -> ET.Element:
    child = ET.SubElement(parent, tag)
    if text is not None:
        child.text = text
    return child


def pain_account(parent: ET.Element, account: str) -> None:
    acct_id = xml_text(parent, "Id")
    cleaned = clean_cell(account).replace(" ", "")
    if re.match(r"^[A-Z]{2}[0-9A-Z]+$", cleaned.upper()):
        xml_text(acct_id, "IBAN", cleaned.upper())
    else:
        othr = xml_text(acct_id, "Othr")
        xml_text(othr, "Id", digits_only(cleaned) or cleaned)


def build_pain_xml(records: list[dict[str, str]], config: dict[str, Any], service_level: str | None = None, version: str = "pain.001.001.03") -> str:
    if not records:
        raise ValueError("Nem találtam konvertálható adatsort.")
    ns = f"urn:iso:std:iso:20022:tech:xsd:{version}"
    ET.register_namespace("", ns)
    doc = ET.Element(f"{{{ns}}}Document")
    cstmr = xml_text(doc, "CstmrCdtTrfInitn")
    now = datetime.now().replace(microsecond=0).isoformat()
    msg_id = clean_cell(records[0].get("message_id")) or f"MSG{datetime.now().strftime('%Y%m%d%H%M%S')}"
    total = sum(Decimal(decimal_amount_string(r.get("amount", ""), 2)) for r in records)

    grp = xml_text(cstmr, "GrpHdr")
    xml_text(grp, "MsgId", msg_id)
    xml_text(grp, "CreDtTm", now)
    xml_text(grp, "NbOfTxs", str(len(records)))
    xml_text(grp, "CtrlSum", f"{total:.2f}")
    initg = xml_text(grp, "InitgPty")
    xml_text(initg, "Nm", clean_cell(records[0].get("debtor_name"))[:70])

    pmt = xml_text(cstmr, "PmtInf")
    xml_text(pmt, "PmtInfId", clean_cell(records[0].get("payment_id")) or f"PMT{datetime.now().strftime('%Y%m%d%H%M%S')}")
    xml_text(pmt, "PmtMtd", "TRF")
    if service_level:
        pti = xml_text(pmt, "PmtTpInf")
        svc = xml_text(pti, "SvcLvl")
        xml_text(svc, "Cd", service_level)
    req_date = parse_date(records[0].get("value_date", ""), "értéknap")
    xml_text(pmt, "ReqdExctnDt", f"{req_date[:4]}-{req_date[4:6]}-{req_date[6:8]}")
    dbtr = xml_text(pmt, "Dbtr")
    xml_text(dbtr, "Nm", clean_cell(records[0].get("debtor_name"))[:70])
    dbtr_acct = xml_text(pmt, "DbtrAcct")
    pain_account(dbtr_acct, records[0].get("sender_account", ""))

    for seq, values in enumerate(records, start=1):
        cdt = xml_text(pmt, "CdtTrfTxInf")
        pmtid = xml_text(cdt, "PmtId")
        xml_text(pmtid, "EndToEndId", clean_cell(values.get("end_to_end_id")) or f"NOTPROVIDED{seq:06d}")
        amt = xml_text(cdt, "Amt")
        instd = xml_text(amt, "InstdAmt", decimal_amount_string(values.get("amount", ""), 2))
        instd.set("Ccy", clean_cell(values.get("currency") or "EUR").upper())
        cb = clean_cell(values.get("cost_bearer") or ("SLEV" if service_level == "SEPA" else "SHAR")).upper()
        xml_text(cdt, "ChrgBr", cb)
        if clean_cell(values.get("beneficiary_bank_swift", "")):
            agent = xml_text(cdt, "CdtrAgt")
            fin = xml_text(agent, "FinInstnId")
            xml_text(fin, "BIC", clean_cell(values.get("beneficiary_bank_swift", "")).upper())
        cdtr = xml_text(cdt, "Cdtr")
        xml_text(cdtr, "Nm", clean_cell(values.get("beneficiary_name", ""))[:70])
        if clean_cell(values.get("beneficiary_country", "")):
            adr = xml_text(cdtr, "PstlAdr")
            xml_text(adr, "Ctry", clean_cell(values.get("beneficiary_country", "")).upper()[:2])
        cdtr_acct = xml_text(cdt, "CdtrAcct")
        pain_account(cdtr_acct, values.get("beneficiary_account", ""))
        if clean_cell(values.get("note", "")):
            rmt = xml_text(cdt, "RmtInf")
            xml_text(rmt, "Ustrd", clean_cell(values.get("note", ""))[:140])

    raw = ET.tostring(doc, encoding="utf-8")
    pretty = minidom.parseString(raw).toprettyxml(indent="  ", encoding="utf-8")
    return pretty.decode("utf-8")


def build_unicredit_pay_text(records: list[dict[str, str]], config: dict[str, Any]) -> str:
    if not records:
        raise ValueError("Nem találtam konvertálható adatsort.")
    first = records[0]
    creation_date = parse_date(config.get("identifier_date") or date.today().strftime("%Y-%m-%d"), "létrehozás dátuma")
    sender_account = fixed_local_account(first.get("sender_account", ""))
    sender_currency = fit_text(first.get("sender_currency") or "HUF", 3)

    header = [" "] * 256
    set_field(header, 1, 2, "23")
    set_field(header, 3, 24, sender_account)
    set_field(header, 27, 3, sender_currency)
    set_field(header, 30, 1, "1")
    set_field(header, 31, 1, "0")
    set_field(header, 32, 8, creation_date)

    lines = ["".join(header)]
    record_sum = 0
    account_sum = 0
    amount_sum = 0
    for sequence, values in enumerate(records, start=1):
        if fixed_local_account(values.get("sender_account", "")) != sender_account:
            raise ValueError("UniCredit PAY fájlban egy csomagon belül csak egy terhelendő számlaszám szerepelhet.")
        if fit_text(values.get("sender_currency") or "HUF", 3) != sender_currency:
            raise ValueError("UniCredit PAY fájlban egy csomagon belül csak egy terhelendő számladeviza szerepelhet.")
        currency = clean_cell(values.get("currency") or "HUF").upper()
        decimals = 0 if currency in {"HUF", "JPY"} else 2
        account = fixed_local_account(values.get("beneficiary_account", ""))
        amount = amount_to_fixed(values.get("amount", ""), 0, 13) + "00" if decimals == 0 else amount_to_fixed(values.get("amount", ""), 2, 15)

        detail = [" "] * 256
        set_field(detail, 1, 2, "43")
        set_field(detail, 3, 6, checksum_number(sequence, 6))
        set_field(detail, 9, 24, account)
        set_field(detail, 33, 32, fit_text(values.get("beneficiary_name", ""), 32))
        set_field(detail, 65, 6, fit_text(values.get("document_no", ""), 6))
        note = fit_text(values.get("note", ""), 96)
        set_field(detail, 71, 32, note[0:32])
        set_field(detail, 103, 32, note[32:64])
        set_field(detail, 135, 32, note[64:96])
        set_field(detail, 167, 15, amount)
        set_field(detail, 182, 3, fit_text(currency, 3))
        set_field(detail, 185, 8, optional_date_or_zero(values.get("value_date", "")))
        set_field(detail, 193, 3, fit_text(values.get("legal_code", ""), 3))
        set_field(detail, 196, 2, fit_text(values.get("beneficiary_country", ""), 2))
        set_field(detail, 205, 3, "000")
        set_field(detail, 208, 3, "000")
        detail_text = "".join(detail)
        lines.append(detail_text)
        record_sum += sequence
        account_sum += checksum_text_number(account, 24)
        amount_sum += int(amount)

    trailer = [" "] * 256
    set_field(trailer, 1, 2, "63")
    set_field(trailer, 3, 6, checksum_number(record_sum, 6))
    set_field(trailer, 9, 24, checksum_number(account_sum, 24))
    set_field(trailer, 167, 15, checksum_number(amount_sum, 15))
    lines.append("".join(trailer))
    return "\r\n".join(lines) + "\r\n"


def build_unicredit_ccy_text(records: list[dict[str, str]], config: dict[str, Any]) -> str:
    if not records:
        raise ValueError("Nem találtam konvertálható adatsort.")
    first = records[0]
    creation_date = parse_date(config.get("identifier_date") or date.today().strftime("%Y-%m-%d"), "létrehozás dátuma")
    sender_account = fixed_local_account(first.get("sender_account", ""))
    sender_currency = fit_text(first.get("sender_currency") or "HUF", 3)

    header = [" "] * 800
    set_field(header, 1, 2, "34")
    set_field(header, 3, 24, sender_account)
    set_field(header, 27, 3, sender_currency)
    set_field(header, 30, 1, "1")
    set_field(header, 31, 1, "0")
    set_field(header, 32, 8, creation_date)

    lines = ["".join(header)]
    record_sum = 0
    amount_sum = 0
    for sequence, values in enumerate(records, start=1):
        if fixed_local_account(values.get("sender_account", "")) != sender_account:
            raise ValueError("UniCredit CCY fájlban egy csomagon belül csak egy terhelendő számlaszám szerepelhet.")
        if fit_text(values.get("sender_currency") or "HUF", 3) != sender_currency:
            raise ValueError("UniCredit CCY fájlban egy csomagon belül csak egy terhelendő számladeviza szerepelhet.")
        currency = clean_cell(values.get("currency") or "EUR").upper()
        decimals = 0 if currency in {"HUF", "JPY"} else 2
        amount = amount_to_fixed(values.get("amount", ""), 0, 13) + "00" if decimals == 0 else amount_to_fixed(values.get("amount", ""), 2, 15)

        note = fit_text(values.get("note", ""), 140)
        detail = [" "] * 800
        set_field(detail, 1, 2, "54")
        set_field(detail, 3, 6, checksum_number(sequence, 6))
        set_field(detail, 9, 11, fit_text(values.get("beneficiary_bank_swift", ""), 11))
        set_field(detail, 20, 33, fit_text(values.get("beneficiary_bank_id", ""), 33))
        set_field(detail, 53, 35, fit_text(values.get("beneficiary_bank_name", ""), 35))
        set_field(detail, 88, 35, fit_text(values.get("beneficiary_bank_address_1", ""), 35))
        set_field(detail, 123, 35, fit_text(values.get("beneficiary_bank_address_2", ""), 35))
        set_field(detail, 158, 35, fit_text(values.get("beneficiary_bank_address_3", ""), 35))
        set_field(detail, 193, 11, fit_text(values.get("correspondent_bank_swift", ""), 11))
        set_field(detail, 204, 33, fit_text(values.get("correspondent_bank_id", ""), 33))
        set_field(detail, 237, 35, fit_text(values.get("correspondent_bank_name", ""), 35))
        set_field(detail, 272, 35, fit_text(values.get("correspondent_bank_address_1", ""), 35))
        set_field(detail, 307, 35, fit_text(values.get("correspondent_bank_address_2", ""), 35))
        set_field(detail, 342, 35, fit_text(values.get("correspondent_bank_address_3", ""), 35))
        set_field(detail, 377, 34, fit_text(values.get("beneficiary_account", ""), 34))
        set_field(detail, 411, 35, fit_text(values.get("beneficiary_name", ""), 35))
        set_field(detail, 446, 35, fit_text(values.get("beneficiary_address_1", ""), 35))
        set_field(detail, 481, 35, fit_text(values.get("beneficiary_address_2", ""), 35))
        set_field(detail, 516, 35, fit_text(values.get("beneficiary_address_3", ""), 35))
        set_field(detail, 551, 35, note[0:35])
        set_field(detail, 586, 35, note[35:70])
        set_field(detail, 621, 35, note[70:105])
        set_field(detail, 656, 35, note[105:140])
        set_field(detail, 691, 3, fit_text(values.get("cost_bearer") or "SHA", 3))
        set_field(detail, 694, 3, fit_text(values.get("payout_currency") or currency, 3))
        set_field(detail, 697, 15, amount)
        set_field(detail, 712, 3, fit_text(currency, 3))
        set_field(detail, 715, 8, optional_date_or_zero(values.get("value_date", "")))
        set_field(detail, 723, 1, yn_flag(values.get("urgent_execution", ""), "N"))
        set_field(detail, 724, 4, "0000")
        set_field(detail, 728, 1, yn_flag(values.get("hold_flag", ""), "N"))
        set_field(detail, 729, 1, yn_flag(values.get("chqb_flag", ""), "N"))
        set_field(detail, 730, 1, yn_flag(values.get("deal_ticket_flag", ""), "N"))
        set_field(detail, 731, 8, optional_date_or_zero(values.get("deal_ticket_date", "")))
        set_field(detail, 739, 6, fit_digits(values.get("deal_ticket_sequence", ""), 6, pad_left=True))
        set_field(detail, 745, 2, fit_text(values.get("beneficiary_country", ""), 2))
        set_field(detail, 747, 3, fit_text(values.get("legal_code", ""), 3))
        set_field(detail, 750, 3, "000")
        set_field(detail, 753, 3, "000")
        lines.append("".join(detail))
        record_sum += sequence
        amount_sum += int(amount)

    trailer = [" "] * 800
    set_field(trailer, 1, 2, "74")
    set_field(trailer, 3, 6, checksum_number(record_sum, 6))
    set_field(trailer, 697, 15, checksum_number(amount_sum, 15))
    lines.append("".join(trailer))
    return "\r\n".join(lines) + "\r\n"


def get_format(format_id: str | None) -> dict[str, Any]:
    if format_id and format_id in FORMATS:
        return FORMATS[format_id]
    return FORMATS["erste_huf_payord"]


def fields_for_format(format_id: str | None) -> dict[str, dict[str, Any]]:
    return get_format(format_id).get("fields", FIELDS)


def validate_required(values: dict[str, str], row_number: int, fields: dict[str, dict[str, Any]]) -> None:
    missing = [info["label"] for key, info in fields.items() if info["required"] and not clean_cell(values.get(key, ""))]
    if missing:
        raise ValueError(f"{row_number}. sor: hiányzó kötelező mező: {', '.join(missing)}")


def convert_records(data: bytes, filename: str, config: dict[str, Any]) -> tuple[str, int]:
    header_row = 1
    format_id = config.get("format") or "erste_huf_payord"
    format_def = get_format(format_id)
    if format_def.get("unsupported"):
        raise ValueError(format_def["unsupported"])
    fields = fields_for_format(format_id)
    sheets, selected, headers, rows = workbook_rows(data, filename)
    if not headers:
        raise ValueError("Nem található fejlécsor.")
    output: list[str] = []
    structured_records: list[dict[str, str]] = []
    sequence = 0
    errors: list[str] = []
    for offset, raw_row in enumerate(rows, start=header_row + 1):
        row = list(raw_row)
        if row_is_empty(row):
            continue
        data_row = row_dict(headers, row)
        values = {key: value_for(data_row, key, config) for key in fields}
        try:
            validate_required(values, offset, fields)
            sequence += 1
            builder = get_format(format_id).get("builder")
            if format_id in {"unicredit_huf_pay", "unicredit_fx_ccy"} or builder in {"pain_huf", "pain_fx", "pain_sepa"}:
                structured_records.append(values)
                continue
            if format_id == "erste_fx_payord":
                line = build_fx_payord_line(values, sequence, config)
            elif format_id == "erste_sepa_payord":
                line = build_sepa_payord_line(values, sequence, config)
            elif format_id == "kh_fx_payord":
                line = build_kh_fx_payord_line(values, sequence, config)
            elif format_id == "otp_huf":
                line = build_otp_huf_payord_line(values, sequence, config)
            elif format_id == "otp_fx":
                line = build_otp_fx_payord_line(values, sequence, config)
            elif format_id == "raiffeisen_huf":
                line = build_raiffeisen_huf_line(values, sequence, config)
            elif format_id in {"raiffeisen_fx", "raiffeisen_sepa"}:
                line = build_raiffeisen_fx_line(values, sequence, config)
            else:
                line = build_payord_line(values, sequence, config)
            output.append(line)
        except ValueError as exc:
            errors.append(f"{offset}. sor: {exc}")
            if len(errors) >= 12:
                break
    if errors:
        raise ValueError("\n".join(errors))
    if format_id == "unicredit_huf_pay":
        return build_unicredit_pay_text(structured_records, config), len(structured_records)
    if format_id == "unicredit_fx_ccy":
        return build_unicredit_ccy_text(structured_records, config), len(structured_records)
    builder = get_format(format_id).get("builder")
    if builder == "pain_sepa":
        return build_pain_xml(structured_records, config, service_level="SEPA"), len(structured_records)
    if builder in {"pain_huf", "pain_fx"}:
        return build_pain_xml(structured_records, config), len(structured_records)
    if not output:
        raise ValueError("Nem találtam konvertálható adatsort.")
    return "\r\n".join(output) + "\r\n", len(output)


def build_template_xlsx(format_id: str | None = None) -> bytes:
    format_def = get_format(format_id)
    fields = fields_for_format(format_id)
    template_example = format_def.get("template_example") or TEMPLATE_EXAMPLE
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Utalasok"

    headers = [info["label"] for info in fields.values()]
    example = [template_example.get(key, "") for key in fields]
    sheet.append(headers)
    sheet.append(example)
    sheet.freeze_panes = "A2"

    header_fill = PatternFill("solid", fgColor="B51F2B")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D7DDE5")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for cell in sheet[2]:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border
    for index, header in enumerate(headers, start=1):
        width = max(14, min(32, len(header) + 4))
        sheet.column_dimensions[openpyxl.utils.get_column_letter(index)].width = width
    sheet.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}2"

    info_sheet = workbook.create_sheet("Mezo_info")
    info_sheet.append(["Mező", "Kötelező", "Megjegyzés"])
    for key, info in fields.items():
        note = info.get("help", "")
        if key == "status":
            note = "Üres = NORMAL, V = VIBER, T = TPLUSONE."
        elif key == "amount":
            note = "Számként vagy szövegként is megadható. A TXT-ben 13 karakterre, balról nullázva kerül."
        elif key == "value_date":
            note = "Dátumként vagy ÉÉÉÉHHNN formában add meg."
        elif key == "identifier":
            note = "Ha üres, a konverter automatikusan generálja."
        info_sheet.append([info["label"], "igen" if info["required"] else "nem", note])
    for cell in info_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    for row in info_sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
    info_sheet.column_dimensions["A"].width = 28
    info_sheet.column_dimensions["B"].width = 12
    info_sheet.column_dimensions["C"].width = 78
    info_sheet.freeze_panes = "A2"

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


HU_WEIGHTS = [9, 7, 3, 1]
IBAN_LENGTHS = {
    "AD": 24, "AE": 23, "AL": 28, "AT": 20, "AZ": 28, "BA": 20, "BE": 16, "BG": 22,
    "BH": 22, "BR": 29, "BY": 28, "CH": 21, "CR": 22, "CY": 28, "CZ": 24, "DE": 22,
    "DK": 18, "DO": 28, "EE": 20, "EG": 29, "ES": 24, "FI": 18, "FO": 18, "FR": 27,
    "GB": 22, "GE": 22, "GI": 23, "GL": 18, "GR": 27, "GT": 28, "HR": 21, "HU": 28,
    "IE": 22, "IL": 23, "IQ": 23, "IS": 26, "IT": 27, "JO": 30, "KW": 30, "KZ": 20,
    "LB": 28, "LC": 32, "LI": 21, "LT": 20, "LU": 20, "LV": 21, "MC": 27, "MD": 24,
    "ME": 22, "MK": 19, "MR": 27, "MT": 31, "MU": 30, "NL": 18, "NO": 15, "PK": 24,
    "PL": 28, "PS": 29, "PT": 25, "QA": 29, "RO": 24, "RS": 22, "SA": 24, "SC": 31,
    "SE": 24, "SI": 19, "SK": 24, "SM": 27, "ST": 25, "SV": 28, "TL": 23, "TN": 24,
    "TR": 26, "UA": 29, "VA": 22, "VG": 24, "XK": 20,
}


def default_bank_registry() -> dict[str, Any]:
    return {
        "source_url": "https://www.mnb.hu/letoltes/sht.xlsx",
        "name_column": "C",
        "prefix_column": "A",
        "updated_at": "",
        "rows": [],
        "detected_name_column": "",
        "detected_prefix_column": "",
        "startup_ok": None,
        "startup_message": "",
    }


def load_json_file(path: Path, default: Any) -> Any:
    if not path.exists():
        return default
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return default


def write_json_file(path: Path, data: Any) -> Any:
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    return data


def load_accounts_file() -> list[dict[str, Any]]:
    data = load_json_file(ACCOUNTS_FILE, [])
    if isinstance(data, dict):
        rows: list[dict[str, Any]] = []
        for company_id, accounts in (data.get("by_company") or {}).items():
            if isinstance(accounts, list):
                for row in accounts:
                    if isinstance(row, dict):
                        rows.append({**row, "company_id": clean_cell(row.get("company_id") or company_id or "default")})
        data = rows
    if not isinstance(data, list):
        return []
    for row in data:
        if isinstance(row, dict):
            row["currency"] = clean_cell(row.get("currency") or "HUF").upper()
            row["company_id"] = clean_cell(row.get("company_id") or "default")
    return data


def save_accounts_file(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return write_json_file(ACCOUNTS_FILE, rows)


def default_company() -> dict[str, str]:
    return {"id": "default", "name": "Alap cég"}


def load_companies_file() -> list[dict[str, Any]]:
    data = load_json_file(COMPANIES_FILE, [])
    if not isinstance(data, list) or not data:
        return [default_company()]
    normalized: list[dict[str, Any]] = []
    seen: set[str] = set()
    for row in data:
        if not isinstance(row, dict):
            continue
        company_id = clean_cell(row.get("id")) or uuid.uuid4().hex
        name = clean_cell(row.get("name")) or "Névtelen cég"
        if company_id in seen:
            continue
        normalized.append({"id": company_id, "name": name})
        seen.add(company_id)
    if "default" not in seen:
        normalized.insert(0, default_company())
    return normalized


def save_companies_file(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return write_json_file(COMPANIES_FILE, rows)


def ensure_company_id(company_id: str | None = None) -> str:
    requested = clean_cell(company_id or "default")
    companies = load_companies_file()
    if any(company.get("id") == requested for company in companies):
        return requested
    return companies[0]["id"] if companies else "default"


def accounts_for_company(company_id: str | None) -> list[dict[str, Any]]:
    cid = ensure_company_id(company_id)
    return [row for row in load_accounts_file() if clean_cell(row.get("company_id") or "default") == cid]


def load_partners_file() -> list[dict[str, Any]]:
    data = load_json_file(PARTNERS_FILE, [])
    if isinstance(data, dict):
        rows: list[dict[str, Any]] = []
        for company_id, partners in (data.get("by_company") or {}).items():
            if isinstance(partners, list):
                for row in partners:
                    if isinstance(row, dict):
                        rows.append({**row, "company_id": clean_cell(row.get("company_id") or company_id or "default")})
        data = rows
    if not isinstance(data, list):
        return []
    rows = [row for row in data if isinstance(row, dict)]
    for row in rows:
        row["company_id"] = clean_cell(row.get("company_id") or "default")
    return rows


def save_partners_file(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return write_json_file(PARTNERS_FILE, rows)


def partners_for_company(company_id: str | None) -> list[dict[str, Any]]:
    cid = ensure_company_id(company_id)
    return [row for row in load_partners_file() if clean_cell(row.get("company_id") or "default") == cid]


def load_bank_registry() -> dict[str, Any]:
    data = load_json_file(BANK_REGISTRY_FILE, default_bank_registry())
    base = default_bank_registry()
    if isinstance(data, dict):
        base.update(data)
    if not isinstance(base.get("rows"), list):
        base["rows"] = []
    return base


def save_bank_registry(data: dict[str, Any]) -> dict[str, Any]:
    base = load_bank_registry()
    for key in ("source_url", "name_column", "prefix_column"):
        if key in data:
            base[key] = clean_cell(data.get(key))
    return write_json_file(BANK_REGISTRY_FILE, base)


def strip_accents(value: str) -> str:
    text = unicodedata.normalize("NFD", str(value or ""))
    return "".join(ch for ch in text if unicodedata.category(ch) != "Mn")


def loose_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", strip_accents(value).lower())


def hu_check_digit(payload: str) -> int:
    total = sum(int(ch) * HU_WEIGHTS[index % len(HU_WEIGHTS)] for index, ch in enumerate(payload))
    return (10 - (total % 10)) % 10


def format_hu_account_digits(digits: str) -> str:
    padded = digits if len(digits) == 24 else digits + "0" * 8
    return "-".join(padded[index:index + 8] for index in range(0, 24, 8))


def registry_lookup(prefix: str, registry: dict[str, Any] | None = None) -> dict[str, str] | None:
    registry = registry or load_bank_registry()
    for row in registry.get("rows", []):
        if clean_cell(row.get("prefix")) == prefix:
            return row
    return None


def validate_hu_account(value: str, registry: dict[str, Any] | None = None) -> dict[str, Any]:
    digits = digits_only(value)
    errors: list[str] = []
    if len(digits) not in (16, 24):
        errors.append("A magyar bankszámlaszám 16 vagy 24 számjegy lehet.")
        return {"ok": False, "errors": errors, "digits": digits, "formatted": "", "bank": None}
    if hu_check_digit(digits[:7]) != int(digits[7]):
        errors.append("Az első 8 számjegy ellenőrzőszáma hibás.")
    second_payload = digits[8:-1]
    if hu_check_digit(second_payload) != int(digits[-1]):
        errors.append("A számlaszám 2-3. nyolcas csoportjának ellenőrzőszáma hibás.")
    bank = registry_lookup(digits[:8], registry)
    if not bank:
        errors.append("Az első 8 számjegy nincs a betöltött hitelesítő táblában.")
    return {
        "ok": not errors,
        "errors": errors,
        "digits": digits,
        "normalized": digits if len(digits) == 24 else digits + "0" * 8,
        "formatted": format_hu_account_digits(digits),
        "prefix": digits[:8],
        "bank": bank,
    }


def normalize_iban(value: str) -> str:
    return re.sub(r"[\s-]+", "", clean_cell(value)).upper()


def iban_mod97_ok(iban: str) -> bool:
    remainder = 0
    for ch in iban[4:] + iban[:4]:
        if ch.isdigit():
            expanded = ch
        elif "A" <= ch <= "Z":
            expanded = str(ord(ch) - 55)
        else:
            return False
        for digit in expanded:
            remainder = (remainder * 10 + int(digit)) % 97
    return remainder == 1


def validate_iban(value: str, registry: dict[str, Any] | None = None) -> dict[str, Any]:
    iban = normalize_iban(value)
    errors: list[str] = []
    if not re.fullmatch(r"[A-Z]{2}[0-9]{2}[A-Z0-9]{1,30}", iban):
        errors.append("Az IBAN formátuma hibás.")
    expected_length = IBAN_LENGTHS.get(iban[:2])
    if expected_length and len(iban) != expected_length:
        errors.append(f"{iban[:2]} IBAN hossza {expected_length} karakter kell legyen.")
    elif len(iban) > 34:
        errors.append("Az IBAN legfeljebb 34 karakter lehet.")
    if not errors and not iban_mod97_ok(iban):
        errors.append("Az IBAN MOD 97-10 ellenőrzése hibás.")
    domestic = None
    if not errors and iban.startswith("HU"):
        domestic = validate_hu_account(iban[4:], registry)
        if not domestic["ok"]:
            errors.extend(f"HU IBAN belföldi rész: {msg}" for msg in domestic["errors"])
    elif not errors:
        errors.append("Külföldi IBAN-t egyelőre nem kezelünk ebben a listában.")
    return {"ok": not errors, "errors": errors, "iban": iban, "domestic": domestic}


def validate_iban_syntax(value: str) -> dict[str, Any]:
    iban = normalize_iban(value)
    errors: list[str] = []
    if not re.fullmatch(r"[A-Z]{2}[0-9]{2}[A-Z0-9]{1,30}", iban):
        errors.append("Az IBAN formátuma hibás.")
    expected_length = IBAN_LENGTHS.get(iban[:2])
    if expected_length and len(iban) != expected_length:
        errors.append(f"{iban[:2]} IBAN hossza {expected_length} karakter kell legyen.")
    elif len(iban) > 34:
        errors.append("Az IBAN legfeljebb 34 karakter lehet.")
    if not errors and not iban_mod97_ok(iban):
        errors.append("Az IBAN MOD 97-10 ellenőrzése hibás.")
    return {"ok": not errors, "errors": errors, "iban": iban}


def validate_account_input(value: str, registry: dict[str, Any] | None = None) -> dict[str, Any]:
    raw = clean_cell(value)
    if re.match(r"^[A-Za-z]{2}", raw.replace(" ", "").replace("-", "")):
        iban_result = validate_iban(raw, registry)
        domestic = iban_result.get("domestic") or {}
        return {
            "ok": iban_result["ok"],
            "errors": iban_result["errors"],
            "input_type": "IBAN",
            "iban": iban_result["iban"],
            "digits": domestic.get("digits", ""),
            "normalized": domestic.get("normalized", ""),
            "formatted": domestic.get("formatted", ""),
            "prefix": domestic.get("prefix", ""),
            "bank": domestic.get("bank"),
        }
    domestic_result = validate_hu_account(raw, registry)
    domestic_result["input_type"] = "HU"
    return domestic_result


def selector_to_index(selector: str, headers: list[str]) -> int | None:
    raw = clean_cell(selector)
    if not raw:
        return None
    if re.fullmatch(r"[A-Za-z]+", raw):
        try:
            return openpyxl.utils.column_index_from_string(raw.upper()) - 1
        except ValueError:
            pass
    if raw.isdigit() and int(raw) > 0:
        return int(raw) - 1
    target = loose_key(raw)
    for index, header in enumerate(headers):
        if loose_key(header) == target:
            return index
    return None


def auto_detect_registry_columns(rows: list[list[Any]]) -> tuple[int, int]:
    scan_rows = rows[:120]
    width = max((len(row) for row in scan_rows), default=0)
    prefix_scores: list[tuple[int, int]] = []
    name_scores: list[tuple[int, int]] = []
    for col in range(width):
        prefix_count = 0
        name_count = 0
        for row in scan_rows[1:]:
            value = clean_cell(row[col] if col < len(row) else "")
            digits = digits_only(value)
            if len(digits) >= 8:
                prefix_count += 1
            if value and len(digits) < max(3, len(value) // 2) and len(value) > 3:
                name_count += 1
        prefix_scores.append((prefix_count, col))
        name_scores.append((name_count, col))
    prefix_col = max(prefix_scores, default=(0, 0))[1]
    name_candidates = [pair for pair in name_scores if pair[1] != prefix_col]
    name_col = max(name_candidates, default=(0, 1 if width > 1 else 0))[1]
    return name_col, prefix_col


def parse_bank_registry(data: bytes, config: dict[str, Any]) -> dict[str, Any]:
    workbook = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    ws = workbook[workbook.sheetnames[0]]
    rows = [list(row) for row in ws.iter_rows(values_only=True)]
    if not rows:
        raise ValueError("A hitelesítő tábla üres.")
    headers = [clean_cell(v) for v in rows[0]]
    detected_name, detected_prefix = auto_detect_registry_columns(rows)
    header_keys = [loose_key(header) for header in headers]
    for index, key in enumerate(header_keys):
        if key in {"branchofficecode", "branchcode"}:
            detected_prefix = index
        if key in {"nameofthebranchoffice", "branchname", "bankname", "name"}:
            detected_name = index
    name_col = selector_to_index(config.get("name_column", ""), headers)
    prefix_col = selector_to_index(config.get("prefix_column", ""), headers)
    if name_col is None:
        name_col = detected_name
    if prefix_col is None:
        prefix_col = detected_prefix
    parsed: dict[str, dict[str, str]] = {}
    for row in rows[1:]:
        name = clean_cell(row[name_col] if name_col < len(row) else "")
        prefix = digits_only(row[prefix_col] if prefix_col < len(row) else "")[:8]
        if len(prefix) == 8 and name:
            parsed[prefix] = {"prefix": prefix, "bank_name": name, "bank_country": "HU"}
    if not parsed:
        raise ValueError("Nem találtam 8 számjegyű prefixet és banknevet a megadott oszlopokkal.")
    return {
        **config,
        "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "detected_name_column": openpyxl.utils.get_column_letter(name_col + 1),
        "detected_prefix_column": openpyxl.utils.get_column_letter(prefix_col + 1),
        "rows": sorted(parsed.values(), key=lambda item: (item["bank_name"], item["prefix"])),
    }


def refresh_bank_registry() -> dict[str, Any]:
    config = load_bank_registry()
    url = clean_cell(config.get("source_url")) or default_bank_registry()["source_url"]
    request = urllib.request.Request(url, headers={"User-Agent": "Banki TXT konverter"})
    with urllib.request.urlopen(request, timeout=30) as response:
        data = response.read()
    parsed = parse_bank_registry(data, config)
    return write_json_file(BANK_REGISTRY_FILE, parsed)


def refresh_bank_registry_at_startup() -> dict[str, Any]:
    previous = load_bank_registry()
    try:
        registry = refresh_bank_registry()
        count = len(registry.get("rows", []))
        registry["startup_ok"] = True
        registry["startup_message"] = f"Hitelesítő tábla indításkor frissítve: {count} prefix, időpont: {registry.get('updated_at', '')}."
        return write_json_file(BANK_REGISTRY_FILE, registry)
    except Exception as exc:
        count = len(previous.get("rows", []))
        previous["startup_ok"] = False
        previous["startup_message"] = f"Hitelesítő tábla nem frissült indításkor, a korábbi helyi tábla maradt használatban ({count} prefix). Hiba: {exc}"
        return write_json_file(BANK_REGISTRY_FILE, previous)


def account_payload_from_body(body: dict[str, Any], registry: dict[str, Any]) -> dict[str, Any]:
    country = clean_cell(body.get("bank_country") or "HU").upper()
    if country != "HU":
        raise ValueError("Külföldi bankszámlákat egyelőre nem kezelünk.")
    currency = clean_cell(body.get("currency")).upper()
    if not currency:
        raise ValueError("A deviza megadása kötelező.")
    if currency not in CURRENCIES:
        raise ValueError(f"Ismeretlen deviza: {currency}")
    validation = validate_account_input(clean_cell(body.get("account_number")), registry)
    if not validation["ok"]:
        raise ValueError("\n".join(validation["errors"]))
    registry_bank = validation.get("bank") or {}
    return {
        "id": clean_cell(body.get("id")) or uuid.uuid4().hex,
        "company_id": ensure_company_id(clean_cell(body.get("company_id"))),
        "bank_country": "HU",
        "bank_name": clean_cell(body.get("bank_name")) or clean_cell(registry_bank.get("bank_name")),
        "currency": currency,
        "account_number": validation["formatted"],
        "account_digits": validation["normalized"],
        "prefix": validation.get("prefix", ""),
        "iban": validation.get("iban", ""),
        "input_type": validation.get("input_type", "HU"),
        "valid": True,
        "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


def account_identity(account: dict[str, Any]) -> str:
    stored = clean_cell(account.get("account_digits"))
    if stored:
        return stored
    digits = digits_only(account.get("account_number", ""))
    if len(digits) == 16:
        return digits + "0" * 8
    return digits


def find_duplicate_account(accounts: list[dict[str, Any]], payload: dict[str, Any]) -> dict[str, Any] | None:
    payload_id = clean_cell(payload.get("id"))
    payload_digits = account_identity(payload)
    payload_company = clean_cell(payload.get("company_id") or "default")
    if not payload_digits:
        return None
    for account in accounts:
        if clean_cell(account.get("id")) == payload_id:
            continue
        if clean_cell(account.get("company_id") or "default") != payload_company:
            continue
        if account_identity(account) == payload_digits:
            return account
    return None


def cleanup_saved_accounts(registry: dict[str, Any] | None = None) -> list[dict[str, Any]]:
    registry = registry or load_bank_registry()
    accounts = load_accounts_file()
    kept: list[dict[str, Any]] = []
    removed: list[dict[str, Any]] = []
    for account in accounts:
        raw_value = clean_cell(account.get("iban")) or clean_cell(account.get("account_number")) or clean_cell(account.get("account_digits"))
        validation = validate_account_input(raw_value, registry)
        if validation["ok"]:
            kept.append(account)
        else:
            removed.append({**account, "validation_errors": validation["errors"]})
    if removed:
        save_accounts_file(kept)
    return removed


def read_json_body(handler: BaseHTTPRequestHandler) -> dict[str, Any]:
    length = int(handler.headers.get("Content-Length", "0") or "0")
    payload = handler.rfile.read(length).decode("utf-8") if length else "{}"
    data = json.loads(payload or "{}")
    if not isinstance(data, dict):
        raise ValueError("Hibás kérés.")
    return data


def import_accounts_from_upload(data: bytes, filename: str, company_id: str | None = None) -> dict[str, Any]:
    _, _, headers, rows = workbook_rows(data, filename)
    if not headers:
        raise ValueError("Nem található fejlécsor.")
    header_map = {loose_key(header): header for header in headers}
    def pick(*names: str) -> str:
        for name in names:
            found = header_map.get(loose_key(name))
            if found:
                return found
        return ""
    country_col = pick("Bank országa", "Ország", "Country")
    bank_col = pick("Bank neve", "Bank", "Bank name")
    account_col = pick("Számlaszám", "Bankszámlaszám", "Account", "IBAN")
    currency_col = pick("Deviza", "Pénznem", "Currency")
    if not account_col:
        raise ValueError("Az import fájlban kell egy Számlaszám vagy IBAN oszlop.")
    if not currency_col:
        raise ValueError("Az import fájlban kell egy Deviza oszlop.")
    registry = load_bank_registry()
    existing = load_accounts_file()
    by_id = {row.get("id"): row for row in existing}
    cid = ensure_company_id(company_id)
    by_digits = {account_identity(row): row for row in existing if account_identity(row) and clean_cell(row.get("company_id") or "default") == cid}
    added = 0
    errors: list[str] = []
    for line_no, raw_row in enumerate(rows, start=2):
        if row_is_empty(list(raw_row)):
            continue
        row = row_dict(headers, list(raw_row))
        try:
            payload = account_payload_from_body({
                "bank_country": row.get(country_col) or "HU",
                "company_id": cid,
                "bank_name": row.get(bank_col) or "",
                "currency": row.get(currency_col) or "",
                "account_number": row.get(account_col) or "",
            }, registry)
            duplicate = by_digits.get(account_identity(payload))
            if duplicate:
                errors.append(f"{line_no}. sor: ez a bankszámla már szerepel: {payload['account_number']}")
                continue
            by_id[payload["id"]] = payload
            by_digits[account_identity(payload)] = payload
            added += 1
        except ValueError as exc:
            errors.append(f"{line_no}. sor: {str(exc).splitlines()[0]}")
    save_accounts_file(list(by_id.values()))
    all_rows = list(by_id.values())
    return {"added": added, "errors": errors, "accounts": [row for row in all_rows if clean_cell(row.get("company_id") or "default") == cid]}


def build_accounts_template_xlsx() -> bytes:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Bankszamlak"
    headers = ["Bank országa", "Bank neve", "Számlaszám", "Deviza"]
    sheet.append(headers)
    sheet.append(["HU", "Minta Bank", "11111111-22222222-33333333", "HUF"])
    header_fill = PatternFill("solid", fgColor="263449")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    sheet.column_dimensions["A"].width = 16
    sheet.column_dimensions["B"].width = 34
    sheet.column_dimensions["C"].width = 32
    sheet.column_dimensions["D"].width = 12
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def normalize_bic(value: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", clean_cell(value).upper())


def validate_bic(value: str) -> bool:
    bic = normalize_bic(value)
    return bool(re.fullmatch(r"[A-Z]{4}[A-Z]{2}[A-Z0-9]{2}([A-Z0-9]{3})?", bic))


def lookup_bic_online(bic: str = "", iban: str = "") -> dict[str, Any]:
    iban = normalize_iban(iban)
    bic = normalize_bic(bic)
    if iban:
        url = f"https://openiban.com/validate/{iban}?getBIC=true&validateBankCode=true"
        try:
            request = urllib.request.Request(url, headers={"User-Agent": "Banki TXT konverter"})
            with urllib.request.urlopen(request, timeout=10) as response:
                data = json.loads(response.read().decode("utf-8"))
            bank = data.get("bankData") or {}
            if bank:
                city = clean_cell(bank.get("city"))
                zip_code = clean_cell(bank.get("zip"))
                return {
                    "found": True,
                    "bic": clean_cell(bank.get("bic")) or bic,
                    "bank_name": clean_cell(bank.get("name")),
                    "bank_address": " ".join(part for part in [zip_code, city] if part),
                    "source": "openiban",
                    "source_message": "OpenIBAN alapján kitöltve. Az OpenIBAN BIC adat csak bizonyos országokra érhető el.",
                }
        except Exception:
            pass
    if bic:
        return {"found": False, "error": "A teljes BIC név/cím adatbázis hivatalosan SWIFTRef/licences forrás. Ehhez most nincs beállított API-kulcs; add meg kézzel, vagy IBAN alapján próbáld."}
    return {"found": False, "error": "Adj meg IBAN-t vagy SWIFT/BIC kódot."}


def partner_payload_from_body(body: dict[str, Any], registry: dict[str, Any]) -> dict[str, Any]:
    name = clean_cell(body.get("name"))
    if not name:
        raise ValueError("A partner neve kötelező.")
    account_number = clean_cell(body.get("account_number"))
    iban = clean_cell(body.get("iban"))
    swift = normalize_bic(body.get("swift_bic", ""))
    country = clean_cell(body.get("country") or "HU").upper()
    bank_name = clean_cell(body.get("bank_name"))
    bank_address = clean_cell(body.get("bank_address"))
    normalized_account = ""
    if account_number:
        validation = validate_hu_account(account_number, registry)
        if not validation["ok"]:
            raise ValueError("\n".join(validation["errors"]))
        normalized_account = validation["normalized"]
        account_number = validation["formatted"]
        bank = validation.get("bank") or {}
        bank_name = bank_name or clean_cell(bank.get("bank_name"))
        country = "HU"
    if iban:
        iban_result = validate_iban_syntax(iban)
        if not iban_result["ok"]:
            raise ValueError("\n".join(iban_result["errors"]))
        iban = iban_result["iban"]
        if iban.startswith("HU") and not account_number:
            validation = validate_hu_account(iban[4:], registry)
            if validation["ok"]:
                normalized_account = validation["normalized"]
                account_number = validation["formatted"]
                bank = validation.get("bank") or {}
                bank_name = bank_name or clean_cell(bank.get("bank_name"))
    if swift and not validate_bic(swift):
        raise ValueError("A SWIFT/BIC 8 vagy 11 karakteres ISO 9362 formátum legyen.")
    if not account_number and not iban:
        raise ValueError("Adj meg magyar számlaszámot vagy IBAN-t.")
    return {
        "id": clean_cell(body.get("id")) or uuid.uuid4().hex,
        "company_id": ensure_company_id(clean_cell(body.get("company_id"))),
        "partner_code": clean_cell(body.get("partner_code")),
        "name": name,
        "account_number": account_number,
        "account_digits": normalized_account,
        "iban": iban,
        "swift_bic": swift,
        "country": country,
        "address": clean_cell(body.get("address")),
        "bank_name": bank_name,
        "bank_address": bank_address,
        "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


def partner_identity(partner: dict[str, Any]) -> str:
    return clean_cell(partner.get("partner_code")) or clean_cell(partner.get("iban")) or clean_cell(partner.get("account_digits")) or clean_cell(partner.get("account_number"))


def import_partners_from_upload(data: bytes, filename: str, company_id: str | None = None) -> dict[str, Any]:
    _, _, headers, rows = workbook_rows(data, filename)
    if not headers:
        raise ValueError("Nem található fejlécsor.")
    header_map = {loose_key(header): header for header in headers}
    def pick(*names: str) -> str:
        for name in names:
            found = header_map.get(loose_key(name))
            if found:
                return found
        return ""
    cols = {
        "partner_code": pick("Partner kód", "Partner kod", "Code"),
        "name": pick("Név", "Nev", "Partner neve", "Name"),
        "account_number": pick("Számlaszám", "Szamlaszam", "Magyar számlaszám"),
        "iban": pick("IBAN"),
        "swift_bic": pick("SWIFT/BIC", "SWIFT", "BIC"),
        "country": pick("Ország", "Orszag", "Country"),
        "address": pick("Partner címe", "Partner cime", "Cím", "Address"),
        "bank_name": pick("Bank neve", "Bank name"),
        "bank_address": pick("Bank címe", "Bank cime", "Bank address"),
    }
    registry = load_bank_registry()
    existing = load_partners_file()
    cid = ensure_company_id(company_id)
    by_id = {row.get("id"): row for row in existing}
    identities = {partner_identity(row): row for row in existing if partner_identity(row) and clean_cell(row.get("company_id") or "default") == cid}
    added = 0
    errors: list[str] = []
    for line_no, raw_row in enumerate(rows, start=2):
        if row_is_empty(list(raw_row)):
            continue
        row = row_dict(headers, list(raw_row))
        try:
            payload = partner_payload_from_body({**{key: row.get(col, "") if col else "" for key, col in cols.items()}, "company_id": cid}, registry)
            ident = partner_identity(payload)
            if ident and ident in identities:
                errors.append(f"{line_no}. sor: ez a partner már szerepel: {ident}")
                continue
            by_id[payload["id"]] = payload
            identities[ident] = payload
            added += 1
        except ValueError as exc:
            errors.append(f"{line_no}. sor: {str(exc).splitlines()[0]}")
    all_rows = list(by_id.values())
    save_partners_file(all_rows)
    return {"added": added, "errors": errors, "partners": [row for row in all_rows if clean_cell(row.get("company_id") or "default") == cid]}


def build_partners_template_xlsx() -> bytes:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Partnerek"
    headers = ["Partner kód", "Név", "Számlaszám", "IBAN", "SWIFT/BIC", "Ország", "Partner címe", "Bank neve", "Bank címe"]
    sheet.append(headers)
    sheet.append(["P001", "Minta Partner Kft", "44444444-55555555-66666666", "", "", "HU", "Minta cím", "", ""])
    sheet.append(["P002", "Minta Partner GmbH", "", "AT451100000421840000", "BKAUATWW", "AT", "Minta cím", "", ""])
    header_fill = PatternFill("solid", fgColor="263449")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    widths = [16, 32, 32, 34, 16, 10, 34, 34, 34]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(index)].width = width
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def default_settings() -> dict[str, Any]:
    return {
        "active_company_id": "default",
        "active_bank": "erste",
        "active_format": "erste_huf_payord",
        "formats": {},
    }


def load_settings_file() -> dict[str, Any]:
    if not SETTINGS_FILE.exists():
        return default_settings()
    try:
        data = json.loads(SETTINGS_FILE.read_text(encoding="utf-8"))
        base = default_settings()
        base.update(data if isinstance(data, dict) else {})
        if not isinstance(base.get("formats"), dict):
            base["formats"] = {}
        if base.get("active_bank") not in BANKS:
            base["active_bank"] = "erste"
        if base.get("active_format") not in FORMATS:
            base["active_format"] = next((fmt_id for fmt_id, fmt in FORMATS.items() if fmt.get("bank") == base.get("active_bank")), "erste_huf_payord")
        if not any(company.get("id") == base.get("active_company_id") for company in load_companies_file()):
            base["active_company_id"] = "default"
        return base
    except Exception:
        return default_settings()


def save_settings_file(data: dict[str, Any]) -> dict[str, Any]:
    cleaned = default_settings()
    cleaned.update(data if isinstance(data, dict) else {})
    if cleaned.get("active_bank") not in BANKS:
        cleaned["active_bank"] = "erste"
    if cleaned.get("active_format") not in FORMATS:
        cleaned["active_format"] = next((fmt_id for fmt_id, fmt in FORMATS.items() if fmt.get("bank") == cleaned.get("active_bank")), "erste_huf_payord")
    if not any(company.get("id") == cleaned.get("active_company_id") for company in load_companies_file()):
        cleaned["active_company_id"] = ensure_company_id("default")
    if not isinstance(cleaned.get("formats"), dict):
        cleaned["formats"] = {}
    for cfg in cleaned["formats"].values():
        if isinstance(cfg, dict):
            cfg.pop("sheet", None)
            cfg.pop("header_row", None)
    SETTINGS_FILE.write_text(json.dumps(cleaned, ensure_ascii=False, indent=2), encoding="utf-8")
    return cleaned


class Handler(BaseHTTPRequestHandler):
    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        path = parsed.path
        if path in {"/", "/index.html"}:
            body = (
                HTML_PAGE
                .replace("__BANKS_JSON__", json.dumps(BANKS, ensure_ascii=False))
                .replace("__FORMATS_JSON__", json.dumps(FORMATS, ensure_ascii=False))
                .replace("__FIELDS_JSON__", json.dumps(FIELDS, ensure_ascii=False))
                .replace("__CURRENCIES_JSON__", json.dumps(CURRENCIES, ensure_ascii=False))
            )
            text_response(self, body)
            return
        if path == "/api/settings":
            json_response(self, load_settings_file())
            return
        if path == "/api/companies":
            settings = load_settings_file()
            json_response(self, {"companies": load_companies_file(), "active_company_id": settings.get("active_company_id", "default")})
            return
        if path == "/api/accounts":
            params = parse_qs(parsed.query)
            company_id = ensure_company_id((params.get("company_id") or [""])[0])
            removed = cleanup_saved_accounts()
            json_response(self, {"accounts": accounts_for_company(company_id), "removed_invalid": len(removed), "company_id": company_id})
            return
        if path == "/api/partners":
            params = parse_qs(parsed.query)
            company_id = ensure_company_id((params.get("company_id") or [""])[0])
            json_response(self, {"partners": partners_for_company(company_id), "company_id": company_id})
            return
        if path == "/api/bic-lookup":
            params = parse_qs(parsed.query)
            json_response(self, lookup_bic_online((params.get("bic") or [""])[0], (params.get("iban") or [""])[0]))
            return
        if path == "/api/bank-registry":
            registry = load_bank_registry()
            rows = registry.get("rows", [])
            json_response(self, {
                **registry,
                "row_count": len(rows),
                "sample": rows[:8],
            })
            return
        if path == "/template.xlsx":
            params = parse_qs(parsed.query)
            format_id = (params.get("format") or [""])[0]
            data = build_template_xlsx(format_id)
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", "attachment; filename=payord_sablon.xlsx")
            self.send_header("Content-Length", str(len(data)))
            self.end_headers()
            self.wfile.write(data)
            return
        if path == "/template.csv":
            params = parse_qs(parsed.query)
            format_id = (params.get("format") or [""])[0]
            content = ";".join(info["label"] for info in fields_for_format(format_id).values()) + "\r\n"
            data = content.encode("utf-8-sig")
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", "text/csv; charset=utf-8")
            self.send_header("Content-Disposition", "attachment; filename=payord_sablon.csv")
            self.send_header("Content-Length", str(len(data)))
            self.end_headers()
            self.wfile.write(data)
            return
        if path == "/accounts-template.xlsx":
            data = build_accounts_template_xlsx()
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", "attachment; filename=bankszamla_sablon.xlsx")
            self.send_header("Content-Length", str(len(data)))
            self.end_headers()
            self.wfile.write(data)
            return
        if path == "/partners-template.xlsx":
            data = build_partners_template_xlsx()
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", "attachment; filename=partner_sablon.xlsx")
            self.send_header("Content-Length", str(len(data)))
            self.end_headers()
            self.wfile.write(data)
            return
        self.send_error(HTTPStatus.NOT_FOUND)

    def do_POST(self) -> None:
        try:
            if self.path == "/api/settings":
                json_response(self, save_settings_file(read_json_body(self)))
                return
            parsed_post = urlparse(self.path)
            post_path = parsed_post.path
            post_params = parse_qs(parsed_post.query)
            if post_path == "/api/companies":
                body = read_json_body(self)
                name = clean_cell(body.get("name"))
                if not name:
                    raise ValueError("A cég neve kötelező.")
                companies = load_companies_file()
                company = {"id": uuid.uuid4().hex, "name": name}
                companies.append(company)
                save_companies_file(companies)
                settings = load_settings_file()
                settings["active_company_id"] = company["id"]
                save_settings_file(settings)
                json_response(self, {"company": company, "companies": companies})
                return
            if post_path == "/api/accounts":
                body = read_json_body(self)
                company_id = ensure_company_id((post_params.get("company_id") or [body.get("company_id") or ""])[0])
                body["company_id"] = company_id
                registry = load_bank_registry()
                payload = account_payload_from_body(body, registry)
                accounts = load_accounts_file()
                duplicate = find_duplicate_account(accounts, payload)
                if duplicate:
                    raise ValueError(f"Ez a bankszámla már szerepel a listában: {duplicate.get('account_number') or payload['account_number']}")
                found = False
                for index, account in enumerate(accounts):
                    if account.get("id") == payload["id"]:
                        accounts[index] = payload
                        found = True
                        break
                if not found:
                    accounts.append(payload)
                save_accounts_file(accounts)
                json_response(self, {"account": payload, "accounts": accounts_for_company(company_id)})
                return
            if post_path == "/api/accounts/delete":
                body = read_json_body(self)
                company_id = ensure_company_id((post_params.get("company_id") or [body.get("company_id") or ""])[0])
                account_id = clean_cell(body.get("id"))
                accounts = [row for row in load_accounts_file() if not (row.get("id") == account_id and clean_cell(row.get("company_id") or "default") == company_id)]
                save_accounts_file(accounts)
                json_response(self, {"accounts": accounts_for_company(company_id)})
                return
            if post_path == "/api/accounts/import":
                form = get_upload(self)
                filename, data = upload_bytes(form)
                company_id = ensure_company_id((post_params.get("company_id") or [""])[0])
                json_response(self, import_accounts_from_upload(data, filename, company_id))
                return
            if post_path == "/api/partners":
                body = read_json_body(self)
                company_id = ensure_company_id((post_params.get("company_id") or [body.get("company_id") or ""])[0])
                body["company_id"] = company_id
                registry = load_bank_registry()
                payload = partner_payload_from_body(body, registry)
                partners = load_partners_file()
                ident = partner_identity(payload)
                for partner in partners:
                    if (
                        partner.get("id") != payload["id"]
                        and ident
                        and partner_identity(partner) == ident
                        and clean_cell(partner.get("company_id") or "default") == company_id
                    ):
                        raise ValueError(f"Ez a partner már szerepel: {ident}")
                found = False
                for index, partner in enumerate(partners):
                    if partner.get("id") == payload["id"] and clean_cell(partner.get("company_id") or "default") == company_id:
                        partners[index] = payload
                        found = True
                        break
                if not found:
                    partners.append(payload)
                save_partners_file(partners)
                json_response(self, {"partner": payload, "partners": partners_for_company(company_id)})
                return
            if post_path == "/api/partners/delete":
                body = read_json_body(self)
                company_id = ensure_company_id((post_params.get("company_id") or [body.get("company_id") or ""])[0])
                partner_id = clean_cell(body.get("id"))
                partners = [row for row in load_partners_file() if not (row.get("id") == partner_id and clean_cell(row.get("company_id") or "default") == company_id)]
                save_partners_file(partners)
                json_response(self, {"partners": partners_for_company(company_id)})
                return
            if post_path == "/api/partners/import":
                form = get_upload(self)
                filename, data = upload_bytes(form)
                company_id = ensure_company_id((post_params.get("company_id") or [""])[0])
                json_response(self, import_partners_from_upload(data, filename, company_id))
                return
            if post_path == "/api/bank-registry":
                json_response(self, save_bank_registry(read_json_body(self)))
                return
            if post_path == "/api/bank-registry/refresh":
                registry = refresh_bank_registry()
                rows = registry.get("rows", [])
                json_response(self, {**registry, "row_count": len(rows), "sample": rows[:8]})
                return
            if post_path == "/api/inspect":
                form = get_upload(self)
                filename, data = upload_bytes(form)
                sheets, selected, headers, rows = workbook_rows(data, filename)
                non_empty_rows = [r for r in rows if not row_is_empty(list(r))]
                sample = [[clean_cell(c) for c in list(r)[: len(headers)]] for r in non_empty_rows[:6]]
                json_response(self, {
                    "sheets": sheets,
                    "sheet": selected,
                    "headers": headers,
                    "guesses": guess_mappings(headers),
                    "sample": sample,
                    "data_rows": len(non_empty_rows),
                })
                return
            if post_path == "/api/convert":
                form = get_upload(self)
                filename, data = upload_bytes(form)
                config = json.loads(field_value(form, "config", "{}") or "{}")
                text, count = convert_records(data, filename, config)
                encoding = config.get("encoding") or "cp1250"
                codec = "cp852" if encoding == "cp852" else ("cp1250" if encoding == "cp1250" else "utf-8")
                payload = text.encode(codec, errors="replace")
                self.send_response(HTTPStatus.OK)
                charset = "ibm852" if codec == "cp852" else ("windows-1250" if codec == "cp1250" else "utf-8")
                self.send_header("Content-Type", f"text/plain; charset={charset}")
                self.send_header("Content-Disposition", "attachment; filename=payord_import.txt")
                self.send_header("X-Filename", "payord_import.txt")
                self.send_header("X-Record-Count", str(count))
                self.send_header("Content-Length", str(len(payload)))
                self.end_headers()
                self.wfile.write(payload)
                return
            self.send_error(HTTPStatus.NOT_FOUND)
        except Exception as exc:
            message = str(exc)
            json_response(self, {"error": message, "errors": [line for line in message.splitlines() if line]}, status=400)

    def log_message(self, fmt: str, *args: Any) -> None:
        print(f"{self.address_string()} - {fmt % args}")


def main() -> None:
    registry = refresh_bank_registry_at_startup()
    print(registry.get("startup_message") or "Hitelesítő tábla állapota betöltve.")
    save_companies_file(load_companies_file())
    save_accounts_file(load_accounts_file())
    removed = cleanup_saved_accounts(registry)
    if removed:
        print(f"{len(removed)} korábban mentett, érvénytelen saját bankszámla eltávolítva.")
    server = ThreadingHTTPServer((HOST, PORT), Handler)
    print(f"PAYORD konverter fut: http://{HOST}:{PORT}")
    server.serve_forever()


if __name__ == "__main__":
    main()
